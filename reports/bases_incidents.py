"""
Génération du fichier "Bases des Incidents pour Automatisations"
à partir des données brutes du ticketing ou depuis la DB.
"""

import re
import pandas as pd
from io import BytesIO
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ── Couleurs ─────────────────────────────────────────────────────────────────
FILL_HDR     = PatternFill('solid', fgColor='4C5A77')   # en-têtes Mobile/Fixe/Core/Transport
FILL_HDR_IGW = PatternFill('solid', fgColor='1E4E79')   # en-têtes IGW
FILL_TITLE   = PatternFill('solid', fgColor='1F4E79')   # ligne titre
FILL_DR2_HDR = PatternFill('solid', fgColor='E7E6E6')   # en-têtes DR2 (gris clair)
FILL_DATA    = PatternFill('solid', fgColor='72AB43')   # TOUTES les lignes données (vert)

FONT_TITLE = Font(name='Century Gothic', size=14, bold=True, color='FFFFFF')
FONT_HDR   = Font(name='Calibri',        size=11, bold=True, color='FFFFFF')
FONT_HDR_IGW = Font(name='Calibri',      size=11, bold=True, color='E7E6E6')
FONT_DR2H  = Font(name='Calibri',        size=11, bold=False, color='000000')
FONT_DATA  = Font(name='Calibri',        size=11)

ALIGN_CTR  = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_LFT  = Alignment(horizontal='left',   vertical='center', wrap_text=False)
ALIGN_LEFT = ALIGN_LFT

BORDER_THIN = Border(
    left=Side(style='thin', color='D0D4DC'),
    right=Side(style='thin', color='D0D4DC'),
    bottom=Side(style='thin', color='D0D4DC'),
)

_MOIS_FR = {
    1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
    5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
    9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE',
}


def _fmt_dur(secs):
    if secs is None:
        return ''
    s = int(abs(secs))
    return f'{s // 3600}:{(s % 3600) // 60:02d}:{s % 60:02d}'


def _fmt_dt(val):
    """Format datetime → 'dd-mm-yyyy HH:MM:SS' string."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d-%m-%Y %H:%M:%S')
    return str(val)


def _clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s.lower() in ('nan', 'none', 'n/a', 'na', ''):
        return ''
    return s


def _parse_dur_str(s):
    """Parse 'HH:MM:SS' or 'H:MM:SS' → seconds. Returns None if invalid."""
    if not s or str(s).strip().lower() in ('n/a', 'na', ''):
        return None
    try:
        parts = str(s).strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        pass
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# PARSERS FICHIERS BRUTS TICKETING
# ═══════════════════════════════════════════════════════════════════════════════

def _read_raw_df(path):
    """Lit le fichier brut (1 feuille, en-têtes ligne 1)."""
    try:
        df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)
    except Exception:
        df = pd.read_excel(path, sheet_name='Feuille 1', header=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how='all')
    return df


def _parse_dt(val):
    if not val or str(val).strip().lower() in ('nan', 'none', 'n/a', 'na', ''):
        return None
    try:
        ts = pd.to_datetime(str(val), dayfirst=True, errors='coerce')
        if pd.isna(ts):
            return None
        return ts.to_pydatetime().replace(tzinfo=None)
    except Exception:
        return None


def parse_raw_mobile(path, mois_filter=None):
    """
    Parse un fichier brut de ticketing réseau mobile.
    Retourne une liste de dicts avec TOUS les champs (y.c. Ingénieur NOC, Alarm text).
    mois_filter: date(YYYY, MM, 1) pour filtrer par mois, None = tout garder.
    """
    df = _read_raw_df(path)

    # Normalisation colonnes
    COL_RENAMES = {
        'Ingénieur NOC': 'ingenieur_noc',
        'Numero du ticket': 'numero_ticket',
        "Nature de l'incident": 'nature',
        'Alarm Time': 'alarm_time_str',
        'Site Parent': 'site_parent',
        'Site Name': 'site_name',
        'Site ID': 'site_id',
        'Région': 'region',
        'Région': 'region',
        'Impact - Equipement': 'impact_equipement',
        'Impact - Service': 'impact_service',
        'Plateforme': 'plateforme',
        'Technologies': 'technologies',
        'Alarm text': 'alarm_text',
        'Cause': 'cause',
        'Escalade': 'escalade',
        'Technicien Informé': 'technicien_informe',
        'Technicien Informé': 'technicien_informe',
        'Durée escalade': 'duree_escalade',
        'Durée escalade': 'duree_escalade',
        'Action': 'action',
        'Technicien de maintenance': 'technicien_maint',
        'Root Cause': 'root_cause',
        'Observation': 'observation',
        'Point bloquant': 'point_bloquant',
        'Cancel Time': 'cancel_time_str',
        'Duration': 'duration_str',
        'Status': 'status',
    }
    df.rename(columns={k: v for k, v in COL_RENAMES.items() if k in df.columns}, inplace=True)

    rows = []
    for _, r in df.iterrows():
        at = _parse_dt(r.get('alarm_time_str', ''))
        ct = _parse_dt(r.get('cancel_time_str', ''))

        # Filtrage par mois
        if mois_filter and at:
            if date(at.year, at.month, 1) != mois_filter:
                continue

        # Duration en secondes
        dur_sec = None
        dur_str = _clean(r.get('duration_str', ''))
        if dur_str:
            dur_sec = _parse_dur_str(dur_str)
        if dur_sec is None and at and ct:
            delta = (ct - at).total_seconds()
            dur_sec = delta if delta >= 0 else None
            if dur_sec is not None:
                dur_str = _fmt_dur(dur_sec)

        rows.append({
            'ingenieur_noc':     _clean(r.get('ingenieur_noc', '')),
            'numero_ticket':     _clean(r.get('numero_ticket', '')),
            'nature':            _clean(r.get('nature', '')),
            'alarm_time':        at,
            'alarm_time_str':    _clean(r.get('alarm_time_str', '')),
            'site_parent':       _clean(r.get('site_parent', '')),
            'site_name':         _clean(r.get('site_name', '')),
            'site_id':           _clean(r.get('site_id', '')),
            'region':            _clean(r.get('region', '')),
            'impact_equipement': _clean(r.get('impact_equipement', '')),
            'impact_service':    _clean(r.get('impact_service', '')),
            'plateforme':        _clean(r.get('plateforme', '')),
            'technologies':      _clean(r.get('technologies', '')),
            'alarm_text':        _clean(r.get('alarm_text', '')),
            'cause':             _clean(r.get('cause', '')),
            'escalade':          _clean(r.get('escalade', '')),
            'technicien_informe': _clean(r.get('technicien_informe', '')),
            'duree_escalade':    _clean(r.get('duree_escalade', '')),
            'action':            _clean(r.get('action', '')),
            'technicien_maint':  _clean(r.get('technicien_maint', '')),
            'root_cause':        _clean(r.get('root_cause', '')),
            'observation':       _clean(r.get('observation', '')),
            'point_bloquant':    _clean(r.get('point_bloquant', '')),
            'cancel_time':       ct,
            'cancel_time_str':   _clean(r.get('cancel_time_str', '')),
            'duration_str':      dur_str,
            'duration_sec':      dur_sec,
            'status':            _clean(r.get('status', '')),
        })
    return rows


def _base_lookup(site_name, site_id):
    """Cherche le champ 'base' dans le modèle Site."""
    try:
        from .models import Site
        if site_name:
            s = Site.objects.filter(site_name__iexact=site_name.strip()).first()
            if s and s.base:
                return s.base
        if site_id:
            s = Site.objects.filter(site_id__iexact=site_id.strip()).first()
            if s and s.base:
                return s.base
    except Exception:
        pass
    return ''


# ═══════════════════════════════════════════════════════════════════════════════
# CHARGEMENT DEPUIS DB
# ═══════════════════════════════════════════════════════════════════════════════

def _db_mobile(mois):
    from .models import Incident
    qs = Incident.objects.filter(domain='mobile', mois_rapport=mois).order_by('alarm_time')
    rows = []
    for inc in qs:
        rows.append({
            'ingenieur_noc': '',
            'numero_ticket': inc.numero_ticket,
            'nature':        inc.nature,
            'alarm_time':    inc.alarm_time.replace(tzinfo=None) if inc.alarm_time else None,
            'site_parent':   inc.site_parent,
            'site_name':     inc.site_name,
            'site_id':       inc.site_id,
            'region':        inc.region,
            'base':          inc.base,
            'impact_equipement': inc.impact_equipement,
            'impact_service':    inc.impact_service,
            'plateforme':    inc.plateforme,
            'technologies':  inc.technologies,
            'alarm_text':    '',
            'cause':         inc.cause,
            'escalade':      inc.escalade,
            'technicien_informe': inc.technicien_informe,
            'duree_escalade': '',
            'action':        inc.action,
            'technicien_maint': inc.technicien_maint,
            'root_cause':    inc.root_cause,
            'observation':   inc.observation,
            'point_bloquant': inc.point_bloquant,
            'cancel_time':   inc.cancel_time.replace(tzinfo=None) if inc.cancel_time else None,
            'duration_sec':  inc.duration_sec,
            'duration_str':  _fmt_dur(inc.duration_sec) if inc.duration_sec else '',
            'status':        inc.status,
        })
    return rows


def _db_domain(domain, mois):
    from .models import Incident
    qs = Incident.objects.filter(domain=domain, mois_rapport=mois).order_by('alarm_time')
    rows = []
    for inc in qs:
        rows.append({
            'numero_ticket':     inc.numero_ticket,
            'nature':            inc.nature,
            'alarm_time':        inc.alarm_time.replace(tzinfo=None) if inc.alarm_time else None,
            'cancel_time':       inc.cancel_time.replace(tzinfo=None) if inc.cancel_time else None,
            'site_parent':       inc.site_parent,
            'site_name':         inc.site_name,
            'site_id':           inc.site_id,
            'region':            inc.region,
            'base':              inc.base,
            'plateforme':        inc.plateforme,
            'technologies':      inc.technologies,
            'impact_equipement': inc.impact_equipement,
            'impact_service':    inc.impact_service,
            'nbre_clients':      inc.nbre_clients,
            'escalade':          inc.escalade,
            'cause':             inc.cause,
            'root_cause':        inc.root_cause,
            'action':            inc.action,
            'technicien_informe': inc.technicien_informe,
            'technicien_maint':  inc.technicien_maint,
            'point_bloquant':    inc.point_bloquant,
            'observation':       inc.observation,
            'duration_sec':      inc.duration_sec,
            'duration_str':      _fmt_dur(inc.duration_sec) if inc.duration_sec else '',
            'status':            inc.status,
        })
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# DÉRIVATION DR2
# ═══════════════════════════════════════════════════════════════════════════════

def derive_dr2(mobile_rows):
    """Extrait les DR2 depuis les lignes Mobile (durée > 3h)."""
    dr2 = []
    n = 1
    for r in mobile_rows:
        dur = r.get('duration_sec')
        if dur is None:
            dur_str = r.get('duration_str', '')
            dur = _parse_dur_str(dur_str)
        if dur is None or dur <= 3 * 3600:
            continue

        at = r.get('alarm_time')
        ct = r.get('cancel_time')
        date_dr2 = None
        if ct:
            date_dr2 = ct.date() if isinstance(ct, datetime) else ct
        elif at:
            date_dr2 = (at + timedelta(hours=3)).date() if isinstance(at, datetime) else at

        dr2.append({
            'n':             n,
            'date_dr2':      date_dr2,
            'numero_ticket': r.get('numero_ticket', ''),
            'site_parent':   r.get('site_parent', ''),
            'site_name':     r.get('site_name', ''),
            'site_id':       r.get('site_id', ''),
            'region':        r.get('region', ''),
            'base':          r.get('base', ''),
            'alarm_time':    at,
            'duree':         _fmt_dur(dur),
            'categorie':     r.get('escalade', ''),
            'cause':         r.get('cause', ''),
            'point_bloquant': r.get('point_bloquant', ''),
            'cancel_time':   ct,
            'observation':   r.get('observation', ''),
        })
        n += 1
    return dr2


def derive_dr2_from_db(mois):
    """Dérive les DR2 depuis les incidents mobile en DB (durée > 3h)."""
    from .models import Incident
    qs = (
        Incident.objects.filter(domain='mobile', mois_rapport=mois)
        .exclude(duration_sec__isnull=True)
        .filter(duration_sec__gt=3 * 3600)
        .order_by('alarm_time')
    )
    rows = []
    n = 1
    for inc in qs:
        at = inc.alarm_time.replace(tzinfo=None) if inc.alarm_time else None
        ct = inc.cancel_time.replace(tzinfo=None) if inc.cancel_time else None
        date_dr2 = ct.date() if ct else (at + timedelta(hours=3)).date() if at else None
        rows.append({
            'n':             n,
            'date_dr2':      date_dr2,
            'numero_ticket': inc.numero_ticket,
            'site_parent':   inc.site_parent,
            'site_name':     inc.site_name,
            'site_id':       inc.site_id,
            'region':        inc.region,
            'base':          inc.base,
            'alarm_time':    at,
            'duree':         _fmt_dur(inc.duration_sec),
            'categorie':     inc.escalade,
            'cause':         inc.cause,
            'point_bloquant': inc.point_bloquant,
            'cancel_time':   ct,
            'observation':   inc.observation,
        })
        n += 1
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATEUR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_cell(ws, row, col, value, fill=FILL_HDR, font=FONT_HDR, align=ALIGN_CTR):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill
    c.font   = font
    c.alignment = align
    c.border = BORDER_THIN
    return c


def _data_cell(ws, row, col, value, alt=False, align=ALIGN_LFT):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_DATA
    c.alignment = align
    c.border = BORDER_THIN
    c.fill = FILL_DATA  # vert uniforme sur toutes les lignes données
    return c


def _title_row(ws, row, text, ncols, row_height=22):
    c = ws.cell(row=row, column=1, value=text)
    c.fill = FILL_TITLE
    c.font = FONT_TITLE
    c.alignment = ALIGN_LFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = row_height


# ── Sheet Mobile ──────────────────────────────────────────────────────────────

def _write_mobile(wb, mobile_rows, mois_label):
    ws = wb.create_sheet('Reseau mobile ')
    ws.sheet_view.showGridLines = True

    HEADERS = [
        'Ingénieur NOC', 'Numero du ticket', "Nature de l'incident", 'Alarm Time',
        'Site Parent', 'Site Name', 'Site ID', 'Région', 'Base',
        'Impact - Equipement', 'Impact - Service', 'Plateforme', 'Technologies',
        'Alarm text', 'Cause', 'Escalade', 'Technicien Informé', 'Durée escalade',
        'Action', 'Technicien de maintenance', 'Root Cause', 'Observation',
        'Point bloquant', 'Cancel Time', 'Duration', 'Status',
    ]
    ncols = len(HEADERS)

    # Row 1: vide
    ws.row_dimensions[1].height = 17.25

    # Row 2: titre
    _title_row(ws, 2, f'Réseau Mobile : {mois_label}', ncols, row_height=18.75)

    # Row 3: vide
    ws.row_dimensions[3].height = 5

    # Row 4: en-têtes
    ws.row_dimensions[4].height = 23.25
    for j, h in enumerate(HEADERS, 1):
        _header_cell(ws, 4, j, h)

    # Rows 5+: données
    for i, r in enumerate(mobile_rows):
        row = i + 5
        alt  = (i % 2 == 1)
        ws.row_dimensions[row].height = 15

        vals = [
            r.get('ingenieur_noc', ''),
            r.get('numero_ticket', ''),
            r.get('nature', ''),
            _fmt_dt(r.get('alarm_time')),
            r.get('site_parent', ''),
            r.get('site_name', ''),
            r.get('site_id', ''),
            r.get('region', ''),
            r.get('base', ''),
            r.get('impact_equipement', ''),
            r.get('impact_service', ''),
            r.get('plateforme', ''),
            r.get('technologies', ''),
            r.get('alarm_text', ''),
            r.get('cause', ''),
            r.get('escalade', ''),
            r.get('technicien_informe', ''),
            r.get('duree_escalade', ''),
            r.get('action', ''),
            r.get('technicien_maint', ''),
            r.get('root_cause', ''),
            r.get('observation', ''),
            r.get('point_bloquant', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('duration_str', ''),
            r.get('status', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v, alt=alt)

    # Largeurs colonnes (calées sur fichier référence)
    _set_col_widths(ws, [
        27, 26, 54, 26, 18, 26, 12, 17, 13,  # A-I (I=Base)
        12, 11, 17, 16, 24, 22, 20, 34, 25,  # J-R
        23, 16, 30, 18, 12, 22, 23, 12,       # S-Z
    ])

    ws.freeze_panes = 'A5'


# ── Sheet DR2 ─────────────────────────────────────────────────────────────────

def _write_dr2(wb, dr2_rows, mois_label):
    ws = wb.create_sheet('DR2')

    HEADERS = [
        None, 'N°', 'DATE DR2', 'Numero ticket', 'SITE PARENT',
        'Site Name', 'Site ID', 'REGION', 'BASE', 'Alarm Time',
        'DUREE', 'CATEGORIE', 'CAUSE', 'POINT BLOQUANTS', 'Cancel Time', 'OBSERVATION',
    ]
    ncols = len(HEADERS)

    # Row 1: vide
    ws.row_dimensions[1].height = 14

    # Row 2: en-têtes DR2 (fond gris clair, texte non-gras, comme le référence)
    ws.row_dimensions[2].height = 22
    for j, h in enumerate(HEADERS, 1):
        if h is not None:
            _header_cell(ws, 2, j, h, fill=FILL_DR2_HDR, font=FONT_DR2H)

    # Rows 3+: données
    for i, r in enumerate(dr2_rows):
        row = i + 3
        alt = (i % 2 == 1)
        ws.row_dimensions[row].height = 15

        dt_str = ''
        dt_dr2 = r.get('date_dr2')
        if dt_dr2:
            if isinstance(dt_dr2, (date, datetime)):
                dt_str = dt_dr2.strftime('%d/%m/%Y')
            else:
                dt_str = str(dt_dr2)

        vals = [
            None,
            r.get('n', i + 1),
            dt_str,
            r.get('numero_ticket', ''),
            r.get('site_parent', ''),
            r.get('site_name', ''),
            r.get('site_id', ''),
            r.get('region', ''),
            r.get('base', ''),
            _fmt_dt(r.get('alarm_time')),
            r.get('duree', ''),
            r.get('categorie', ''),
            r.get('cause', ''),
            r.get('point_bloquant', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('observation', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v, alt=alt,
                       align=ALIGN_CTR if j in (2, 3, 7, 11) else ALIGN_LFT)

    # Largeurs colonnes DR2 (calées sur fichier référence)
    _set_col_widths(ws, [11, 6, 14, 22, 19, 28, 20, 20, 13, 26, 25, 46, 42, 24, 18, 80])
    ws.freeze_panes = 'B3'


# ── Sheet Réseau Fixe ─────────────────────────────────────────────────────────

def _write_fixe(wb, fixe_rows, mois_label):
    ws = wb.create_sheet('Reseau Fixe')

    HEADERS = [
        "Nature de l'incident", 'Alarm Time', 'Site Name', 'Plateforme',
        'Impact - Equipement', 'Impact - Service', 'Nbre de client Impactés',
        'Escalade', 'Root Cause', 'Action', 'Technicien de maintenance',
        'Cancel Time', 'Duration', 'Status', 'Commentaire',
    ]
    ncols = len(HEADERS)

    _title_row(ws, 2, f'Réseau Fixe : {mois_label}', ncols)
    ws.row_dimensions[3].height = 17
    ws.row_dimensions[4].height = 27
    ws.row_dimensions[5].height = 29
    for j, h in enumerate(HEADERS, 1):
        _header_cell(ws, 5, j, h)

    for i, r in enumerate(fixe_rows):
        row = i + 6
        ws.row_dimensions[row].height = 19
        vals = [
            r.get('nature', ''),
            _fmt_dt(r.get('alarm_time')),
            r.get('site_name', ''),
            r.get('plateforme', ''),
            r.get('impact_equipement', ''),
            r.get('impact_service', ''),
            r.get('nbre_clients', ''),
            r.get('escalade', ''),
            r.get('root_cause', ''),
            r.get('action', ''),
            r.get('technicien_maint', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('duration_str', ''),
            r.get('status', ''),
            r.get('observation', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v)

    # Largeurs colonnes Fixe (calées sur fichier référence)
    _set_col_widths(ws, [36, 28, 24, 16, 20, 118, 26, 20, 55, 51, 25, 25, 19, 19, 33])
    ws.freeze_panes = 'A6'


# ── Sheet Transport ───────────────────────────────────────────────────────────

def _write_transport(wb, transport_rows, mois_label):
    ws = wb.create_sheet('Transport')

    HEADERS_ROW = [
        None, 'Numero du ticket', "Nature de l'incident", 'Alarm Time',
        'Site Parent', 'Site Name', 'Site ID', 'Région',
        'Impact - Equipement', 'Impact - Service', 'Plateforme', 'Technologies',
        'Cause', 'Escalade', 'Technicien Informé', 'Durée escalade',
        'Action', 'Technicien de maintenance', 'Root Cause', 'Observation',
        'Point bloquant', 'Cancel Time', 'Duration', 'Status',
    ]
    ncols = len(HEADERS_ROW)

    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 22
    for j, h in enumerate(HEADERS_ROW, 1):
        if h is not None:
            _header_cell(ws, 2, j, h)

    for i, r in enumerate(transport_rows):
        row = i + 3
        ws.row_dimensions[row].height = 15
        vals = [
            None,
            r.get('numero_ticket', ''),
            r.get('nature', ''),
            _fmt_dt(r.get('alarm_time')),
            r.get('site_parent', ''),
            r.get('site_name', ''),
            r.get('site_id', ''),
            r.get('region', ''),
            r.get('impact_equipement', ''),
            r.get('impact_service', ''),
            r.get('plateforme', ''),
            r.get('technologies', ''),
            r.get('cause', ''),
            r.get('escalade', ''),
            r.get('technicien_informe', ''),
            r.get('duree_escalade', ''),
            r.get('action', ''),
            r.get('technicien_maint', ''),
            r.get('root_cause', ''),
            r.get('observation', ''),
            r.get('point_bloquant', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('duration_str', ''),
            r.get('status', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v)

    # Largeurs colonnes Transport (calées sur fichier référence)
    _set_col_widths(ws, [8, 21, 107, 18, 38, 49, 12, 17, 74, 145, 17, 17, 12, 10, 64, 17, 71, 33, 13, 256, 20, 18, 10, 7])
    ws.freeze_panes = 'B3'


# ── Sheet IGW ─────────────────────────────────────────────────────────────────

def _write_igw(wb, igw_rows, mois_label):
    ws = wb.create_sheet('IGW')

    HEADERS = [
        'ALARM TIME', "NATURE DE L'INCIDENT", 'LIEN', 'LIEN INTERNET',
        'IMPACTS', 'ESCALADE', "CAUSES DE L'INCIDENT", 'PRIORITES',
        'CANCEL TIME', 'DURATION', 'ACTIONS DONE', 'STATUS', 'OBSERVATIONS',
    ]
    ncols = len(HEADERS)

    _title_row(ws, 2, f' Transmission Internationale (IGW) : {mois_label}', ncols)
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 17
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 22
    for j, h in enumerate(HEADERS, 1):
        # IGW utilise fond bleu foncé #1E4E79 avec texte gris clair
        _header_cell(ws, 5, j, h, fill=FILL_HDR_IGW, font=FONT_HDR_IGW,
                     align=Alignment(horizontal='center', vertical='center', wrap_text=True))

    for i, r in enumerate(igw_rows):
        row = i + 6
        ws.row_dimensions[row].height = 30
        vals = [
            _fmt_dt(r.get('alarm_time')),
            r.get('nature', ''),
            r.get('site_name', ''),
            r.get('plateforme', ''),
            r.get('impact_service', ''),
            r.get('escalade', ''),
            r.get('cause', ''),
            r.get('priority', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('duration_str', ''),
            r.get('action', ''),
            r.get('status', ''),
            r.get('observation', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v,
                       align=Alignment(horizontal='left', vertical='center', wrap_text=True))

    # Largeurs colonnes IGW (calées sur fichier référence)
    _set_col_widths(ws, [22, 58, 23, 19, 23, 19, 30, 16, 19, 14, 27, 11, 51])
    ws.freeze_panes = 'A6'


# ── Sheet Core ────────────────────────────────────────────────────────────────

def _write_core(wb, core_rows, mois_label):
    ws = wb.create_sheet('Core')

    HEADERS_ROW = [
        None, "Nature de l'incident", 'Alarm Time', 'ESPC',
        'Impact - Service', 'Escalade', 'Action', 'Technicien Informé',
        'Root Cause', 'Cancel Time', 'Duration', 'Status', 'Commentaire',
    ]
    ncols = len(HEADERS_ROW)

    # Row 2: titre (à partir de col 2 pour le core)
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 22
    c = ws.cell(row=2, column=2, value=f'Core Network : {mois_label}')
    c.fill = FILL_TITLE
    c.font = FONT_TITLE
    c.alignment = ALIGN_LFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncols)

    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[5].height = 22
    for j, h in enumerate(HEADERS_ROW, 1):
        if h is not None:
            _header_cell(ws, 5, j, h)

    for i, r in enumerate(core_rows):
        row = i + 6
        ws.row_dimensions[row].height = 21
        vals = [
            None,
            r.get('nature', ''),
            _fmt_dt(r.get('alarm_time')),
            r.get('site_name', ''),
            r.get('impact_service', ''),
            r.get('escalade', ''),
            r.get('action', ''),
            r.get('technicien_informe', ''),
            r.get('root_cause', ''),
            _fmt_dt(r.get('cancel_time')),
            r.get('duration_str', ''),
            r.get('status', ''),
            r.get('observation', ''),
        ]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, row, j, v)

    # Largeurs colonnes Core (calées sur fichier référence)
    _set_col_widths(ws, [8, 74, 18, 19, 30, 31, 28, 18, 21, 18, 12, 12, 17])
    ws.freeze_panes = 'B6'


# ═══════════════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════

def generate_bases_incidents(
    mois,                    # date(YYYY, MM, 1)
    mobile_file=None,        # path ou None → depuis DB
    fixe_file=None,
    transport_file=None,
    igw_file=None,
    core_file=None,
):
    """
    Génère le fichier "BASES DES INCIDENTS" multi-onglets.
    Pour chaque domaine:
    - Si un fichier est fourni: on le parse directement (préserve Ingénieur NOC, Alarm text, etc.)
    - Sinon: on charge depuis la DB pour le mois donné.
    Retourne un BytesIO.
    """
    mois_label = f'{_MOIS_FR.get(mois.month, str(mois.month))} {mois.year}'

    # ── Mobile ───────────────────────────────────────────────────────────────
    if mobile_file:
        mobile_rows = parse_raw_mobile(mobile_file, mois_filter=mois)
        # Lookup Base depuis Site model
        for r in mobile_rows:
            if not r.get('base'):
                r['base'] = _base_lookup(r.get('site_name', ''), r.get('site_id', ''))
        # Dériver DR2 depuis les lignes mobile
        dr2_rows = derive_dr2(mobile_rows)
    else:
        mobile_rows = _db_mobile(mois)
        for r in mobile_rows:
            if not r.get('base'):
                r['base'] = _base_lookup(r.get('site_name', ''), r.get('site_id', ''))
        dr2_rows = derive_dr2_from_db(mois)

    # ── Fixe ─────────────────────────────────────────────────────────────────
    if fixe_file:
        # Pour les autres plateformes, même parseur générique (headers ligne 1)
        fixe_rows = _parse_generic_raw(fixe_file, mois, 'fixe')
    else:
        fixe_rows = _db_domain('fixe', mois)

    # ── Transport ─────────────────────────────────────────────────────────────
    if transport_file:
        transport_rows = _parse_generic_raw(transport_file, mois, 'transport')
    else:
        transport_rows = _db_domain('transport', mois)

    # ── IGW ───────────────────────────────────────────────────────────────────
    if igw_file:
        igw_rows = _parse_generic_raw(igw_file, mois, 'igw')
    else:
        igw_rows = _db_domain('igw', mois)

    # ── Core ──────────────────────────────────────────────────────────────────
    if core_file:
        core_rows = _parse_generic_raw(core_file, mois, 'core')
    else:
        core_rows = _db_domain('core', mois)

    # ── Assemblage Excel ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    # Supprimer la feuille vide par défaut
    wb.remove(wb.active)

    _write_mobile(wb, mobile_rows, mois_label)
    _write_dr2(wb, dr2_rows, mois_label)
    _write_fixe(wb, fixe_rows, mois_label)
    _write_transport(wb, transport_rows, mois_label)
    _write_igw(wb, igw_rows, mois_label)
    _write_core(wb, core_rows, mois_label)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(mobile_rows), len(dr2_rows)


def generate_platform_bases_incidents(platform, mois, source_file=None):
    """
    Génère le fichier Bases des Incidents pour une seule plateforme.
    - source_file: chemin vers le fichier brut uploadé (None = depuis DB)
    - mois: date(YYYY,MM,1) — utilisé pour filtrer la DB ou nommer le fichier
    Retourne (BytesIO, nb_rows).
    """
    mois_label = f'{_MOIS_FR.get(mois.month, str(mois.month))} {mois.year}'
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if platform == 'mobile-dr2':
        if source_file:
            # Pas de filtre mois pour le fichier (toutes les données du fichier)
            mobile_rows = parse_raw_mobile(source_file, mois_filter=None)
            for r in mobile_rows:
                if not r.get('base'):
                    r['base'] = _base_lookup(r.get('site_name', ''), r.get('site_id', ''))
            dr2_rows = derive_dr2(mobile_rows)
        else:
            mobile_rows = _db_mobile(mois)
            for r in mobile_rows:
                if not r.get('base'):
                    r['base'] = _base_lookup(r.get('site_name', ''), r.get('site_id', ''))
            dr2_rows = derive_dr2_from_db(mois)
        _write_mobile(wb, mobile_rows, mois_label)
        _write_dr2(wb, dr2_rows, mois_label)
        nb = len(mobile_rows)

    elif platform == 'fixe':
        rows = (_parse_generic_raw(source_file, None, 'fixe') if source_file
                else _db_domain('fixe', mois))
        _write_fixe(wb, rows, mois_label)
        nb = len(rows)

    elif platform == 'transmission':
        rows = (_parse_generic_raw(source_file, None, 'transport') if source_file
                else _db_domain('transport', mois))
        _write_transport(wb, rows, mois_label)
        nb = len(rows)

    elif platform == 'igw':
        rows = (_parse_generic_raw(source_file, None, 'igw') if source_file
                else _db_domain('igw', mois))
        _write_igw(wb, rows, mois_label)
        nb = len(rows)

    elif platform == 'core':
        rows = (_parse_generic_raw(source_file, None, 'core') if source_file
                else _db_domain('core', mois))
        _write_core(wb, rows, mois_label)
        nb = len(rows)

    else:
        return None, 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, nb


def _parse_generic_raw(path, mois_filter, platform):
    """
    Parseur générique pour fichier brut d'une plateforme (headers ligne 1).
    Tente de mapper automatiquement les colonnes connues.
    """
    try:
        df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)
    except Exception:
        return []
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how='all')

    ALIASES = {
        'Alarm Time': 'alarm_time_str',
        'ALARM TIME': 'alarm_time_str',
        'Cancel Time': 'cancel_time_str',
        'CANCEL TIME': 'cancel_time_str',
        "Nature de l'incident": 'nature',
        "NATURE DE L'INCIDENT": 'nature',
        "NATURE DE L'INCIDENT ": 'nature',
        'Numero du ticket': 'numero_ticket',
        'Site Name': 'site_name',
        'LIEN': 'site_name',
        'Site Parent': 'site_parent',
        'SITE PARENT': 'site_parent',
        'Site ID': 'site_id',
        'Région': 'region',
        'REGION': 'region',
        'Plateforme': 'plateforme',
        'LIEN INTERNET': 'plateforme',
        'Technologies': 'technologies',
        'Cause': 'cause',
        "CAUSES DE L'INCIDENT": 'cause',
        "CAUSES DE L'INCIDENT ": 'cause',
        'Escalade': 'escalade',
        'ESCALADE': 'escalade',
        'Action': 'action',
        'ACTIONS DONE': 'action',
        'Root Cause': 'root_cause',
        'Technicien Informé': 'technicien_informe',
        'Technicien Informé': 'technicien_informe',
        'Technicien de maintenance': 'technicien_maint',
        'Impact - Service': 'impact_service',
        'IMPACTS': 'impact_service',
        'IMPACTS ': 'impact_service',
        'Impact - Equipement': 'impact_equipement',
        'Nbre de client Impactés': 'nbre_clients',
        'Observation': 'observation',
        'OBSERVATIONS': 'observation',
        'Point bloquant': 'point_bloquant',
        'Status': 'status',
        'STATUS': 'status',
        'STATUS ': 'status',
        'Duration': 'duration_str',
        'DURATION': 'duration_str',
        'Commentaire': 'observation',
        'Commentaire ': 'observation',
        'ESPC': 'site_name',
        'ESPC ': 'site_name',
        'PRIORITES': 'priority',
    }
    df.rename(columns={k: v for k, v in ALIASES.items() if k in df.columns}, inplace=True)

    rows = []
    for _, r in df.iterrows():
        at = _parse_dt(r.get('alarm_time_str', ''))
        ct = _parse_dt(r.get('cancel_time_str', ''))

        if mois_filter and at:
            if date(at.year, at.month, 1) != mois_filter:
                continue

        dur_sec = None
        dur_str = _clean(r.get('duration_str', ''))
        if dur_str:
            dur_sec = _parse_dur_str(dur_str)
        if dur_sec is None and at and ct:
            delta = (ct - at).total_seconds()
            dur_sec = delta if delta >= 0 else None
            if dur_sec is not None:
                dur_str = _fmt_dur(dur_sec)

        row_dict = {
            'numero_ticket':     _clean(r.get('numero_ticket', '')),
            'nature':            _clean(r.get('nature', '')),
            'alarm_time':        at,
            'cancel_time':       ct,
            'site_parent':       _clean(r.get('site_parent', '')),
            'site_name':         _clean(r.get('site_name', '')),
            'site_id':           _clean(r.get('site_id', '')),
            'region':            _clean(r.get('region', '')),
            'base':              _clean(r.get('base', '')),
            'plateforme':        _clean(r.get('plateforme', '')),
            'technologies':      _clean(r.get('technologies', '')),
            'impact_equipement': _clean(r.get('impact_equipement', '')),
            'impact_service':    _clean(r.get('impact_service', '')),
            'nbre_clients':      _clean(r.get('nbre_clients', '')),
            'escalade':          _clean(r.get('escalade', '')),
            'cause':             _clean(r.get('cause', '')),
            'root_cause':        _clean(r.get('root_cause', '')),
            'action':            _clean(r.get('action', '')),
            'technicien_informe': _clean(r.get('technicien_informe', '')),
            'duree_escalade':    '',
            'technicien_maint':  _clean(r.get('technicien_maint', '')),
            'point_bloquant':    _clean(r.get('point_bloquant', '')),
            'observation':       _clean(r.get('observation', '')),
            'priority':          _clean(r.get('priority', '')),
            'duration_str':      dur_str,
            'duration_sec':      dur_sec,
            'status':            _clean(r.get('status', '')),
        }
        rows.append(row_dict)
    return rows
