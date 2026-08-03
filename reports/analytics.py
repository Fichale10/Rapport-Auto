"""Module Analytics — analyses automatiques d'incidents (page /analytics/).

Pipeline :
  1. Chargement : fichier Excel/CSV importé OU export API ticketing (NetXcare).
  2. Normalisation des colonnes vers le schéma canonique :
     region, base, site, equipement, cause, classification, escalade,
     duration_sec, date, status.
  3. Enrichissement depuis la table Site (région / base / classification
     manquantes, jointure par nom de site).
  4. compute() : KPIs + les 10 axes d'analyse, restitués en dict JSON-safe
     (consommé par le template + Chart.js et par les exports Excel / PDF).
"""
from __future__ import annotations

import datetime as _dt
import io
import json
import math
import re
import unicodedata

import pandas as pd

EMPTY_LABEL = '(Non renseigné)'

# Libellés français pour l'export Excel / PDF
FR_LABELS = {
    'region':         'Région',
    'base':           'Base',
    'site':           'Site',
    'equipement':     'Équipement en défaut',
    'cause':          'Cause',
    'classification': 'Classification du site',
    'escalade':       'Escalade',
    'duration_sec':   'Durée (s)',
    'date':           'Date',
    'status':         'Statut',
}


# ═════════════════════════ 1. Normalisation ═════════════════════════════════

def _norm(label) -> str:
    """Minuscule, sans accents ni ponctuation — pour comparer les en-têtes."""
    s = str(label or '')
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', ' ', s.lower()).strip()


# Alias reconnus pour chaque colonne canonique (formats fichier libre + API)
_ALIASES: dict[str, list[str]] = {
    'region':         ['region', 'regions', 'dr', 'direction regionale'],
    'base':           ['base', 'base technique'],
    'site':           ['site', 'site name', 'nom du site', 'nom site', 'site impacte'],
    'equipement':     ['equipement en defaut', 'equipement defaillant', 'impact equipement',
                      'impact equipment', 'equipement en panne', 'equipement', 'equipment'],
    'cause':          ['cause', 'type de cause', 'cause de l incident', 'root cause',
                      'cause racine'],
    'classification': ['classification du site', 'classification technique', 'classification',
                      'classif tech', 'classif', 'categorie site'],
    'escalade':       ['escalade', 'escalation', 'entite responsable'],
    'duration':       ['duree de l incident', 'duree incident', 'duration', 'duree',
                      'duree totale', 'temps d indisponibilite', 'outage'],
    'date':           ['alarm time', 'date de l incident', 'date debut', 'date', 'debut incident', 'debut'],
    'status':         ['status', 'statut', 'etat'],
}


def _map_columns(df: pd.DataFrame) -> dict[str, str]:
    """Associe chaque colonne canonique à la colonne source correspondante."""
    norm_cols: dict[str, str] = {}
    for c in df.columns:
        n = _norm(c)
        if n and n not in norm_cols:
            norm_cols[n] = c
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for canon, aliases in _ALIASES.items():
        found = None
        for a in aliases:                              # correspondance exacte
            col = norm_cols.get(a)
            if col and col not in used:
                found = col
                break
        if not found:                                  # correspondance partielle
            for a in aliases:
                for n, orig in norm_cols.items():
                    if orig in used:
                        continue
                    if n.startswith(a + ' ') or (len(a) > 6 and a in n):
                        found = orig
                        break
                if found:
                    break
        if found:
            mapping[canon] = found
            used.add(found)
    return mapping


def _dur_to_sec(v) -> float:
    """Durée → secondes. Gère HH:MM:SS (heures > 24 acceptées), Timedelta,
    datetime.time, « 2j 3h 15m » et numérique (interprété en MINUTES)."""
    if v is None:
        return 0.0
    if isinstance(v, pd.Timedelta):
        return max(v.total_seconds(), 0.0)
    if isinstance(v, _dt.timedelta):
        return max(v.total_seconds(), 0.0)
    if isinstance(v, _dt.time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return 0.0
        return max(float(v), 0.0) * 60.0
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none', '-', 'n/a'):
        return 0.0
    m = re.fullmatch(r'(\d+):(\d{1,2})(?::(\d{1,2}))?(?:\.\d+)?', s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3) or 0)
    total, found = 0.0, False
    for num, unit in re.findall(r'(\d+(?:[.,]\d+)?)\s*(jours?|j|d|heures?|hrs?|h|min(?:utes?)?|mn|m|sec(?:ondes?)?|s)',
                                s.lower()):
        n = float(num.replace(',', '.'))
        found = True
        if unit[0] in ('j', 'd'):
            total += n * 86400
        elif unit[0] == 'h':
            total += n * 3600
        elif unit[0] == 'm':
            total += n * 60
        else:
            total += n
    if found:
        return total
    try:
        return max(float(s.replace(',', '.')), 0.0) * 60.0
    except ValueError:
        return 0.0


_NAN_RE = re.compile(r'(?i)^(nan|nat|none|null|n/?a|<na>|-)$')


def _to_display(v) -> str:
    """Valeur source → chaîne lisible JSON-safe (dates, durées, nombres…)."""
    if v is None:
        return ''
    try:
        if pd.isna(v):          # NaN, NaT, pd.NA (scalaires uniquement)
            return ''
    except (TypeError, ValueError):
        pass
    if isinstance(v, (pd.Timestamp, _dt.datetime)):
        return v.strftime('%d/%m/%Y %H:%M:%S')
    if isinstance(v, _dt.date):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, (pd.Timedelta, _dt.timedelta)):
        return _fmt(v.total_seconds())
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return '' if _NAN_RE.fullmatch(s) else s


def _clean_str_series(s: pd.Series) -> pd.Series:
    # fillna AVANT astype : en pandas 3.x, astype(str) préserve les NaN
    s = s.fillna('').astype(str).str.strip()
    return s.where(~s.str.fullmatch(_NAN_RE).fillna(False), '')


def _enrich_from_sites(out: pd.DataFrame) -> None:
    """Complète région / base / classification manquantes via la table Site."""
    try:
        from .models import Site
        rows = list(Site.objects.all().values(
            'site_name', 'region', 'base', 'zone', 'classif_tech', 'type_site'))
    except Exception:
        return
    lut: dict[str, dict] = {}
    for s in rows:
        key = str(s.get('site_name') or '').strip().upper()
        if key:
            lut[key] = s
    if not lut:
        return
    keys = out['site'].str.upper().str.strip()
    plans = (
        ('region',         ('region',)),
        ('base',           ('base', 'zone')),
        ('classification', ('classif_tech', 'type_site')),
    )
    for col, fields in plans:
        need = out[col] == ''
        if not need.any():
            continue

        def _lookup(k, _fields=fields):
            rec = lut.get(k)
            if not rec:
                return ''
            for f in _fields:
                v = str(rec.get(f) or '').strip()
                if v:
                    return v
            return ''
        out.loc[need, col] = keys[need].map(_lookup)


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame source (fichier ou API) → schéma canonique Analytics."""
    if df is None or df.empty:
        raise ValueError('La source ne contient aucune ligne.')
    mapping = _map_columns(df)
    if 'site' not in mapping:
        raise ValueError(
            "Colonne « Site » introuvable. Colonnes attendues : Région, Base, Site, "
            "Équipement en défaut, Cause, Classification du site, Durée de l'incident, Escalade.")

    out = pd.DataFrame(index=df.index)
    for canon in ('region', 'base', 'site', 'equipement', 'cause',
                  'classification', 'escalade', 'status'):
        if canon in mapping:
            out[canon] = _clean_str_series(df[mapping[canon]])
        else:
            out[canon] = ''

    out['duration_sec'] = (df[mapping['duration']].map(_dur_to_sec)
                           if 'duration' in mapping else 0.0)

    # Cause : repli sur « Root Cause » NetXcare quand la colonne Cause est vide
    if (out['cause'] == '').any():
        rc_col = next((c for c in df.columns
                       if c != mapping.get('cause')
                       and _norm(c) in ('root cause', 'cause racine')), None)
        if rc_col is not None:
            rc = _clean_str_series(df[rc_col])
            need = out['cause'] == ''
            out.loc[need, 'cause'] = rc[need]

    if 'date' in mapping:
        d = pd.to_datetime(df[mapping['date']], dayfirst=True, format='mixed', errors='coerce')
        out['date'] = d.dt.strftime('%Y-%m-%d').fillna('')
    else:
        out['date'] = ''

    # Champs NetXcare supplémentaires (varient selon la plateforme) : toutes
    # les colonnes source non mappées et renseignées sont conservées telles
    # quelles pour l'exploration détaillée (section 8).
    mapped_srcs = set(mapping.values())
    # L'horodatage complet (Alarm Time…) reste visible dans le détail — le
    # champ canonique « date » ne garde que le jour.
    mapped_srcs.discard(mapping.get('date'))
    extras_data: dict[str, pd.Series] = {}
    for c in df.columns:
        name = str(c).strip()
        if c in mapped_srcs or not name or name.lower().startswith('unnamed'):
            continue
        s = df[c].map(_to_display)
        if (s != '').any():
            extras_data[name] = s
    if extras_data:
        names = list(extras_data)
        out['extra'] = [{n: v for n, v in zip(names, vals) if v}
                        for vals in zip(*extras_data.values())]
    else:
        out['extra'] = [{} for _ in range(len(df))]

    out = out[out['site'] != ''].copy()
    if out.empty:
        raise ValueError('Aucune ligne avec un nom de site exploitable.')

    _enrich_from_sites(out)

    for c in ('region', 'base', 'equipement', 'cause', 'classification', 'escalade'):
        out[c] = out[c].replace('', EMPTY_LABEL)
    return out.reset_index(drop=True)


# ═════════════════════ 2. Chargement des sources ════════════════════════════

def read_uploaded(uploaded_file) -> pd.DataFrame:
    """Lit un fichier importé (xlsx / xls / csv) en DataFrame brut.

    Gère aussi les classeurs multi-feuilles (« Base de connaissances » — une
    feuille par mois, ex. JANVIER/FEVRIER/…) : concatène toutes les feuilles
    dont l'en-tête contient une colonne Site exploitable, en ignorant les
    feuilles de synthèse/tableau croisé sans en-tête utilisable (ex. « Feuil1 »
    avec des colonnes « Unnamed: N »)."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    if name.endswith('.csv'):
        try:
            return pd.read_csv(uploaded_file, sep=None, engine='python', encoding='utf-8-sig')
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, sep=None, engine='python', encoding='latin-1')
    if name.endswith(('.xlsx', '.xls', '.xlsm')):
        xls = pd.ExcelFile(uploaded_file)
        if len(xls.sheet_names) == 1:
            return pd.read_excel(xls, sheet_name=xls.sheet_names[0])
        frames = []
        for sh in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sh)
            if df.empty or not any('site' in _norm(c) for c in df.columns):
                continue
            frames.append(df)
        if frames:
            return pd.concat(frames, ignore_index=True, sort=False)
        # Aucune feuille avec colonne Site reconnaissable : comportement d'origine
        return pd.read_excel(xls, sheet_name=xls.sheet_names[0])
    raise ValueError('Format non supporté — importez un fichier .xlsx, .xls ou .csv.')


# Alarmes retenues pour le réseau Mobile — même filtre que le rapport
# officiel (treatement.process_file) : le trafic BTS/WCDMA génère un grand
# nombre d'alarmes annexes non représentatives d'une indisponibilité de site.
_MOBILE_ALARMES_A_GARDER = [
    "BTS O&M LINK FAILURE / WCDMA BASE STATION OUT OF USE",
    "WCDMA BASE STATION OUT OF USE",
    "BTS O&M LINK FAILURE",
    "ALL RFMS MISSING",
    "WCDMA CELL OUT OF USE",
]


def _filter_mobile_alarms(df: pd.DataFrame) -> pd.DataFrame:
    """Restreint aux types d'alarme officiels du réseau Mobile (colonne « Alarm text »)."""
    if 'Alarm text' not in df.columns:
        return df
    return df[df['Alarm text'].astype(str).str.strip().isin(_MOBILE_ALARMES_A_GARDER)].copy()


def _prepare_api_rows(df: pd.DataFrame, d1: _dt.date, d2: _dt.date) -> pd.DataFrame:
    """Aligne les données API sur la logique des rapports officiels (treatement.py) :

    1. borne les durées à la période demandée (un ticket commencé avant ou
       toujours ouvert est compté de 00:00:00 à 23:59:59 au maximum) ;
    2. dédoublonne les pannes : l'export NetXcare contient une ligne par site
       enfant / équipement pour un même incident → on garde une seule ligne
       par (site racine, heure d'alarme), comme la synthèse journalière.
    """
    if 'Alarm Time' not in df.columns:
        return df
    at = pd.to_datetime(df['Alarm Time'], dayfirst=True, format='mixed', errors='coerce')
    ct = pd.to_datetime(df.get('Cancel Time', pd.Series(index=df.index, dtype='object')),
                        dayfirst=True, format='mixed', errors='coerce')
    debut = pd.Timestamp(f'{d1.isoformat()} 00:00:00')
    fin   = pd.Timestamp(f'{d2.isoformat()} 23:59:59')

    # Tickets actifs pendant la période uniquement
    keep = at.notna() & (at <= fin) & (ct.isna() | (ct >= debut))
    df, at, ct = df[keep].copy(), at[keep], ct[keep]

    # Durée bornée à la période (recalculée depuis Alarm/Cancel Time)
    at_c = at.clip(lower=debut)
    ct_c = ct.fillna(fin).clip(upper=fin).where(lambda s: s >= at_c, at_c)
    df['Duration'] = ct_c - at_c            # Timedelta → géré par _dur_to_sec
    df['Alarm Time'] = at_c

    # Dédoublonnage : même site racine + même heure d'alarme = même panne
    if 'Site Name' in df.columns:
        racine = None
        if 'Site Parent' in df.columns:
            racine = _clean_str_series(df['Site Parent']).replace('', pd.NA)
            racine = racine.fillna(_clean_str_series(df['Site Name']))
        else:
            racine = _clean_str_series(df['Site Name'])
        avant = len(df)
        df = df.loc[~pd.DataFrame({'r': racine, 't': at_c}).duplicated(keep='first')]
        if len(df) != avant:
            import logging
            logging.getLogger(__name__).info(
                'Analytics API: dédoublonnage %d → %d lignes', avant, len(df))
    return df


def prepare_source_dataframe(df: pd.DataFrame, filename: str = '') -> pd.DataFrame:
    """Prépare un fichier importé de type export NetXcare (colonnes Alarm Time /
    Cancel Time / Site Name) : borne les durées à la période et dédoublonne,
    comme pour la source API. Période déduite du nom de fichier
    (…_YYYYMMDD_YYYYMMDD…), sinon des dates min/max du fichier.
    Les fichiers déjà agrégés (sans Alarm Time) sont retournés tels quels."""
    cols = {_norm(c) for c in df.columns}
    if 'alarm time' not in cols or 'site name' not in cols:
        return df

    d1 = d2 = None
    m = re.search(r'(\d{8})[_-](\d{8})', filename or '')
    if m:
        try:
            d1 = _dt.datetime.strptime(m.group(1), '%Y%m%d').date()
            d2 = _dt.datetime.strptime(m.group(2), '%Y%m%d').date()
        except ValueError:
            d1 = d2 = None
    if d1 is None:
        col_at = next(c for c in df.columns if _norm(c) == 'alarm time')
        at = pd.to_datetime(df[col_at], dayfirst=True, format='mixed', errors='coerce')
        if at.notna().any():
            d1, d2 = at.min().date(), at.max().date()
        else:
            return df
    return _prepare_api_rows(df, d1, d2)


def fetch_api_dataframe(date_debut: str, date_fin: str, network: str = 'mobile') -> pd.DataFrame:
    """Récupère l'historique des incidents via l'API ticketing (NetXcare).

    Découpe la période en tranches de 30 jours pour éviter les erreurs
    « 413 Payload Too Large » sur les longues plages.
    """
    from django.conf import settings
    from .api_client import TicketingApiClient
    from .api_import import json_to_dataframe

    api_url  = getattr(settings, 'TICKETING_API_URL', '') or ''
    api_user = getattr(settings, 'TICKETING_API_USERNAME', '') or ''
    api_pass = getattr(settings, 'TICKETING_API_PASSWORD', '') or ''
    if not (api_url and api_user and api_pass):
        raise RuntimeError(
            'API NetXcare non configurée — renseignez TICKETING_API_URL, '
            'TICKETING_API_USERNAME et TICKETING_API_PASSWORD dans .env.')

    d1 = _dt.date.fromisoformat(str(date_debut)[:10])
    d2 = _dt.date.fromisoformat(str(date_fin)[:10])
    if d2 < d1:
        d1, d2 = d2, d1

    client = TicketingApiClient(api_url)
    client.login(api_user, api_pass)
    net = None if network in ('', 'all') else network
    plateformes_id = client.get_plateformes_ids_for_network(net) if net else None

    # Plage élargie (-7 j / +1 j) pour capter les pannes multi-jours commencées
    # avant la période — les durées sont ensuite bornées par _prepare_api_rows.
    q1 = d1 - _dt.timedelta(days=7)
    q2 = d2 + _dt.timedelta(days=1)
    rows: list[dict] = []
    cur = q1
    while cur <= q2:
        end = min(cur + _dt.timedelta(days=29), q2)
        rows.extend(client.export_data(
            cur.isoformat(), end.isoformat(),
            plateformes_id=plateformes_id, network=net))
        cur = end + _dt.timedelta(days=1)

    if not rows:
        raise ValueError(
            f'Aucune donnée retournée par l’API pour {d1} → {d2} (réseau : {network}).')
    raw = json_to_dataframe(rows)
    if network == 'mobile':
        raw = _filter_mobile_alarms(raw)
        if raw.empty:
            raise ValueError(
                f'Aucune alarme Mobile retenue (BTS O&M LINK FAILURE / WCDMA BASE '
                f'STATION OUT OF USE / ALL RFMS MISSING) pour {d1} → {d2}.')
    return _prepare_api_rows(raw, d1, d2)


# ═════════════════════ 3. Persistance session ═══════════════════════════════

def save_normalized(df: pd.DataFrame, path: str) -> None:
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(df.to_dict('records'), fh, ensure_ascii=False)


def load_normalized(path: str) -> pd.DataFrame:
    with open(path, 'r', encoding='utf-8') as fh:
        records = json.load(fh)
    df = pd.DataFrame(records)
    if not df.empty:
        df['duration_sec'] = pd.to_numeric(df['duration_sec'], errors='coerce').fillna(0.0)
        # Robustesse : aucun NaN ne doit subsister dans les colonnes texte
        for c in ('region', 'base', 'equipement', 'cause', 'classification', 'escalade'):
            if c in df.columns:
                df[c] = df[c].fillna(EMPTY_LABEL).replace('', EMPTY_LABEL)
        for c in ('site', 'status', 'date'):
            if c in df.columns:
                df[c] = df[c].fillna('')
        if 'extra' in df.columns:
            df['extra'] = df['extra'].map(lambda v: v if isinstance(v, dict) else {})
        else:
            df['extra'] = [{} for _ in range(len(df))]
    return df


# ═════════════════════ 4. Calcul des analyses ═══════════════════════════════

def _fmt(sec: float) -> str:
    sec = max(float(sec or 0), 0)
    return f'{int(sec // 3600)}:{int(sec % 3600 // 60):02d}:{int(sec % 60):02d}'


def _h(sec: float) -> float:
    return round(float(sec or 0) / 3600.0, 2)


def apply_filters(df: pd.DataFrame, *, date_debut: str = '', date_fin: str = '',
                  regions=(), causes=(), sites=()) -> pd.DataFrame:
    f = df
    if date_debut:
        f = f[(f['date'] != '') & (f['date'] >= str(date_debut)[:10])]
    if date_fin:
        f = f[(f['date'] != '') & (f['date'] <= str(date_fin)[:10])]
    if regions:
        f = f[f['region'].isin(list(regions))]
    if causes:
        f = f[f['cause'].isin(list(causes))]
    if sites:
        f = f[f['site'].isin(list(sites))]
    return f


def _top_label_per_group(f: pd.DataFrame, group_col: str, label_col: str,
                         by: str = 'duration_sec') -> dict[str, str]:
    """Pour chaque valeur de group_col : libellé de label_col dominant."""
    if f.empty:
        return {}
    if by == 'count':
        g = f.groupby([group_col, label_col]).size()
    else:
        g = f.groupby([group_col, label_col])['duration_sec'].sum()
    idx = g.groupby(level=0).idxmax()
    return {k: v[1] for k, v in idx.items()}


# Champs canoniques utilisables comme dimension de croisement (section 8)
_CANON_KEYS = ('date', 'region', 'base', 'site', 'equipement', 'cause',
               'classification', 'escalade', 'status')


def field_series(f: pd.DataFrame, key: str) -> pd.Series:
    """Série de valeurs (str) d'un champ canonique ou d'un champ NetXcare."""
    if key in _CANON_KEYS and key in f.columns:
        return f[key].fillna('').astype(str)
    if 'extra' in f.columns:
        return f['extra'].map(
            lambda d: str(d.get(key, '') or '').strip() if isinstance(d, dict) else '')
    return pd.Series('', index=f.index, dtype=object)


def drill_fields(df: pd.DataFrame) -> list[dict]:
    """Champs disponibles pour le croisement : canoniques + NetXcare (extra).
    Les champs de travail usuels (Site Parent, Site, Région, Plateforme…)
    sont proposés en premier."""
    fields = [{'key': k, 'label': FR_LABELS.get(k, k)}
              for k in _CANON_KEYS if k in df.columns]
    if 'extra' in df.columns:
        seen: set[str] = set()
        for d in df['extra']:
            if isinstance(d, dict):
                seen.update(str(k) for k in d)
        fields += [{'key': k, 'label': k} for k in sorted(seen)]
    priority = ['Site Parent', 'site', 'region', 'Plateforme', 'Technologies',
                'Alarm text', 'cause', 'escalade', 'Root Cause', 'base',
                'equipement', 'classification']
    rank = {k: i for i, k in enumerate(priority)}
    fields.sort(key=lambda f: (rank.get(f['key'], len(priority)), f['label']))
    return fields


def drill_pivot(f: pd.DataFrame, dims: list[str], filters=(),
                limit: int = 300) -> dict:
    """Croisement libre : agrège incidents + outage sur 1 à 4 champs choisis
    (canoniques ou NetXcare) ; filtres (champ, valeur) indépendants du
    croisement, applicables à n'importe quel champ."""
    dims = [d for d in dims if d][:4] or ['region', 'site', 'cause']
    labels = [FR_LABELS.get(d, d) for d in dims]
    filters = [(k, v) for k, v in filters if k]
    base = {'dims': dims, 'labels': labels, 'rows': [], 'total_groups': 0,
            'filter_options': [[] for _ in filters]}
    if f.empty:
        return base

    fseries = {k: field_series(f, k).replace('', EMPTY_LABEL) for k, _ in filters}

    # Valeurs proposées pour chaque filtre : les AUTRES filtres appliqués
    filter_options: list[list[str]] = []
    for i, (k, _v) in enumerate(filters):
        mask = pd.Series(True, index=f.index)
        for j, (k2, v2) in enumerate(filters):
            if j != i and v2:
                mask &= fseries[k2] == v2
        filter_options.append(sorted(str(x) for x in fseries[k][mask].unique()))
    base['filter_options'] = filter_options

    mask = pd.Series(True, index=f.index)
    for k, v in filters:
        if v:
            mask &= fseries[k] == v
    ff = f[mask]
    if ff.empty:
        return base

    sub = pd.DataFrame({f'_d{i}': field_series(ff, d).replace('', EMPTY_LABEL)
                        for i, d in enumerate(dims)})
    sub['duration_sec'] = pd.to_numeric(ff['duration_sec'], errors='coerce').fillna(0.0).values

    g = (sub.groupby([f'_d{i}' for i in range(len(dims))])['duration_sec']
         .agg(['size', 'sum']).sort_values('sum', ascending=False))
    total_groups = int(len(g))
    rows = []
    for idx, r in g.head(limit).iterrows():
        keyvals = list(idx) if isinstance(idx, tuple) else [idx]
        rows.append({'k': [str(x) for x in keyvals], 'n': int(r['size']),
                     'outage': _fmt(r['sum']), 'outage_h': _h(r['sum'])})
    return {**base, 'rows': rows, 'total_groups': total_groups}


def drill_details(f: pd.DataFrame, filters=(), limit: int = 300) -> dict:
    """Incidents individuels correspondant aux filtres (champ, valeur) — avec
    TOUS les champs NetXcare renseignés, colonnes dynamiques selon la plateforme."""
    for key, val in filters:
        if not val:
            continue
        s = field_series(f, key).replace('', EMPTY_LABEL)
        f = f[s == val]
    total = int(len(f))
    f = f.sort_values('duration_sec', ascending=False).head(limit)

    rows_d: list[dict] = []
    extra_keys: list[str] = []
    seen: set[str] = set()
    has_extra = 'extra' in f.columns
    for _, r in f.iterrows():
        d: dict = {}
        for k in _CANON_KEYS:
            v = str(r.get(k) or '')
            d[FR_LABELS.get(k, k)] = '' if v == EMPTY_LABEL else v
        d['Durée'] = _fmt(r.get('duration_sec') or 0)
        extra = r.get('extra') if has_extra else None
        if isinstance(extra, dict):
            for k, v in extra.items():
                v = str(v or '').strip()
                if not v:
                    continue
                d[str(k)] = v
                if k not in seen:
                    seen.add(k)
                    extra_keys.append(str(k))
        rows_d.append(d)

    columns = [FR_LABELS.get(k, k) for k in _CANON_KEYS] + ['Durée'] + extra_keys
    columns = [c for c in columns if any(row.get(c) for row in rows_d)]
    return {
        'columns': columns,
        'rows':    [[row.get(c, '') for c in columns] for row in rows_d],
        'total':   total,
        'shown':   len(rows_d),
    }


def compute(df: pd.DataFrame, *, date_debut: str = '', date_fin: str = '',
            regions=(), causes=(), sites=()) -> dict:
    """Calcule KPIs + les 10 axes d'analyse. Retourne un dict JSON-safe."""
    dates_ok = df[df['date'] != '']['date']
    filters = {
        'regions':  sorted(x for x in df['region'].unique() if x != EMPTY_LABEL),
        'causes':   sorted(x for x in df['cause'].unique() if x != EMPTY_LABEL),
        'sites':    sorted(df['site'].unique()),
        'date_min': str(dates_ok.min()) if not dates_ok.empty else '',
        'date_max': str(dates_ok.max()) if not dates_ok.empty else '',
    }

    f = apply_filters(df, date_debut=date_debut, date_fin=date_fin,
                      regions=regions, causes=causes, sites=sites)
    if f.empty:
        return {'empty': True, 'filters': filters}

    # ── KPIs ────────────────────────────────────────────────────────────────
    n_inc  = int(len(f))
    outage = float(f['duration_sec'].sum())
    mttr   = outage / n_inc if n_inc else 0.0
    cause_out = f.groupby('cause')['duration_sec'].sum().sort_values(ascending=False)
    f_dates = f[f['date'] != '']['date']
    kpi = {
        'incidents': n_inc,
        'outage':    _fmt(outage),
        'outage_h':  _h(outage),
        'mttr':      _fmt(mttr),
        'sites':     int(f['site'].nunique()),
        'regions':   int(f['region'].nunique()),
        'top_cause': str(cause_out.index[0]) if not cause_out.empty else '—',
        'period':    (f'{f_dates.min()} → {f_dates.max()}' if not f_dates.empty else '—'),
    }

    # ── 1. Région & Site — outage (histogramme empilé) ──────────────────────
    # Chaque région est ordonnée par outage total (desc). À l'intérieur de
    # chaque barre, les sites sont empilés selon LEUR PROPRE classement
    # d'indisponibilité (le site le plus indisponible de la région en bas).
    reg_out  = f.groupby('region')['duration_sec'].sum().sort_values(ascending=False)
    regions_ = [str(r) for r in reg_out.index]
    site_out = f.groupby('site')['duration_sec'].sum().sort_values(ascending=False)
    pv = f.pivot_table(index='region', columns='site', values='duration_sec',
                       aggfunc='sum', fill_value=0)
    TOP_N = 6
    a1_datasets = [{'label': f'Site #{i + 1}', 'data': [], 'siteNames': []}
                   for i in range(TOP_N)]
    autres = []
    for r in regions_:
        row = pv.loc[r] if r in pv.index else pd.Series(dtype=float)
        ranked = row[row > 0].sort_values(ascending=False)
        for i in range(TOP_N):
            if i < len(ranked):
                a1_datasets[i]['data'].append(_h(ranked.iloc[i]))
                a1_datasets[i]['siteNames'].append(str(ranked.index[i]))
            else:
                a1_datasets[i]['data'].append(0)
                a1_datasets[i]['siteNames'].append('')
        tot = float(reg_out.get(r, 0))
        tops = float(ranked.head(TOP_N).sum())
        autres.append(_h(max(tot - tops, 0)))
    if any(v > 0 for v in autres):
        a1_datasets.append({'label': 'Autres sites', 'data': autres})
    a1 = {'labels': regions_, 'datasets': a1_datasets}

    # ── 2. Site & Cause — outage (barres horizontales, top 15) ──────────────
    cause_by_site = _top_label_per_group(f, 'site', 'cause')
    n_by_site = f.groupby('site').size()
    a2 = [{'site': str(s), 'outage_h': _h(v), 'outage': _fmt(v),
           'n': int(n_by_site.get(s, 0)), 'cause': str(cause_by_site.get(s, '—'))}
          for s, v in site_out.head(15).items()]

    # ── 3. Cause & Site — incidents + outage (double axe) ───────────────────
    g3 = f.groupby('cause').agg(n=('cause', 'size'), outage=('duration_sec', 'sum'))
    g3 = g3.sort_values('outage', ascending=False)
    site_by_cause = _top_label_per_group(f, 'cause', 'site')
    a3 = {
        'labels':   [str(c) for c in g3.head(12).index],
        'counts':   [int(v) for v in g3.head(12)['n']],
        'outage_h': [_h(v) for v in g3.head(12)['outage']],
        'table':    [{'cause': str(c), 'n': int(r['n']), 'outage': _fmt(r['outage']),
                      'outage_h': _h(r['outage']), 'top_site': str(site_by_cause.get(c, '—'))}
                     for c, r in g3.iterrows()],
    }

    # ── 4. Région & Site — synthèse régionale ───────────────────────────────
    g4 = f.groupby('region').agg(n=('region', 'size'), outage=('duration_sec', 'sum'),
                                 sites=('site', 'nunique'))
    g4 = g4.sort_values('outage', ascending=False)
    top_site_by_reg = _top_label_per_group(f, 'region', 'site')
    a4 = [{'region': str(r), 'sites': int(row['sites']), 'n': int(row['n']),
           'outage': _fmt(row['outage']), 'outage_h': _h(row['outage']),
           'mttr': _fmt(row['outage'] / row['n'] if row['n'] else 0),
           'top_site': str(top_site_by_reg.get(r, '—'))}
          for r, row in g4.iterrows()]

    # ── 5. Classification & Site — incidents + cause principale ─────────────
    g5 = f.groupby('classification').agg(n=('classification', 'size'),
                                         outage=('duration_sec', 'sum'),
                                         sites=('site', 'nunique'))
    g5 = g5.sort_values('n', ascending=False)
    cause_by_classif = _top_label_per_group(f, 'classification', 'cause', by='count')
    a5 = [{'classification': str(c), 'sites': int(row['sites']), 'n': int(row['n']),
           'outage': _fmt(row['outage']), 'outage_h': _h(row['outage']),
           'top_cause': str(cause_by_classif.get(c, '—'))}
          for c, row in g5.iterrows()]

    # ── 6. Cause — Pareto 80/20 (outage cumulé) ─────────────────────────────
    total_out = float(cause_out.sum()) or 1.0
    cum, a6_labels, a6_h, a6_cum = 0.0, [], [], []
    for c, v in cause_out.head(15).items():
        cum += float(v)
        a6_labels.append(str(c))
        a6_h.append(_h(v))
        a6_cum.append(round(cum / total_out * 100, 1))
    a6 = {'labels': a6_labels, 'outage_h': a6_h, 'cum_pct': a6_cum}

    # ── 7. Équipement en défaut — top 10 des moins fiables ──────────────────
    f7 = f[f['equipement'] != EMPTY_LABEL]
    if f7.empty:
        f7 = f
    g7 = f7.groupby('equipement').agg(n=('equipement', 'size'),
                                      outage=('duration_sec', 'sum'))
    g7 = g7.sort_values(['n', 'outage'], ascending=False).head(10)
    a7 = [{'equipement': str(e), 'n': int(row['n']),
           'outage': _fmt(row['outage']), 'outage_h': _h(row['outage'])}
          for e, row in g7.iterrows()]

    # ── 8. Croisement libre — calculé à la demande via drill_pivot (AJAX) ───

    # ── 9. Escalade & Cause — répartition des responsabilités ───────────────
    esc_out = f.groupby('escalade')['duration_sec'].sum().sort_values(ascending=False)
    g9 = f.groupby(['escalade', 'cause']).agg(n=('escalade', 'size'),
                                              outage=('duration_sec', 'sum'))
    g9 = g9.sort_values('outage', ascending=False).head(20)
    a9 = {
        'labels':   [str(e) for e in esc_out.index],
        'outage_h': [_h(v) for v in esc_out.values],
        'table':    [{'escalade': str(e), 'cause': str(c), 'n': int(row['n']),
                      'outage': _fmt(row['outage']),
                      'pct': round(float(row['outage']) / total_out * 100, 1)}
                     for (e, c), row in g9.iterrows()],
    }

    # ── 10. Base & Site — performance opérationnelle par base ───────────────
    g10 = f.groupby('base').agg(n=('base', 'size'), outage=('duration_sec', 'sum'),
                                sites=('site', 'nunique'))
    g10 = g10.sort_values('outage', ascending=False)
    top_site_by_base = _top_label_per_group(f, 'base', 'site')
    a10 = [{'base': str(b), 'sites': int(row['sites']), 'n': int(row['n']),
            'outage': _fmt(row['outage']), 'outage_h': _h(row['outage']),
            'mttr': _fmt(row['outage'] / row['n'] if row['n'] else 0),
            'top_site': str(top_site_by_base.get(b, '—'))}
           for b, row in g10.iterrows()]

    return {
        'empty': False, 'filters': filters, 'kpi': kpi,
        'a1': a1, 'a2': a2, 'a3': a3, 'a4': a4, 'a5': a5,
        'a6': a6, 'a7': a7, 'a9': a9, 'a10': a10,
    }


# ═════════════════════ 5. Exports Excel / PDF ═══════════════════════════════

def build_excel(df_filtered: pd.DataFrame, res: dict) -> io.BytesIO:
    """Classeur Excel : données filtrées + une feuille par axe d'analyse."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as xw:
        kpi = res.get('kpi', {})
        pd.DataFrame([
            ('Incidents', kpi.get('incidents')),
            ('Indisponibilité totale', kpi.get('outage')),
            ('Indisponibilité (heures)', kpi.get('outage_h')),
            ('MTTR', kpi.get('mttr')),
            ('Sites impactés', kpi.get('sites')),
            ('Régions', kpi.get('regions')),
            ('Cause dominante', kpi.get('top_cause')),
            ('Période', kpi.get('period')),
        ], columns=['Indicateur', 'Valeur']).to_excel(xw, sheet_name='KPIs', index=False)

        cols = [c for c in ('date', 'region', 'base', 'site', 'equipement', 'cause',
                            'classification', 'escalade', 'status', 'duration_sec')
                if c in df_filtered.columns]
        df_data = df_filtered[cols].copy()
        # Durée lisible HH:MM:SS (la valeur en secondes reste disponible à côté)
        df_data.insert(cols.index('duration_sec'), 'duration_hms',
                       df_data['duration_sec'].map(_fmt))
        df_data.rename(columns={**FR_LABELS, 'duration_hms': 'Durée (HH:MM:SS)'}
                       ).to_excel(xw, sheet_name='Données', index=False)

        def _sheet(name, rows, columns):
            pd.DataFrame(rows).reindex(columns=list(columns)).rename(
                columns=columns).to_excel(xw, sheet_name=name, index=False)

        _sheet('4_Synthèse régions', res['a4'],
               {'region': 'Région', 'sites': 'Sites', 'n': 'Incidents',
                'outage': 'Outage', 'outage_h': 'Outage (h)', 'mttr': 'MTTR',
                'top_site': 'Site le plus impacté'})
        _sheet('2_Top sites', res['a2'],
               {'site': 'Site', 'n': 'Incidents', 'outage': 'Outage',
                'outage_h': 'Outage (h)', 'cause': 'Cause principale'})
        _sheet('3_Causes', res['a3']['table'],
               {'cause': 'Cause', 'n': 'Incidents', 'outage': 'Outage',
                'outage_h': 'Outage (h)', 'top_site': 'Site le plus impacté'})
        _sheet('5_Classification', res['a5'],
               {'classification': 'Classification', 'sites': 'Sites', 'n': 'Incidents',
                'outage': 'Outage', 'outage_h': 'Outage (h)', 'top_cause': 'Cause principale'})
        pd.DataFrame({'Cause': res['a6']['labels'],
                      'Outage (h)': res['a6']['outage_h'],
                      'Cumul (%)': res['a6']['cum_pct']}).to_excel(
            xw, sheet_name='6_Pareto causes', index=False)
        _sheet('7_Équipements', res['a7'],
               {'equipement': 'Équipement en défaut', 'n': 'Incidents',
                'outage': 'Outage', 'outage_h': 'Outage (h)'})
        _sheet('9_Escalade x Cause', res['a9']['table'],
               {'escalade': 'Escalade', 'cause': 'Cause', 'n': 'Incidents',
                'outage': 'Outage', 'pct': '% outage total'})
        _sheet('10_Bases', res['a10'],
               {'base': 'Base', 'sites': 'Sites', 'n': 'Incidents',
                'outage': 'Outage', 'outage_h': 'Outage (h)', 'mttr': 'MTTR',
                'top_site': 'Site le plus impacté'})
    buf.seek(0)
    return buf


# Sections du dashboard qui affichent un graphique Chart.js (les autres
# n'ont qu'un tableau HTML, déjà exportable côté client via SheetJS).
SECTION_CHART_TITLES = {
    'a1': "Indisponibilité par Région & Site",
    'a6': 'Pareto des causes (80/20)',
    'a2': 'Indisponibilité par Site & Cause (Top 15)',
    'a3': 'Corrélation Cause & Site',
    'a7': 'Top 10 des équipements les moins fiables',
    'a9': 'Répartition des responsabilités — Escalade & Cause',
}


def build_section_excel(res: dict, key: str) -> io.BytesIO:
    """Classeur Excel pour UNE case du dashboard, données + graphique natif
    Excel reproduisant le graphique Chart.js affiché sur la page."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    if key not in SECTION_CHART_TITLES:
        raise ValueError(f'Section inconnue : {key}')
    title = SECTION_CHART_TITLES[key]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Données'

    def _style_header(sheet, ncols):
        for c in range(1, ncols + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = PatternFill('solid', fgColor='003087')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        sheet.freeze_panes = 'A2'

    if key == 'a1':
        a1 = res['a1']
        cols = [d['label'] for d in a1['datasets']]
        ws.append(['Région'] + cols)
        for i, region in enumerate(a1['labels']):
            ws.append([region] + [d['data'][i] for d in a1['datasets']])
        _style_header(ws, len(cols) + 1)
        n = len(a1['labels'])
        chart = BarChart()
        chart.type, chart.grouping, chart.overlap = 'col', 'stacked', 100
        chart.title = title
        chart.y_axis.title = 'Outage (h)'
        chart.add_data(Reference(ws, min_col=2, max_col=1 + len(cols), min_row=1, max_row=1 + n),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        chart.width, chart.height = 24, 12
        ws.add_chart(chart, f'A{n + 4}')

        # Détail : à quel site correspond chaque rang, par région
        ws2 = wb.create_sheet('Détail Sites')
        ws2.append(['Région', 'Rang', 'Site', 'Outage (h)'])
        for i, region in enumerate(a1['labels']):
            for d in a1['datasets']:
                if d['label'] == 'Autres sites':
                    continue
                name = d.get('siteNames', [''] * n)[i]
                if name:
                    ws2.append([region, d['label'], name, d['data'][i]])
        _style_header(ws2, 4)

    elif key == 'a6':
        a6 = res['a6']
        ws.append(['Cause', 'Outage (h)', 'Cumul (%)'])
        for c, h, cum in zip(a6['labels'], a6['outage_h'], a6['cum_pct']):
            ws.append([c, h, cum])
        _style_header(ws, 3)
        n = len(a6['labels'])
        bar = BarChart()
        bar.type = 'col'
        bar.title = title
        bar.y_axis.title = 'Outage (h)'
        bar.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=1 + n), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        line = LineChart()
        line.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=1 + n), titles_from_data=True)
        line.y_axis.axId = 200
        line.y_axis.title = 'Cumul (%)'
        bar.y_axis.crosses = 'max'
        bar += line
        bar.width, bar.height = 24, 12
        ws.add_chart(bar, f'A{n + 4}')

    elif key == 'a2':
        rows = res['a2']
        ws.append(['Site', 'Incidents', 'Outage (h)', 'Cause principale'])
        for r in rows:
            ws.append([r['site'], r['n'], r['outage_h'], r['cause']])
        _style_header(ws, 4)
        n = len(rows)
        bar = BarChart()
        bar.type = 'bar'   # horizontal, comme sur la page
        bar.title = title
        bar.x_axis.title = 'Outage (h)'
        bar.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=1 + n), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        bar.width, bar.height = 22, 18
        ws.add_chart(bar, f'A{n + 4}')

    elif key == 'a3':
        t = res['a3']['table']
        ws.append(['Cause', 'Incidents', 'Outage (h)', 'Site le plus impacté'])
        for r in t:
            ws.append([r['cause'], r['n'], r['outage_h'], r['top_site']])
        _style_header(ws, 4)
        n = len(t)
        bar = BarChart()
        bar.type = 'col'
        bar.title = title
        bar.y_axis.title = 'Incidents'
        bar.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=1 + n), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        line = LineChart()
        line.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=1 + n), titles_from_data=True)
        line.y_axis.axId = 200
        line.y_axis.title = 'Outage (h)'
        bar.y_axis.crosses = 'max'
        bar += line
        bar.width, bar.height = 24, 12
        ws.add_chart(bar, f'A{n + 4}')

    elif key == 'a7':
        rows = res['a7']
        ws.append(['Équipement en défaut', 'Incidents', 'Outage (h)'])
        for r in rows:
            ws.append([r['equipement'], r['n'], r['outage_h']])
        _style_header(ws, 3)
        n = len(rows)
        bar = BarChart()
        bar.type = 'bar'   # horizontal, comme sur la page
        bar.title = title
        bar.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1 + n), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        bar.width, bar.height = 22, 16
        ws.add_chart(bar, f'A{n + 4}')

    elif key == 'a9':
        a9 = res['a9']
        ws.append(['Escalade', 'Outage (h)'])
        for lbl, h in zip(a9['labels'], a9['outage_h']):
            ws.append([lbl, h])
        _style_header(ws, 2)
        n = len(a9['labels'])
        pie = DoughnutChart()
        pie.title = title
        pie.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=1 + n), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
        pie.width, pie.height = 16, 12
        ws.add_chart(pie, f'A{n + 4}')

        ws2 = wb.create_sheet('Détail Escalade x Cause')
        ws2.append(['Escalade', 'Cause', 'Incidents', 'Outage', '% du total'])
        for r in a9['table']:
            ws2.append([r['escalade'], r['cause'], r['n'], r['outage'], r['pct']])
        _style_header(ws2, 5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(res: dict, source: str, generated_on: str) -> io.BytesIO:
    """Rapport PDF de synthèse (reportlab) — KPIs + tableaux principaux."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    NAVY   = colors.HexColor('#003087')
    YELLOW = colors.HexColor('#FFC72C')
    GREY   = colors.HexColor('#f0f4ff')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title='Analytics — Rapport d’analyses automatiques')
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Title'], textColor=NAVY, fontSize=17)
    h2 = ParagraphStyle('h2', parent=styles['Heading2'], textColor=NAVY, fontSize=12,
                        spaceBefore=10, spaceAfter=4)
    meta = ParagraphStyle('meta', parent=styles['Normal'], fontSize=8,
                          textColor=colors.HexColor('#555555'))
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7.5, leading=9)

    def _table(headers, rows, widths=None):
        data = [[Paragraph(f'<b>{h}</b>', cell) for h in headers]]
        for r in rows:
            data.append([Paragraph(str(v), cell) for v in r])
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GREY]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#c9d4ea')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]))
        # en-tête : texte blanc dans les Paragraph
        data[0] = [Paragraph(f'<font color="white"><b>{h}</b></font>', cell) for h in headers]
        return t

    story = [
        Paragraph('Analytics — Analyses automatiques des incidents', h1),
        Paragraph(f'Source : {source} &nbsp;&nbsp;|&nbsp;&nbsp; Généré le {generated_on}', meta),
        Spacer(1, 6),
    ]

    kpi = res.get('kpi', {})
    story.append(_table(
        ['Incidents', 'Indisponibilité', 'MTTR', 'Sites impactés', 'Régions', 'Cause dominante'],
        [[kpi.get('incidents', 0), kpi.get('outage', '—'), kpi.get('mttr', '—'),
          kpi.get('sites', 0), kpi.get('regions', 0), kpi.get('top_cause', '—')]]))

    story.append(Paragraph('Synthèse régionale (incidents & indisponibilité)', h2))
    story.append(_table(
        ['Région', 'Sites', 'Incidents', 'Outage', 'MTTR', 'Site le plus impacté'],
        [[r['region'], r['sites'], r['n'], r['outage'], r['mttr'], r['top_site']]
         for r in res['a4']]))

    story.append(Paragraph('Top 10 sites par indisponibilité', h2))
    story.append(_table(
        ['Site', 'Incidents', 'Outage', 'Cause principale'],
        [[r['site'], r['n'], r['outage'], r['cause']] for r in res['a2'][:10]]))

    story.append(Paragraph('Pareto des causes (80/20)', h2))
    story.append(_table(
        ['Cause', 'Outage (h)', 'Cumul (%)'],
        list(zip(res['a6']['labels'], res['a6']['outage_h'], res['a6']['cum_pct']))))

    story.append(Paragraph('Top 10 équipements les moins fiables', h2))
    story.append(_table(
        ['Équipement en défaut', 'Incidents', 'Outage'],
        [[r['equipement'], r['n'], r['outage']] for r in res['a7']]))

    story.append(Paragraph('Répartition des responsabilités (Escalade × Cause)', h2))
    story.append(_table(
        ['Escalade', 'Cause', 'Incidents', 'Outage', '% outage total'],
        [[r['escalade'], r['cause'], r['n'], r['outage'], f"{r['pct']} %"]
         for r in res['a9']['table'][:12]]))

    story.append(Paragraph('Performance opérationnelle par base', h2))
    story.append(_table(
        ['Base', 'Sites', 'Incidents', 'Outage', 'MTTR', 'Site le plus impacté'],
        [[r['base'], r['sites'], r['n'], r['outage'], r['mttr'], r['top_site']]
         for r in res['a10']]))

    story.append(Spacer(1, 8))
    story.append(Paragraph('Classification des sites (incidents & cause principale)', h2))
    story.append(_table(
        ['Classification', 'Sites', 'Incidents', 'Outage', 'Cause principale'],
        [[r['classification'], r['sites'], r['n'], r['outage'], r['top_cause']]
         for r in res['a5']]))

    doc.build(story)
    buf.seek(0)
    return buf
