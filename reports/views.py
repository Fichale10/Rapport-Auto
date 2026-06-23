import calendar
import json
import time  # noqa
import os
import math
from collections import defaultdict
from datetime import date, timedelta
from urllib import request

from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.safestring import mark_safe
from django.http import FileResponse, Http404, JsonResponse
from django.conf import settings
from django.contrib import messages

from .models import UploadedReport
from .forms import UploadForm

from treatement import process_file

ESC_MAPPING = {
    'ENERGIE':          'ENERGIE',
    'RAN-FIELD O':      'RAN',
    'RAN':              'RAN',
    'TRANS FH-FIELD O': 'TRANS FH',
    'TRANS FH':         'TRANS FH',
    'TRANS IP':         'TRANS IP',
}

NB_SITES = {
    'ENERGIE':  1227,
    'RAN':      1227,
    'TRANS FH': 996,
    'TRANS IP': 582,
}

REGION_TOTAL_SITES = [
    ('LOME',     'Lomé',      398),
    ('MARITIME', 'Maritime',  201),
    ('PLATEAUX', 'Plateaux',  169),
    ('CENTRALE', 'Centrale',  143),
    ('KARA',     'Kara',      175),
    ('SAVANES',  'Savanes',   127),
]

MOIS_FR_LONG = [
    '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
]
MOIS_FR_SHORT = [
    '', 'Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin',
    'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc',
]

SPARK_ESCALADES = [
    ('ENERGIE',         '#FFC72C'),
    ('TRANS FO',        '#e53e3e'),
    ('TRANS IP',        '#22c55e'),
    ('RAN-FIELD O',     '#8b5cf6'),
    ('TRANS FH-FIELD O','#f97316'),
]


def _filter_reports_by_month(queryset, year=None, month=None):
    today = date.today()
    year = year or today.year
    month = month or today.month
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    return queryset.filter(date_rapport__lte=month_end).filter(
        Q(date_fin__gte=month_start)
        | Q(date_fin__isnull=True, date_rapport__gte=month_start)
    )


def _filter_reports_by_period(queryset, period):
    today = date.today()
    if period == 'week':
        return queryset.filter(uploaded_at__date__gte=today - timedelta(days=6))
    if period == 'month':
        return queryset.filter(uploaded_at__date__gte=today - timedelta(days=29))
    if period == 'quarter':
        return queryset.filter(uploaded_at__date__gte=today - timedelta(days=89))
    if period == 'half':
        return queryset.filter(uploaded_at__date__gte=today - timedelta(days=179))
    if period == 'year':
        return queryset.filter(uploaded_at__date__gte=today - timedelta(days=364))
    return queryset  # 'all' ou 'custom'


def _shift_month(d, delta):
    m = d.month - 1 + delta
    y = d.year + m // 12
    m = m % 12 + 1
    return date(y, m, 1)


def _evol_time_buckets(period, today=None, custom_start=None, custom_end=None):
    """Intervalles fixes : 7 jours, 7 semaines, 3/6/12 mois, ou plage custom."""
    today = today or date.today()
    buckets = []

    if period == 'custom' and custom_start and custom_end:
        delta = (custom_end - custom_start).days
        if delta <= 7:
            for i in range(delta + 1):
                d = custom_start + timedelta(days=i)
                buckets.append({'label': d.strftime('%d/%m'), 'start': d, 'end': d})
        elif delta <= 90:
            d = custom_start
            while d <= custom_end:
                we = min(d + timedelta(days=6), custom_end)
                buckets.append({'label': f"S{d.isocalendar()[1]}", 'start': d, 'end': we})
                d += timedelta(weeks=1)
        else:
            d = custom_start.replace(day=1)
            while d <= custom_end:
                me = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])
                buckets.append({
                    'label': f"{MOIS_FR_SHORT[d.month]} {str(d.year)[2:]}",
                    'start': d,
                    'end': min(me, custom_end),
                })
                d = _shift_month(d, 1)
        return buckets

    if period == 'week':
        # 7 buckets journaliers (J-6 → aujourd'hui)
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            buckets.append({'label': d.strftime('%d/%m'), 'start': d, 'end': d})
    elif period == 'month':
        # 4 buckets hebdomadaires glissants (J-29 → aujourd'hui)
        for i in range(3, -1, -1):
            ws = today - timedelta(days=i * 7 + 6)
            we = today - timedelta(days=i * 7)
            buckets.append({'label': f"S{ws.isocalendar()[1]}", 'start': ws, 'end': we})
    elif period == 'quarter':
        # 3 buckets mensuels — les 3 mois qui couvrent J-89 → aujourd'hui
        start_approx = today - timedelta(days=89)
        base = start_approx.replace(day=1)
        for i in range(3):
            ms = _shift_month(base, i)
            me = date(ms.year, ms.month, calendar.monthrange(ms.year, ms.month)[1])
            buckets.append({
                'label': f"{MOIS_FR_SHORT[ms.month]} {ms.year}",
                'start': ms,
                'end': me,
            })
    elif period == 'half':
        # 6 buckets mensuels — les 6 mois qui couvrent J-179 → aujourd'hui
        start_approx = today - timedelta(days=179)
        base = start_approx.replace(day=1)
        for i in range(6):
            ms = _shift_month(base, i)
            me = date(ms.year, ms.month, calendar.monthrange(ms.year, ms.month)[1])
            buckets.append({
                'label': f"{MOIS_FR_SHORT[ms.month]} {str(ms.year)[2:]}",
                'start': ms,
                'end': me,
            })
    elif period == 'year':
        # 12 buckets mensuels — les 12 mois qui couvrent J-364 → aujourd'hui
        start_approx = today - timedelta(days=364)
        base = start_approx.replace(day=1)
        for i in range(12):
            ms = _shift_month(base, i)
            me = date(ms.year, ms.month, calendar.monthrange(ms.year, ms.month)[1])
            buckets.append({
                'label': f"{MOIS_FR_SHORT[ms.month]} {str(ms.year)[2:]}",
                'start': ms,
                'end': me,
            })
    else:
        # 'all' → 12 derniers mois glissants
        base = _shift_month(today.replace(day=1), -11)
        for i in range(12):
            ms = _shift_month(base, i)
            me = date(ms.year, ms.month, calendar.monthrange(ms.year, ms.month)[1])
            buckets.append({
                'label': f"{MOIS_FR_SHORT[ms.month]} {str(ms.year)[2:]}",
                'start': ms,
                'end': me,
            })

    return buckets


def _build_spark_evolution(reports_qs, period, custom_start=None, custom_end=None):
    all_reports = list(reports_qs)
    buckets = _evol_time_buckets(period, custom_start=custom_start, custom_end=custom_end)
    labels = []
    total_vals = []
    esc_vals = {esc: [] for esc, _ in SPARK_ESCALADES}
    range_reports = []

    for b in buckets:
        matching = [
            r for r in all_reports
            if b['start'] <= r.date_rapport <= b['end']
        ]
        range_reports.extend(matching)
        labels.append(b['label'])
        total_vals.append(sum(r.total_incidents for r in matching))
        for esc, _ in SPARK_ESCALADES:
            esc_vals[esc].append(sum(_inc_for_escalade(r, esc) for r in matching))

    spark_series = [{
        'name': 'Total',
        'color': '#003087',
        'is_total': True,
        'values': total_vals,
    }]
    for esc, color in SPARK_ESCALADES:
        spark_series.append({
            'name': esc,
            'color': color,
            'is_total': False,
            'values': esc_vals[esc],
        })

    period_label = ''
    if buckets:
        period_label = (
            f"{buckets[0]['start'].strftime('%d/%m/%Y')}"
            f" → {buckets[-1]['end'].strftime('%d/%m/%Y')}"
        )

    unique_range = {r.pk: r for r in range_reports}
    range_list = sorted(unique_range.values(), key=lambda r: r.date_rapport)
    evol_incidents = sum(r.total_incidents for r in unique_range.values())
    evol_unresolved = sum(r.unresolved_count or 0 for r in unique_range.values())
    evol_resolved = evol_incidents - evol_unresolved
    latest = range_list[-1] if range_list else None
    return labels, spark_series, period_label, evol_incidents, evol_resolved, evol_unresolved, latest, bool(labels)


def _inc_for_escalade(report, esc):
    for row in (report.synthesis_json or []):
        if row.get('Escalade') == esc:
            return int(row.get('Inc count', 0) or 0)
    return 0


def _period_label(report):
    start = report.date_rapport
    end = report.date_fin or start
    if report.period_type == UploadedReport.PERIOD_DAY:
        return start.strftime('%d/%m/%Y')
    return f"Du {start.strftime('%d/%m/%Y')} au {end.strftime('%d/%m/%Y')}"


def home(request):
    if request.user.is_superuser:
        all_reports = UploadedReport.objects.filter(processed=True).order_by('-uploaded_at')
    else:
        all_reports = UploadedReport.objects.filter(processed=True, user=request.user).order_by('-uploaded_at')

    total_reports    = all_reports.count()
    total_unresolved = sum(r.unresolved_count for r in all_reports if r.unresolved_count)
    total_outage_h = round(sum(r.total_duration_sec for r in all_reports) / 3600, 1)

    today = date.today()
    month_reports = list(_filter_reports_by_month(all_reports))
    month_reports_count = len(month_reports)
    month_incidents = sum(r.total_incidents for r in month_reports)
    month_unresolved = sum(r.unresolved_count or 0 for r in month_reports)
    month_resolved = month_incidents - month_unresolved
    month_label = f'{MOIS_FR_LONG[today.month]} {today.year}'

    prev_month = today.month - 1
    prev_year = today.year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1
    prev_month_incidents = sum(
        r.total_incidents for r in _filter_reports_by_month(all_reports, prev_year, prev_month)
    )
    month_trend_pct = None
    if prev_month_incidents > 0:
        month_trend_pct = round(
            (month_incidents - prev_month_incidents) / prev_month_incidents * 100, 1,
        )
    elif month_incidents > 0:
        month_trend_pct = 100.0

    # ── Période unifiée pour Évolution + Synthèse ──────────────────────────
    period = request.GET.get('period', 'week')
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    custom_start  = custom_end = None

    if date_from_str and date_to_str:
        try:
            from datetime import datetime as _dt
            dt_from = _dt.fromisoformat(date_from_str)
            dt_to   = _dt.fromisoformat(date_to_str)
            custom_start = dt_from.date()
            custom_end   = dt_to.date()
            period = 'custom'
        except (ValueError, TypeError):
            date_from_str = date_to_str = ''

    if period == 'custom' and custom_start and custom_end:
        base_qs = all_reports.filter(
            uploaded_at__date__gte=custom_start,
            uploaded_at__date__lte=custom_end,
        )
    else:
        base_qs = all_reports

    (
        spark_labels,
        spark_series,
        evol_period_label,
        evol_incidents,
        evol_resolved,
        evol_unresolved,
        evol_latest_report,
        show_spark_chart,
    ) = _build_spark_evolution(base_qs, period, custom_start, custom_end)

    last_report = all_reports.first()
    if not evol_latest_report:
        evol_latest_report = last_report

    # ── Synthèse par Escalade agrégée (même période) ───────────────────────
    synth_qs = _filter_reports_by_period(base_qs, period)  # 'custom' → return as-is

    esc_data = defaultdict(lambda: {
        'inc': 0, 'duree_sec': 0, 'outage_sec': 0,
        'mttr_sec': 0, 'unresolved': 0, 'has_data': False,
    })

    def _hms_to_sec(s):
        try:
            parts = str(s).split(':')
            if len(parts) == 3:
                return int(float(parts[0])) * 3600 + int(float(parts[1])) * 60 + int(float(parts[2]))
        except Exception:
            pass
        return 0

    def _sec_to_hms(sec):
        h = int(sec // 3600)
        m = int((sec % 3600) // 60)
        s = int(sec % 60)
        return f"{h}:{m:02d}:{s:02d}"

    for r in synth_qs:
        if not r.synthesis_json:
            continue
        for row in r.synthesis_json:
            esc = row.get('Escalade', '')
            if not esc or esc == 'TOTAL':
                continue
            inc = row.get('Inc count', 0) or 0
            esc_data[esc]['inc']        += inc
            esc_data[esc]['duree_sec']  += _hms_to_sec(row.get('DUREE', '0:00:00'))
            esc_data[esc]['outage_sec'] += _hms_to_sec(row.get('OUTAGE', '0:00:00'))
            esc_data[esc]['mttr_sec']   += _hms_to_sec(row.get('MTTR', '0:00:00'))
            if inc > 0:
                esc_data[esc]['has_data'] = True
            status = str(row.get('Status', ''))
            if 'Non resolu' in status:
                try:
                    esc_data[esc]['unresolved'] += int(status.split()[0])
                except Exception:
                    esc_data[esc]['unresolved'] += 1

    synth_rows = []
    total_inc = total_duree = total_outage = 0
    esc_order = [
        'ENERGIE', 'TRANS FH-FIELD O', 'RAN-FIELD O', 'ENERGIE / TRANS / RAN',
        'TRANS / RAN', 'INFRA', 'PROJET', 'TRANS FO', 'TRANS FTTM', 'TRANS IP',
        'ENVIRONNEMENT', 'BSS',
    ]
    all_escs = list(esc_order) + [e for e in esc_data if e not in esc_order]

    for esc in all_escs:
        if esc not in esc_data:
            continue
        d = esc_data[esc]
        inc = d['inc']
        unres = d['unresolved']
        if inc == 0:
            status = 'na'
        elif unres > 0:
            status = 'unresolved'
        else:
            status = 'resolved'

        synth_rows.append({
            'escalade':   esc,
            'inc_count':  inc if inc > 0 else 0,
            'duree':      _sec_to_hms(d['duree_sec']) if d['duree_sec'] else '0:00:00',
            'mttr':       _sec_to_hms(d['mttr_sec']) if d['mttr_sec'] else '0:00:00',
            'outage':     _sec_to_hms(d['outage_sec']) if d['outage_sec'] else '0:00:00',
            'status':     status,
            'unresolved': unres,
            'is_total':   False,
        })
        total_inc    += inc
        total_duree  += d['duree_sec']
        total_outage += d['outage_sec']

    if synth_rows:
        synth_rows.append({
            'escalade':   'TOTAL',
            'inc_count':  total_inc,
            'duree':      _sec_to_hms(total_duree),
            'mttr':       '—',
            'outage':     _sec_to_hms(total_outage),
            'status':     '',
            'unresolved': 0,
            'is_total':   True,
        })

    synth_total_inc = total_inc

    # ── Statut Sites par Région ────────────────────────────────────────────
    region_impacted_sets = defaultdict(set)
    for r in all_reports:
        for region, sites in (r.region_sites_json or {}).items():
            if isinstance(sites, list):
                region_impacted_sets[region].update(sites)

    statut_sites = []
    for key, label, total in REGION_TOTAL_SITES:
        impacted = len(region_impacted_sets.get(key, set()))
        pct = round(impacted / total * 100, 1) if total > 0 else 0
        statut_sites.append({
            'key':      key,
            'region':   label,
            'total':    total,
            'impacted': impacted,
            'pct':      pct,
            'is_total': False,
        })
    grand_total  = sum(t for _, _, t in REGION_TOTAL_SITES)
    grand_impact = sum(r['impacted'] for r in statut_sites)
    grand_pct    = round(grand_impact / grand_total * 100, 1) if grand_total > 0 else 0
    statut_sites.append({
        'region':   'TOTAL',
        'total':    grand_total,
        'impacted': grand_impact,
        'pct':      grand_pct,
        'is_total': True,
    })

    # ── Points de sites réels pour la carte (depuis Site, importés du KML) ──
    from .models import Site
    impacted_names = set()
    for sites in region_impacted_sets.values():
        impacted_names.update(sites)

    sites_geo = [
        {
            'name':   s.site_name,
            'lat':    s.latitude,
            'lon':    s.longitude,
            'region': s.region,
        }
        for s in Site.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
        if s.site_name in impacted_names
    ]

    return render(request, 'reports/home.html', {
        'period':           period,
        'date_from':        date_from_str,
        'date_to':          date_to_str,
        'synth_rows':       synth_rows,
        'synth_total_inc':  synth_total_inc,
        'total_reports':        total_reports,
        'total_unresolved':     total_unresolved,
        'total_outage_h':       total_outage_h,
        'month_label':          month_label,
        'month_reports_count':  month_reports_count,
        'month_incidents':      month_incidents,
        'month_resolved':       month_resolved,
        'month_unresolved':     month_unresolved,
        'month_trend_pct':      month_trend_pct,
        'evol_period':          period,
        'show_spark_chart':     show_spark_chart,
        'spark_labels':         mark_safe(json.dumps(spark_labels)),
        'spark_series':         mark_safe(json.dumps(spark_series)),
        'evol_incidents':       evol_incidents,
        'evol_resolved':        evol_resolved,
        'evol_unresolved':      evol_unresolved,
        'evol_period_label':    evol_period_label,
        'evol_latest_report':   evol_latest_report,
        'last_report':          last_report,
        'statut_sites':         statut_sites,
        'sites_geo_json':       mark_safe(json.dumps(sites_geo)),
    })


def upload(request):
    from datetime import timedelta
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            report = form.save(commit=False)
            report.original_filename = request.FILES['file'].name
            report.user = request.user
            report.save()
            return redirect('process_report', pk=report.pk)
    else:
        form = UploadForm()
    today      = date.today()
    yesterday  = today - timedelta(days=1)
    active_tab = request.GET.get('tab', 'excel')
    return render(request, 'reports/upload.html', {
        'form':       form,
        'today':      today,
        'yesterday':  yesterday,
        'active_tab': active_tab,
    })


def process_report(request, pk):
    report = get_object_or_404(UploadedReport, pk=pk)

    if report.processed:
        return redirect('results', pk=report.pk)

    if request.method == 'GET':
        return render(request, 'reports/processing.html', {'report': report})

    start = time.time()
    input_path = report.file.path
    date_debut_str = report.date_rapport.strftime('%Y-%m-%d')
    date_fin_str = report.date_fin.strftime('%Y-%m-%d') if report.date_fin else None

    results_dir = os.path.join(settings.MEDIA_ROOT, 'results')
    os.makedirs(results_dir, exist_ok=True)
    output_path = os.path.join(results_dir, f'{report.pk}_detailed.xlsx')

    try:
        df_export, df_dedup, df_synthese = process_file(
            input_path, date_debut_str, date_fin=date_fin_str,
        )
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    df_export.to_excel(output_path, index=False)
    synthesis_path = output_path.replace('.xlsx', '_Synthese.xlsx')
    df_synthese.to_excel(synthesis_path, index=False)

    import pandas as pd
    df_original = pd.read_excel(input_path)
    report.total_rows = len(df_original)
    report.filtered_rows = len(df_export)

    try:
        total_row = df_synthese[df_synthese['Escalade'] == 'TOTAL']
        report.total_incidents = int(total_row['Inc count'].values[0]) if len(total_row) > 0 else 0
    except (IndexError, KeyError):
        report.total_incidents = 0

    unresolved = 0
    for _, row in df_synthese.iterrows():
        status = str(row.get('Status', ''))
        if 'Non resolu' in status:
            try:
                unresolved += int(status.split()[0])
            except (ValueError, IndexError):
                pass
    report.unresolved_count = unresolved

    def _parse_duration(x):
        try:
            parts = str(x).split(':')
            if len(parts) == 3:
                return int(float(parts[0])) * 3600 + int(float(parts[1])) * 60 + int(float(parts[2]))
        except (ValueError, AttributeError):
            pass
        return 0

    report.total_duration_sec = float(
        df_export['Duration'].astype(str).apply(_parse_duration).sum()
        if 'Duration' in df_export.columns and len(df_export) > 0 else 0
    )
    report.processing_time_sec = round(time.time() - start, 2)

    import json
    import numpy as np

    class _NpEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, np.integer): return int(obj)
            if isinstance(obj, np.floating): return float(obj)
            return super().default(obj)

    report.synthesis_json = json.loads(
        json.dumps(df_synthese.to_dict('records'), cls=_NpEncoder)
    )

    site_col = next((c for c in ('Site Name', 'Site name', 'SITE NAME') if c in df_dedup.columns), None)
    if site_col and len(df_dedup) > 0:
        top_sites = (
            df_dedup[site_col].dropna().astype(str)
            .value_counts().head(10).reset_index()
        )
        top_sites.columns = ['name', 'count']
        report.top_sites_json = json.loads(json.dumps(top_sites.to_dict('records'), cls=_NpEncoder))
    else:
        report.top_sites_json = []

    region_col = next((c for c in ('Région', 'Region', 'REGION', 'region') if c in df_dedup.columns), None)
    if region_col and site_col and len(df_dedup) > 0:
        region_sites = {}
        for region, grp in df_dedup.groupby(region_col):
            sites = grp[site_col].dropna().astype(str).unique().tolist()
            region_sites[str(region).strip()] = sites
        report.region_sites_json = region_sites
    else:
        report.region_sites_json = {}

    cause_col = next((c for c in ('Root Cause', 'Cause') if c in df_export.columns), None)
    if cause_col and 'Duration' in df_export.columns:
        cause_duration = defaultdict(float)
        for _, row in df_export.iterrows():
            cause = str(row.get(cause_col, '')).strip()
            dur   = _parse_duration(str(row.get('Duration', '')))
            if cause and cause != 'nan' and dur > 0:
                cause_duration[cause] += dur
        top_causes = sorted(cause_duration.items(), key=lambda x: x[1], reverse=True)[:10]
        report.top_causes_json = json.loads(json.dumps([
            {'name': k, 'duration_sec': v} for k, v in top_causes
        ], cls=_NpEncoder))
    else:
        report.top_causes_json = []

    report.detailed_file.name = os.path.relpath(output_path, settings.MEDIA_ROOT)
    if os.path.exists(synthesis_path):
        report.synthesis_file.name = os.path.relpath(synthesis_path, settings.MEDIA_ROOT)

    report.processed = True

    outage_j = defaultdict(lambda: defaultdict(float))

    if 'Alarm Time' in df_export.columns and 'Escalade' in df_export.columns and 'Duration' in df_export.columns:
        for _, row in df_export.iterrows():
            alarm_time = row.get('Alarm Time')
            escalade   = str(row.get('Escalade', '')).strip()
            duration   = row.get('Duration', '')

            esc_key = ESC_MAPPING.get(escalade)
            if not esc_key:
                continue

            try:
                if hasattr(alarm_time, 'date'):
                    day = alarm_time.date().isoformat()
                else:
                    day = pd.to_datetime(alarm_time).date().isoformat()
            except Exception:
                continue

            dur_sec = _parse_duration(str(duration))
            if dur_sec > 0:
                outage_j[esc_key][day] += dur_sec

    report.outage_journalier_json = {
        esc: dict(jours)
        for esc, jours in outage_j.items()
    }
    report.save()

    return JsonResponse({
        'done': True,
        'redirect': f'/results/{report.pk}/',
        'total_incidents': report.total_incidents,
        'processing_time': report.processing_time_sec,
    })


def process_status(request, pk):
    report = get_object_or_404(UploadedReport, pk=pk)
    if report.processed:
        return JsonResponse({'done': True, 'redirect': f'/results/{report.pk}/'})
    return JsonResponse({'done': False})


def delete_report(request, pk):
    report = get_object_or_404(UploadedReport, pk=pk)

    if report.user != request.user and not request.user.is_superuser:
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    if request.method == 'POST':
        files_to_delete = []
        if report.file:
            files_to_delete.append(report.file.path)
        if report.detailed_file:
            try: files_to_delete.append(report.detailed_file.path)
            except Exception: pass
        if report.synthesis_file:
            try: files_to_delete.append(report.synthesis_file.path)
            except Exception: pass

        for filepath in files_to_delete:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except Exception:
                pass

        report.delete()
        return JsonResponse({'success': True, 'redirect': '/history/'})

    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)


def results(request, pk):
    report = get_object_or_404(UploadedReport, pk=pk, processed=True)
    if report.user != request.user and not request.user.is_superuser:
        return redirect('history')
    return render(request, 'reports/results.html', {'report': report, 'period_label': _period_label(report)})


def download_file(request, pk, file_type):
    report = get_object_or_404(UploadedReport, pk=pk, processed=True)
    if report.user != request.user and not request.user.is_superuser:
        raise Http404("Non autorisé")
    safe_label = _period_label(report).replace('/', '-').replace(' ', '_')
    if file_type == 'detailed' and report.detailed_file:
        file_field = report.detailed_file
        filename = f"Rapport_Detail_{safe_label}.xlsx"
    elif file_type == 'synthesis' and report.synthesis_file:
        file_field = report.synthesis_file
        filename = f"Rapport_Synthese_{safe_label}.xlsx"
    else:
        raise Http404("File not found")
    return FileResponse(open(file_field.path, 'rb'), as_attachment=True, filename=filename)


def history(request):
    from datetime import timedelta

    if request.user.is_superuser:
        all_reports = UploadedReport.objects.filter(processed=True).order_by('-date_rapport')
    else:
        all_reports = UploadedReport.objects.filter(processed=True, user=request.user).order_by('-date_rapport')

    period_filter = request.GET.get('period', 'all')

    if period_filter == 'day':
        filtered_reports = list(all_reports.filter(date_fin__isnull=True))
    elif period_filter == 'week':
        filtered_reports = [
            r for r in all_reports
            if r.date_fin and 5 <= (r.date_fin - r.date_rapport).days <= 10
        ]
    elif period_filter == 'month':
        filtered_reports = [
            r for r in all_reports
            if r.date_fin and 20 <= (r.date_fin - r.date_rapport).days <= 45
        ]
    elif period_filter == 'year':
        filtered_reports = [
            r for r in all_reports
            if r.date_fin and (r.date_fin - r.date_rapport).days > 45
        ]
    else:
        filtered_reports = list(all_reports)

    total_reports    = len(filtered_reports)
    total_incidents  = sum(r.total_incidents for r in filtered_reports)
    total_unresolved = sum(r.unresolved_count for r in filtered_reports if r.unresolved_count)
    total_resolved   = sum(1 for r in filtered_reports if r.unresolved_count == 0)

    return render(request, 'reports/history.html', {
        'reports':          filtered_reports,
        'period_filter':    period_filter,
        'total_reports':    total_reports,
        'total_incidents':  total_incidents,
        'total_unresolved': total_unresolved,
        'total_resolved':   total_resolved,
    })


# ── Palette couleurs ──────────────────────────────────────────────────────────
DONUT_COLORS = [
    '#003087', '#e05a2b', '#FF9800', '#2196F3', '#FFC72C',
    '#4CAF50', '#9C27B0', '#00BCD4', '#8BC34A', '#FF5722',
    '#607D8B', '#795548', '#009688', '#0047cc', '#F44336',
]
DONUT_DARK = [
    '#001245', '#8b3318', '#b36a00', '#0d5ca8', '#a07800',
    '#2e7d32', '#6a1b9a', '#00838f', '#558b2f', '#bf360c',
    '#37474f', '#4e342e', '#00695c', '#002fa0', '#b71c1c',
]


def _make_donut_svg(data, total_h):
    if not data:
        return ''
    total = sum(d['outage_h'] for d in data)
    if total == 0:
        return ''

    CX, CY  = 320, 175
    RX, RY  = 230, 90
    DEPTH   = 58
    W       = 680
    VIEW_PAD = 120
    FULL_W  = W + VIEW_PAD * 2

    n        = len(data)
    COLS     = 3
    LEG_ROWS = math.ceil(n / COLS)
    H        = CY + RY + DEPTH + 44 + LEG_ROWS * 52 + 20

    def pt(a, r=1.0):
        return (CX + r * RX * math.cos(a), CY + r * RY * math.sin(a))

    defs = '<defs>'
    for i, d in enumerate(data):
        c  = DONUT_COLORS[i % len(DONUT_COLORS)]
        dk = DONUT_DARK[i  % len(DONUT_DARK)]
        defs += (
            f'<linearGradient id="pg{i}" x1="0" y1="0" x2="0.2" y2="1">'
            f'<stop offset="0%" stop-color="{c}"/>'
            f'<stop offset="100%" stop-color="{dk}"/>'
            f'</linearGradient>'
        )
    defs += (
        '<filter id="pshadow" x="-30%" y="-30%" width="160%" height="180%">'
        '<feDropShadow dx="0" dy="8" stdDeviation="10" flood-color="#00000022"/>'
        '</filter>'
        '</defs>'
    )

    slices = []
    angle  = -math.pi / 2
    for i, d in enumerate(data):
        sweep = (d['outage_h'] / total) * 2 * math.pi
        slices.append({
            'a1':    angle,
            'a2':    angle + sweep,
            'mid':   angle + sweep / 2,
            'color': DONUT_COLORS[i % len(DONUT_COLORS)],
            'dark':  DONUT_DARK[i  % len(DONUT_DARK)],
            'grad':  f'url(#pg{i})',
            'pct':   d['pct'],
            'h':     d['outage_h'],
            'name':  d['name'],
        })
        angle += sweep

    sides = ''
    for s in reversed(slices):
        a1, a2 = s['a1'], s['a2']
        vs = max(a1, 0)
        ve = min(a2, math.pi)
        if vs < ve:
            N = 40
            tp, bp = [], []
            for k in range(N + 1):
                a = vs + (ve - vs) * k / N
                x = CX + RX * math.cos(a)
                y = CY + RY * math.sin(a)
                tp.append(f'{x:.2f},{y:.2f}')
                bp.append(f'{x:.2f},{y + DEPTH:.2f}')
            bp.reverse()
            sides += (
                f'<path d="M{" L".join(tp + bp)}Z" '
                f'fill="{s["dark"]}" stroke="rgba(255,255,255,0.15)" stroke-width="0.8"/>'
            )
        for a in [a1, a2]:
            if 0 <= a <= math.pi:
                ox, oy = pt(a)
                sides += (
                    f'<path d="M{CX:.2f},{CY:.2f} L{ox:.2f},{oy:.2f} '
                    f'L{ox:.2f},{oy+DEPTH:.2f} L{CX:.2f},{CY+DEPTH:.2f} Z" '
                    f'fill="{s["dark"]}" opacity="0.40" '
                    f'stroke="rgba(255,255,255,0.08)" stroke-width="0.6"/>'
                )

    bot_ellipse = (
        f'<ellipse cx="{CX}" cy="{CY+DEPTH}" rx="{RX}" ry="{RY}" '
        f'fill="none" stroke="rgba(0,0,0,0.05)" stroke-width="1"/>'
    )

    tops = ''
    for s in slices:
        a1, a2 = s['a1'], s['a2']
        x1, y1 = pt(a1)
        x2, y2 = pt(a2)
        large  = 1 if (a2 - a1) > math.pi else 0
        tops += (
            f'<path d="M{CX:.2f},{CY:.2f} L{x1:.2f},{y1:.2f} '
            f'A{RX},{RY} 0 {large},1 {x2:.2f},{y2:.2f} Z" '
            f'fill="{s["grad"]}" stroke="rgba(255,255,255,0.25)" stroke-width="1.2"/>'
        )

    highlights = ''
    for s in slices:
        hs = max(s['a1'], -math.pi / 2)
        he = min(s['a2'], -math.pi / 2 + 0.85)
        if hs < he:
            hx1, hy1 = pt(hs, 0.96)
            hx2, hy2 = pt(he, 0.96)
            highlights += (
                f'<path d="M{hx1:.2f},{hy1:.2f} '
                f'A{RX*0.96:.1f},{RY*0.96:.1f} 0 0,1 {hx2:.2f},{hy2:.2f}" '
                f'fill="none" stroke="rgba(255,255,255,0.35)" '
                f'stroke-width="1.6" stroke-linecap="round"/>'
            )

    INNER_THRESHOLD = 8
    labels = ''
    EXT_R  = 1.42
    LINE_R = 1.12

    ext_items = []
    for s in slices:
        if s['pct'] < INNER_THRESHOLD:
            tx, ty = pt(s['mid'], EXT_R)
            anchor = 'start' if math.cos(s['mid']) >= 0 else 'end'
            offset = 22 if anchor == 'start' else -22
            ext_items.append({'s': s, 'tx': tx + offset, 'ty': ty, 'anchor': anchor})

    MIN_GAP   = 26
    left_ext  = sorted([e for e in ext_items if e['anchor'] == 'end'],   key=lambda e: e['ty'])
    right_ext = sorted([e for e in ext_items if e['anchor'] == 'start'], key=lambda e: e['ty'])

    for grp in [left_ext, right_ext]:
        for k in range(1, len(grp)):
            if grp[k]['ty'] - grp[k-1]['ty'] < MIN_GAP:
                grp[k]['ty'] = grp[k-1]['ty'] + MIN_GAP

    ext_items = left_ext + right_ext

    for s in slices:
        if s['pct'] >= INNER_THRESHOLD:
            lx, ly = pt(s['mid'], 0.56)
            name_short = s['name'][:12] + ('…' if len(s['name']) > 12 else '')
            labels += (
                f'<rect x="{lx-38:.1f}" y="{ly-20:.1f}" width="76" height="38" rx="8" '
                f'fill="rgba(0,0,0,0.35)"/>'
                f'<text x="{lx:.1f}" y="{ly-4:.1f}" text-anchor="middle" '
                f'font-family="Arial,sans-serif" font-size="10" font-weight="700" '
                f'fill="rgba(255,255,255,0.90)">{name_short}</text>'
                f'<text x="{lx:.1f}" y="{ly+11:.1f}" text-anchor="middle" '
                f'font-family="Arial,sans-serif" font-size="15" font-weight="900" '
                f'fill="white">{s["pct"]}%</text>'
            )

    for e in ext_items:
        s      = e['s']
        ox, oy = pt(s['mid'], LINE_R)
        px, py = pt(s['mid'], 1.00)
        tx, ty = e['tx'], e['ty']
        anchor = e['anchor']
        name_short = s['name'][:14] + ('…' if len(s['name']) > 14 else '')
        labels += (
            f'<line x1="{px:.1f}" y1="{py:.1f}" x2="{tx:.1f}" y2="{ty:.1f}" '
            f'stroke="{s["color"]}" stroke-width="1.5" opacity="0.9"/>'
            f'<circle cx="{px:.1f}" cy="{py:.1f}" r="2" fill="{s["color"]}"/>'
            f'<text x="{tx:.1f}" y="{ty-4:.1f}" text-anchor="{anchor}" '
            f'font-family="Arial,sans-serif" font-size="13" font-weight="700" '
            f'fill="{s["dark"]}">{name_short}</text>'
            f'<text x="{tx:.1f}" y="{ty+11:.1f}" text-anchor="{anchor}" '
            f'font-family="Arial,sans-serif" font-size="12" font-weight="600" '
            f'fill="#4b5563">{s["pct"]}% · {s["h"]}h</text>'
        )

    badge = (
        f'<rect x="{CX-40}" y="{CY-18}" width="80" height="34" rx="9" '
        f'fill="white" opacity="0.95" stroke="rgba(0,48,135,0.14)" stroke-width="1"/>'
        f'<text x="{CX}" y="{CY}" text-anchor="middle" '
        f'font-family="Arial,sans-serif" font-size="17" font-weight="900" fill="#003087">'
        f'{total_h}h</text>'
        f'<text x="{CX}" y="{CY+13}" text-anchor="middle" '
        f'font-family="Arial,sans-serif" font-size="12" font-weight="700" '
        f'fill="#9ca3af" letter-spacing="1.5">TOTAL</text>'
    )

    leg_top = CY + RY + DEPTH + 44
    COL_W   = int((W - 40) / COLS)
    legend  = ''
    for i, s in enumerate(slices):
        col  = i % COLS
        row  = i // COLS
        x    = 20 + col * COL_W - VIEW_PAD
        y    = leg_top + row * 52
        name = s['name'][:18] + ('…' if len(s['name']) > 18 else '')
        legend += (
            f'<rect x="{x}" y="{y}" width="{COL_W-12}" height="42" rx="8" '
            f'fill="{s["color"]}" opacity="0.09" '
            f'stroke="{s["color"]}" stroke-opacity="0.20"/>'
            f'<circle cx="{x+13}" cy="{y+12}" r="6.5" fill="{s["color"]}"/>'
            f'<text x="{x+25}" y="{y+16}" '
            f'font-family="Arial,sans-serif" font-size="12.5" font-weight="800" '
            f'fill="{s["dark"]}">{name}</text>'
            f'<text x="{x+13}" y="{y+32}" '
            f'font-family="Arial,sans-serif" font-size="11.5" font-weight="600" '
            f'fill="#4b5563">{s["h"]}h · {s["pct"]}%</text>'
        )

    return (
        f'<svg width="100%" '
        f'viewBox="-{VIEW_PAD} 0 {FULL_W} {H}" '
        f'xmlns="http://www.w3.org/2000/svg" overflow="visible">'
        f'{defs}'
        f'<g filter="url(#pshadow)">{sides}{bot_ellipse}{tops}</g>'
        f'{highlights}{labels}{badge}{legend}'
        f'</svg>'
    )


def export_pdf(request, pk):
    import datetime
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO

    report = get_object_or_404(UploadedReport, pk=pk, processed=True)
    if report.user != request.user and not request.user.is_superuser:
        raise Http404("Non autorisé")

    YAS_BLUE   = colors.HexColor('#003087')
    YAS_YELLOW = colors.HexColor('#FFC72C')
    LIGHT_BLUE = colors.HexColor('#e8f0ff')
    LIGHT_GRAY = colors.HexColor('#f8faff')

    buffer = BytesIO()
    W = A4[0] - 40*mm

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=15*mm, bottomMargin=20*mm,
    )
    styles = getSampleStyleSheet()
    elements = []
    now = datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')

    def p(text, size=9, color='#333333', bold=False, align='LEFT'):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        return Paragraph(
            f'<font size="{size}" color="{color}">{text}</font>',
            ParagraphStyle('_', fontName=fn, alignment={'LEFT':0,'CENTER':1,'RIGHT':2}[align])
        )

    ht = Table([[
        p('<b>YAS</b>', size=24, color='#003087'),
        Table([
            [p('● RAPPORT D\'INCIDENTS RÉSEAU', size=8, color='#FFC72C')],
            [p('<b>ReportXCare</b>', size=14, color='#003087')],
            [p(f'📅 {_period_label(report)}', size=9, color="#888888")],
        ], colWidths=[W*0.55]),
    ]], colWidths=[W*0.35, W*0.65])
    ht.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.white),
        ('BOX',        (0,0), (-1,-1), 0.5, colors.HexColor('#e8edf5')),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',    (0,0), (-1,-1), 8),
        ('LINEBELOW',  (0,0), (-1,-1), 3, YAS_YELLOW),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 5*mm))

    fname = report.original_filename
    if len(fname) > 40: fname = fname[:37] + '...'
    mt = Table([[
        Table([[p('FICHIER SOURCE', size=7, color='#aaaaaa')],[p(f'<b>{fname}</b>', size=9, color='#003087')]], colWidths=[W*0.33]),
        Table([[p('GÉNÉRÉ PAR', size=7, color='#aaaaaa')],[p(f'<b>{request.user.get_full_name() or request.user.username}</b>', size=9, color='#003087')]], colWidths=[W*0.25]),
        Table([[p('DATE', size=7, color='#aaaaaa')],[p(f'<b>{now}</b>', size=9, color='#003087')]], colWidths=[W*0.28]),
        Table([[p('DURÉE', size=7, color='#aaaaaa')],[p(f'<b>{report.processing_time_sec}s</b>', size=9, color='#003087')]], colWidths=[W*0.14]),
    ]], colWidths=[W*0.33, W*0.25, W*0.28, W*0.14])
    mt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT_GRAY),
        ('BOX',        (0,0), (-1,-1), 0.5, colors.HexColor('#e8edf5')),
        ('INNERGRID',  (0,0), (-1,-1), 0.5, colors.HexColor('#e8edf5')),
        ('PADDING',    (0,0), (-1,-1), 6),
        ('VALIGN',     (0,0), (-1,-1), 'TOP'),
    ]))
    elements.append(mt)
    elements.append(Spacer(1, 5*mm))

    elements.append(p('RÉSUMÉ DU TRAITEMENT', size=9, color='#003087', bold=True))
    elements.append(HRFlowable(width='100%', thickness=2, color=YAS_YELLOW, spaceAfter=4))
    elements.append(Spacer(1, 2*mm))

    unresolved_color = '#e53e3e' if report.unresolved_count > 0 else '#22c55e'
    kpi_w = W / 4
    kt = Table([[
        Table([[p(str(report.total_rows), size=20, color='#003087', bold=True)],[p('LIGNES FICHIER', size=7, color='#888888')]]),
        Table([[p(str(report.filtered_rows), size=20, color='#003087', bold=True)],[p('ALARMES FILTRÉES', size=7, color='#888888')]]),
        Table([[p(str(report.total_incidents), size=20, color='#FFC72C', bold=True)],[p('INCIDENTS', size=7, color='#ffffff')]]),
        Table([[p(str(report.unresolved_count), size=20, color=unresolved_color, bold=True)],[p('NON RÉSOLUS', size=7, color='#888888')]]),
    ]], colWidths=[kpi_w]*4)
    kt.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (1,0), LIGHT_GRAY),
        ('BACKGROUND',  (2,0), (2,0), YAS_BLUE),
        ('BACKGROUND',  (3,0), (3,0), LIGHT_GRAY),
        ('BOX',         (0,0), (-1,-1), 0.5, colors.HexColor('#e8edf5')),
        ('INNERGRID',   (0,0), (-1,-1), 0.5, colors.HexColor('#e8edf5')),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',     (0,0), (-1,-1), 10),
    ]))
    elements.append(kt)
    elements.append(Spacer(1, 5*mm))

    if report.synthesis_json:
        elements.append(p('SYNTHÈSE PAR ESCALADE', size=9, color='#003087', bold=True))
        elements.append(HRFlowable(width='100%', thickness=2, color=YAS_YELLOW, spaceAfter=4))
        elements.append(Spacer(1, 2*mm))

        cw = [W*0.30, W*0.12, W*0.16, W*0.16, W*0.16, W*0.10]

        def cell(txt, size=9, color='#333333', bold=False, align='CENTER'):
            fn = 'Helvetica-Bold' if bold else 'Helvetica'
            return Paragraph(
                f'<font size="{size}" color="{color}">{txt}</font>',
                ParagraphStyle('_', fontName=fn, alignment={'LEFT':0,'CENTER':1,'RIGHT':2}[align])
            )

        header_row = [
            cell('Escalade',  size=8, color='#ffffff', bold=True, align='LEFT'),
            cell('Incidents', size=8, color='#ffffff', bold=True),
            cell('Durée',     size=8, color='#ffffff', bold=True),
            cell('MTTR',      size=8, color='#ffffff', bold=True),
            cell('Outage',    size=8, color='#ffffff', bold=True),
            cell('Statut',    size=8, color='#ffffff', bold=True),
        ]

        data_rows = [header_row]
        total_idx = None

        for i, row in enumerate(report.synthesis_json):
            esc = row.get('Escalade', '')
            inc = row.get('Inc count', 0)
            is_total = (esc == 'TOTAL')
            if is_total:
                total_idx = i + 1

            status = str(row.get('Status', ''))
            if 'Non resolu' in status:
                status_txt = f'⚠ {status}'
                s_color = '#7a5a00'
            elif status == 'Résolu':
                status_txt = '✓ Résolu'
                s_color = '#166534'
            elif status == 'N/A':
                status_txt = 'N/A'
                s_color = '#aaaaaa'
            else:
                status_txt = status
                s_color = '#333333'

            esc_color = '#003087' if is_total else '#111111'
            data_rows.append([
                cell(esc,                    size=9, color=esc_color, bold=is_total, align='LEFT'),
                cell(str(inc),               size=9, color=esc_color, bold=is_total),
                cell(row.get('DUREE',''),    size=8, color='#555555'),
                cell(row.get('MTTR',''),     size=8, color='#555555'),
                cell(row.get('OUTAGE',''),   size=9, color='#003087', bold=is_total),
                cell(status_txt,             size=7, color=s_color),
            ])

        st = Table(data_rows, colWidths=cw, repeatRows=1)
        ts = TableStyle([
            ('BACKGROUND',     (0,0), (-1,0),  YAS_BLUE),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_GRAY]),
            ('GRID',           (0,0), (-1,-1), 0.3, colors.HexColor('#e8edf5')),
            ('ALIGN',          (1,0), (-1,-1), 'CENTER'),
            ('ALIGN',          (0,0), (0,-1),  'LEFT'),
            ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING',        (0,0), (-1,-1), 5),
            ('LEFTPADDING',    (0,0), (0,-1),  8),
        ])
        if total_idx:
            ts.add('BACKGROUND', (0, total_idx), (-1, total_idx), LIGHT_BLUE)
            ts.add('LINEABOVE',  (0, total_idx), (-1, total_idx), 1.5, YAS_BLUE)
        st.setStyle(ts)
        elements.append(st)
        elements.append(Spacer(1, 6*mm))

    ft = Table([[
        p('YAS Togo — Rapport généré automatiquement par le système NOC', size=8, color='#ffffff'),
        p(f'ISOC • Confidentiel • {now}', size=8, color='#ffffff', align='RIGHT'),
    ]], colWidths=[W*0.60, W*0.40])
    ft.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), YAS_BLUE),
        ('PADDING',    (0,0), (-1,-1), 8),
    ]))
    elements.append(ft)

    doc.build(elements)
    buffer.seek(0)

    filename = f"Rapport_{report.original_filename.replace('.xlsx','')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def notifications(request):
    if request.user.is_superuser:
        reports = UploadedReport.objects.filter(processed=True, unresolved_count__gt=0).order_by('-date_rapport')
    else:
        reports = UploadedReport.objects.filter(processed=True, user=request.user, unresolved_count__gt=0).order_by('-date_rapport')

    data = [{
        'pk':       str(r.pk),
        'filename': r.original_filename[:40],
        'date':     r.date_rapport.strftime('%d/%m/%Y'),
        'count':    r.unresolved_count,
    } for r in reports[:10]]

    return JsonResponse({
        'total': reports.count(),
        'items': data,
    })


def register_view(request):
    if request.user.is_authenticated:
        return redirect('/')

    if request.method == 'POST':
        from django.contrib.auth.models import User
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password1  = request.POST.get('password1', '')
        password2  = request.POST.get('password2', '')

        if not password1:
            messages.error(request, 'Le mot de passe est obligatoire.')
        elif not username:
            messages.error(request, "L'identifiant est obligatoire.")
        elif User.objects.filter(username__iexact=username).exists():
            messages.error(request, 'Cet identifiant est déjà utilisé.')
        elif password1 != password2:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
        else:
            from django.contrib.auth.password_validation import validate_password
            from django.core.exceptions import ValidationError
            try:
                validate_password(password1)
            except ValidationError as e:
                for err in e.messages:
                    messages.error(request, err)
                return render(request, 'accounts/register.html')
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name,
            )
            user.is_active = False
            user.save()
            messages.success(request, 'Inscription envoyée ! Votre compte sera activé par un administrateur.')
            return redirect('accounts:login')

    return render(request, 'accounts/register.html')


def comparer(request):
    if request.user.is_superuser:
        all_reports = UploadedReport.objects.filter(processed=True).order_by('-uploaded_at')
    else:
        all_reports = UploadedReport.objects.filter(processed=True, user=request.user).order_by('-uploaded_at')

    pk1 = request.GET.get('r1')
    pk2 = request.GET.get('r2')

    r1 = r2 = None
    diff = None

    if pk1 and pk2:
        r1 = get_object_or_404(all_reports, pk=pk1)
        r2 = get_object_or_404(all_reports, pk=pk2)

        def parse_hms(s):
            try:
                parts = str(s).split(':')
                if len(parts) == 3:
                    return int(float(parts[0])) * 3600 + int(float(parts[1])) * 60 + int(float(parts[2]))
            except Exception:
                pass
            return 0

        def kpi_diff(v1, v2):
            if v1 == 0 and v2 == 0:
                return 0, 'neutral'
            if v1 == 0:
                return '+∞', 'worse'
            delta = v2 - v1
            pct   = round(delta / v1 * 100, 1)
            trend = 'better' if delta < 0 else ('worse' if delta > 0 else 'neutral')
            return (f'+{pct}%' if pct > 0 else f'{pct}%'), trend

        inc_delta,  inc_trend  = kpi_diff(r1.total_incidents,  r2.total_incidents)
        unr_delta,  unr_trend  = kpi_diff(r1.unresolved_count, r2.unresolved_count)
        rows_delta, rows_trend = kpi_diff(r1.total_rows,       r2.total_rows)
        filt_delta, filt_trend = kpi_diff(r1.filtered_rows,    r2.filtered_rows)

        def synth_map(report):
            m = {}
            for row in (report.synthesis_json or []):
                esc = row.get('Escalade', '')
                if esc and esc != 'TOTAL':
                    m[esc] = {
                        'count':      row.get('Inc count', 0),
                        'outage_sec': parse_hms(row.get('OUTAGE', '0:00:00')),
                        'status':     row.get('Status', ''),
                    }
            return m

        s1, s2   = synth_map(r1), synth_map(r2)
        all_escs = sorted(set(list(s1.keys()) + list(s2.keys())))

        esc_rows = []
        for esc in all_escs:
            v1 = s1.get(esc, {'count': 0, 'outage_sec': 0, 'status': '—'})
            v2 = s2.get(esc, {'count': 0, 'outage_sec': 0, 'status': '—'})
            delta = v2['count'] - v1['count']
            esc_rows.append({
                'name':       esc,
                'count1':     v1['count'],
                'count2':     v2['count'],
                'outage1_h':  round(v1['outage_sec'] / 3600, 1),
                'outage2_h':  round(v2['outage_sec'] / 3600, 1),
                'status1':    v1['status'],
                'status2':    v2['status'],
                'delta':      delta,
                'trend':      'better' if delta < 0 else ('worse' if delta > 0 else 'neutral'),
            })

        def sites_map(report):
            return {s['name']: s['count'] for s in (report.top_sites_json or [])}

        ts1, ts2  = sites_map(r1), sites_map(r2)
        all_sites = sorted(set(list(ts1.keys()) + list(ts2.keys())),
                           key=lambda s: max(ts1.get(s, 0), ts2.get(s, 0)), reverse=True)[:10]

        site_rows = [{
            'name':   s,
            'count1': ts1.get(s, 0),
            'count2': ts2.get(s, 0),
            'delta':  ts2.get(s, 0) - ts1.get(s, 0),
        } for s in all_sites]

        diff = {
            'inc_delta':  inc_delta,  'inc_trend':  inc_trend,
            'unr_delta':  unr_delta,  'unr_trend':  unr_trend,
            'rows_delta': rows_delta, 'rows_trend': rows_trend,
            'filt_delta': filt_delta, 'filt_trend': filt_trend,
            'esc_rows':   esc_rows,
            'site_rows':  site_rows,
        }

    return render(request, 'reports/comparer.html', {
        'all_reports': all_reports,
        'r1':   r1,
        'r2':   r2,
        'pk1':  pk1,
        'pk2':  pk2,
        'diff': diff,
    })


def export_statistiques(request):
    import openpyxl
    import openpyxl.chart.label
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart3D, LineChart, Reference, Series
    from openpyxl.chart.series import DataPoint
    from django.http import HttpResponse
    from collections import defaultdict
    from datetime import timedelta
    import io

    # ── Récupère les rapports ────────────────────────────────────────────
    if request.user.is_superuser:
        base_qs = UploadedReport.objects.filter(processed=True).order_by('-uploaded_at')
    else:
        base_qs = UploadedReport.objects.filter(processed=True, user=request.user).order_by('-uploaded_at')

    report_pk     = request.GET.get('report')
    period_filter = request.GET.get('period', 'latest')

    if report_pk:
        reports      = base_qs.filter(pk=report_pk)
        period_label = 'Rapport sélectionné'
    elif period_filter == 'latest' or period_filter not in ('day','3days','week','2weeks','month','quarter','half','year','all','custom'):
        first   = base_qs.first()
        reports = base_qs.filter(pk=first.pk) if first else base_qs.none()
        period_label = 'Dernier rapport'
    else:
        today      = date.today()
        labels_map = {
            'day': "Aujourd'hui", '3days': '3 jours', 'week': 'Semaine',
            '2weeks': '2 semaines', 'month': 'Mois', 'quarter': '3 mois',
            'half': '6 mois', 'year': 'Année', 'all': 'Tout',
            'custom': 'Personnalisé',
        }
        period_label = labels_map.get(period_filter, period_filter)
        if period_filter == 'day':
            reports = base_qs.filter(uploaded_at__date=today)
        elif period_filter == '3days':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=3))
        elif period_filter == 'week':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=7))
        elif period_filter == '2weeks':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=14))
        elif period_filter == 'month':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=30))
        elif period_filter == 'quarter':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=90))
        elif period_filter == 'half':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=180))
        elif period_filter == 'year':
            reports = base_qs.filter(uploaded_at__year=today.year)
        elif period_filter == 'custom':
            date_from = request.GET.get('date_from')
            date_to   = request.GET.get('date_to')
            if date_from and date_to:
                reports = base_qs.filter(uploaded_at__date__gte=date_from, uploaded_at__date__lte=date_to)
                period_label = f'{date_from} → {date_to}'
            else:
                reports = base_qs
        else:
            reports = base_qs

    # ── Helpers ──────────────────────────────────────────────────────────
    def parse_hms(s):
        try:
            parts = str(s).split(':')
            if len(parts) == 3:
                return int(float(parts[0]))*3600 + int(float(parts[1]))*60 + int(float(parts[2]))
        except Exception:
            pass
        return 0

    # ── Calcul données ───────────────────────────────────────────────────
    escalade_data = defaultdict(lambda: {'count': 0, 'outage_sec': 0, 'duree_sec': 0})
    for r in reports:
        if not r.synthesis_json:
            continue
        for row in r.synthesis_json:
            esc = row.get('Escalade', '')
            if esc == 'TOTAL' or not esc:
                continue
            escalade_data[esc]['count']      += row.get('Inc count', 0)
            escalade_data[esc]['outage_sec'] += parse_hms(row.get('OUTAGE', '0:00:00'))
            escalade_data[esc]['duree_sec']  += parse_hms(row.get('DUREE',  '0:00:00'))

    escalades_sorted = sorted(escalade_data.items(), key=lambda x: x[1]['count'], reverse=True)
    total_outage_sec = sum(v['outage_sec'] for v in escalade_data.values())
    total_incidents  = sum(v['count'] for v in escalade_data.values())
    total_duree_sec  = sum(v['duree_sec'] for v in escalade_data.values())

    site_data = defaultdict(int)
    for r in reports:
        for s in (r.top_sites_json or []):
            site_data[s['name']] += s['count']
    sites_top10 = sorted(site_data.items(), key=lambda x: x[1], reverse=True)[:10]

    outage_data = [
        (k, round(v['outage_sec']/3600, 1),
         round(v['outage_sec']/total_outage_sec*100) if total_outage_sec else 0)
        for k, v in escalades_sorted if v['outage_sec'] > 0
    ]

    # Causes durée
    cause_dur_data = defaultdict(float)
    for r in reports:
        for c in (r.top_causes_json or []):
            cause_dur_data[c['name']] += c['duration_sec']
    causes_dur_top10 = sorted(cause_dur_data.items(), key=lambda x: x[1], reverse=True)[:10]
    max_cause_dur    = causes_dur_top10[0][1] if causes_dur_top10 else 1

    # Causes nombre
    cause_nb_data = defaultdict(int)
    for r in reports:
        if not r.detailed_file:
            continue
        file_name = r.detailed_file.name or ''
        if not (('results/' in file_name or 'results\\' in file_name) and file_name.endswith('_detailed.xlsx')):
            continue
        try:
            import pandas as pd
            df = pd.read_excel(r.detailed_file.path)
            cause_col2 = next((c for c in df.columns if c.strip() in ('Root Cause', 'Cause')), None)
            if cause_col2:
                for val in df[cause_col2].dropna().astype(str):
                    val = val.strip()
                    if val and val != 'nan':
                        cause_nb_data[val] += 1
        except Exception:
            continue
    causes_nb_top10 = sorted(cause_nb_data.items(), key=lambda x: x[1], reverse=True)[:10]
    max_cause_nb    = causes_nb_top10[0][1] if causes_nb_top10 else 1

    # Sites dégradés
    site_duration = defaultdict(float)
    for r in reports:
        if not r.detailed_file:
            continue
        file_name = r.detailed_file.name or ''
        if not (('results/' in file_name or 'results\\' in file_name) and file_name.endswith('_detailed.xlsx')):
            continue
        try:
            import pandas as pd
            df = pd.read_excel(r.detailed_file.path)
            site_col = next((c for c in df.columns if c.strip().lower() == 'site name'), None)
            if site_col and 'Duration' in df.columns:
                for _, row in df.iterrows():
                    site = str(row[site_col]).strip()
                    dur  = parse_hms(row['Duration'])
                    if site and site != 'nan':
                        site_duration[site] += dur
        except Exception:
            continue
    degraded_top10 = sorted(site_duration.items(), key=lambda x: x[1], reverse=True)[:10]
    max_degraded   = degraded_top10[0][1] if degraded_top10 else 1

    # Disponibilité
    import datetime as _dt
    if period_filter == 'custom':
        date_from  = request.GET.get('date_from')
        date_to    = request.GET.get('date_to')
        cutoff     = date.fromisoformat(date_from) if date_from else None
        cutoff_end = date.fromisoformat(date_to)   if date_to   else None
        semaine_labels_exp, dispo_table_exp, _ = _calc_disponibilite(base_qs, cutoff_date=cutoff, cutoff_end=cutoff_end)
    elif report_pk:
        first_r    = reports.first()
        cutoff     = first_r.date_rapport if first_r else None
        cutoff_end = (first_r.date_fin or first_r.date_rapport) if first_r else None
        semaine_labels_exp, dispo_table_exp, _ = _calc_disponibilite(reports, cutoff_date=cutoff, cutoff_end=cutoff_end)
    else:
        period_days_map = {'day':1,'3days':3,'week':7,'2weeks':14,'month':30,'quarter':90,'half':180,'year':365}
        nb_days = period_days_map.get(period_filter)
        cutoff  = (date.today() - _dt.timedelta(days=nb_days)) if nb_days else None
        semaine_labels_exp, dispo_table_exp, _ = _calc_disponibilite(base_qs, cutoff_date=cutoff)

    # ── Styles ───────────────────────────────────────────────────────────
    YAS_BLUE    = '003087'
    YAS_YELLOW  = 'FFC72C'
    LIGHT_BLUE  = 'E8F0FF'
    LIGHT_GRAY  = 'F8FAFF'
    MID_GRAY    = 'E8EDF5'
    WHITE       = 'FFFFFF'
    ORANGE      = 'E05A2B'
    PURPLE      = '7B1FA2'
    GREEN_OK    = 'E8F5E9'
    YELLOW_WARN = 'FFF8E1'
    RED_BAD     = 'FFEBEE'

    thin   = Side(style='thin', color=MID_GRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def h_font(sz=11):  return Font(name='Calibri', bold=True, color=WHITE, size=sz)
    def h_fill():       return PatternFill('solid', fgColor=YAS_BLUE)
    def s_fill():       return PatternFill('solid', fgColor=LIGHT_BLUE)
    def a_fill():       return PatternFill('solid', fgColor=LIGHT_GRAY)
    def w_fill():       return PatternFill('solid', fgColor=WHITE)
    def c_align():      return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def l_align():      return Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def style_hdr(ws, row, ncols, height=24):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = h_font(); cell.fill = h_fill()
            cell.alignment = c_align(); cell.border = border
        ws.row_dimensions[row].height = height

    def style_row(ws, row, ncols, alt=False, bold=False, color='1A1A2E', bg=None):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font      = Font(name='Calibri', size=10, bold=bold, color=color)
            cell.fill      = PatternFill('solid', fgColor=bg) if bg else (a_fill() if alt else w_fill())
            cell.alignment = l_align() if c == 1 else c_align()
            cell.border    = border
        ws.row_dimensions[row].height = 20

    def style_total(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name='Calibri', bold=True, size=11, color=WHITE)
            cell.fill = h_fill()
            cell.alignment = l_align() if c == 1 else c_align()
            cell.border = border
        ws.row_dimensions[row].height = 22

    def add_banner(ws, title, ncols=6):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws['A1']
        c.value = f'  📊  YAS NOC — {title}'
        c.font  = Font(name='Calibri', bold=True, size=14, color=YAS_BLUE)
        c.alignment = l_align(); c.fill = s_fill()
        ws.row_dimensions[1].height = 30
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws['A2']
        c.value = f'  Période : {period_label}  |  Généré le {date.today().strftime("%d/%m/%Y")}'
        c.font  = Font(name='Calibri', size=10, color='666666', italic=True)
        c.alignment = l_align()
        c.fill = PatternFill('solid', fgColor='F0F4FF')
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 8

    def add_section(ws, row, title, ncols, color=YAS_BLUE):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=f'  {title}')
        c.font = Font(name='Calibri', bold=True, size=11, color=WHITE)
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = l_align(); c.border = border
        ws.row_dimensions[row].height = 20

    SLICE_COLORS = ['003087','E05A2B','FF9800','2196F3','FFC72C','4CAF50','9C27B0','00BCD4','8BC34A','FF5722']
    ESC_COLORS   = {'ENERGIE':'003087','RAN':'4CAF50','TRANS FH':'FFC72C','TRANS IP':'2196F3'}
    DISPO_ESCS   = ['ENERGIE','RAN','TRANS FH','TRANS IP']

    # ════════════════════════════════════════════════════════════════════
    wb  = openpyxl.Workbook()

    # ── ONGLET 0 : TABLEAU DE BORD ───────────────────────────────────────
    ws0 = wb.active
    ws0.title = '🏠 Tableau de Bord'
    ws0.sheet_view.showGridLines = False
    for col, w in enumerate([3,22,14,14,14,14,14,3], 1):
        ws0.column_dimensions[get_column_letter(col)].width = w

    ws0.row_dimensions[1].height = 10
    ws0.merge_cells('B2:G2')
    c = ws0['B2']
    c.value = 'YAS NOC — REPORTXCARE'
    c.font  = Font(name='Calibri', bold=True, size=20, color=WHITE)
    c.alignment = c_align(); c.fill = h_fill()
    ws0.row_dimensions[2].height = 44

    ws0.merge_cells('B3:G3')
    c = ws0['B3']
    c.value = f'  Période : {period_label}  |  Généré le {date.today().strftime("%d/%m/%Y")}'
    c.font  = Font(name='Calibri', size=11, color='888888', italic=True)
    c.alignment = c_align(); c.fill = s_fill()
    ws0.row_dimensions[3].height = 22
    ws0.row_dimensions[4].height = 16

    # KPIs
    nb_reports = reports.count()
    kpis = [
        ('📋 Rapports',  str(nb_reports),                        YAS_BLUE,  LIGHT_BLUE),
        ('⚡ Incidents',  f'{total_incidents:,}'.replace(',',' '), 'E05A2B',  'FFF0EB'),
        ('⏱ Outage (h)', str(round(total_outage_sec/3600, 1)),   '7B1FA2',  'F3E5F5'),
        ('📡 Escalades',  str(len([e for e in escalades_sorted if e[1]["count"]>0])), '003087', LIGHT_BLUE),
    ]
    for i, (lbl, val, fcol, bcol) in enumerate(kpis):
        col = 2 + i
        for r, v, sz, bold in [(5, lbl, 9, False), (6, val, 20, True), (7, '', 9, False)]:
            cell = ws0.cell(row=r, column=col, value=v)
            cell.font      = Font(name='Calibri', size=sz, bold=bold, color=fcol if r==6 else '888888')
            cell.fill      = PatternFill('solid', fgColor=bcol)
            cell.alignment = c_align()
            cell.border    = border
        for r in [5,6,7]: ws0.row_dimensions[r].height = 18 if r!=6 else 36

    ws0.row_dimensions[8].height = 16

    # Sommaire
    add_section(ws0, 9, '📑  Contenu du classeur — Navigation', 6)
    for c, h in enumerate(['Onglet','Description','Données incluses','Nb lignes'], 2):
        cell = ws0.cell(row=10, column=c, value=h)
        cell.font = Font(name='Calibri', bold=True, size=10, color=YAS_BLUE)
        cell.fill = s_fill(); cell.alignment = c_align(); cell.border = border
    ws0.row_dimensions[10].height = 20

    sheets_info = [
        ('📊 Escalades',     'Classement par type d\'escalade',           'Incidents, Outage, Durée, %',   str(len(escalades_sorted))),
        ('📡 Sites',         'Récurrence des sites en panne',             'Top 10 occurrences',            '10'),
        ('🥧 Outage Métier', 'Répartition outage par métier + camembert', 'Heures, Pourcentages',          str(len(outage_data))),
        ('🔧 Causes Durée',  'Causes triées par durée d\'outage',         'Top 10 causes + graphique',     str(len(causes_dur_top10))),
        ('🔢 Causes Nombre', 'Causes triées par nombre d\'incidents',     'Top 10 causes + graphique',     str(len(causes_nb_top10))),
        ('🔴 Dégradés',      'Sites avec le plus d\'outage cumulé',       'Top 10 + niveau criticité',     str(len(degraded_top10))),
        ('📈 Disponibilité', 'Disponibilité % équipements/semaine',       '4 équipements, Min, Moyenne',   str(len(DISPO_ESCS))),
    ]
    for i, (ong, desc, data, nb) in enumerate(sheets_info):
        r = 11 + i
        for c, v in enumerate([ong, desc, data, nb], 2):
            cell = ws0.cell(row=r, column=c, value=v)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = a_fill() if i%2 else w_fill()
            cell.alignment = l_align() if c == 2 else c_align()
            cell.border    = border
        ws0.row_dimensions[r].height = 18

    # ── ONGLET 1 : ESCALADES ────────────────────────────────────────────
    ws1 = wb.create_sheet('📊 Escalades')
    ws1.sheet_view.showGridLines = False
    for col, w in enumerate([32,12,12,12,10,10], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    add_banner(ws1, 'Classement des Escalades', ncols=6)
    headers1 = ['Escalade','Incidents','Outage (h)','Durée (h)','% Outage','% Incidents']
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=c, value=h)
    style_hdr(ws1, 4, 6)

    for i, (esc, v) in enumerate(escalades_sorted):
        r = 5 + i
        pct_out = round(v['outage_sec']/total_outage_sec*100, 1) if total_outage_sec else 0
        pct_inc = round(v['count']/total_incidents*100, 1) if total_incidents else 0
        for c, val in enumerate([esc, v['count'], round(v['outage_sec']/3600,1), round(v['duree_sec']/3600,1), f'{pct_out}%', f'{pct_inc}%'], 1):
            ws1.cell(row=r, column=c, value=val)
        style_row(ws1, r, 6, alt=(i%2==1))

    tr1 = 5 + len(escalades_sorted)
    for c, v in enumerate(['TOTAL', total_incidents, round(total_outage_sec/3600,1), round(total_duree_sec/3600,1), '100%','100%'], 1):
        ws1.cell(row=tr1, column=c, value=v)
    style_total(ws1, tr1, 6)

    bc1 = BarChart()
    bc1.type = 'bar'; bc1.title = 'Incidents par Escalade'; bc1.style = 10
    bc1.width = 20; bc1.height = 14
    d1 = Reference(ws1, min_col=2, min_row=4, max_row=4+len(escalades_sorted))
    l1 = Reference(ws1, min_col=1, min_row=5, max_row=4+len(escalades_sorted))
    bc1.add_data(d1, titles_from_data=True); bc1.set_categories(l1)
    bc1.series[0].graphicalProperties.solidFill = YAS_BLUE
    ws1.add_chart(bc1, 'H4')

    # ── ONGLET 2 : SITES ────────────────────────────────────────────────
    ws2 = wb.create_sheet('📡 Sites')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 20

    add_banner(ws2, 'Récurrence des Sites (Top 10)', ncols=3)
    for c, h in enumerate(['Site','Occurrences','Barre visuelle'], 1):
        ws2.cell(row=4, column=c, value=h)
    style_hdr(ws2, 4, 3)

    max_site = sites_top10[0][1] if sites_top10 else 1
    for i, (site, cnt) in enumerate(sites_top10):
        r = 5 + i
        pct = round(cnt/max_site*100)
        bar = '█' * (pct//5) + '░' * (20 - pct//5)
        ws2.cell(row=r, column=1, value=site)
        ws2.cell(row=r, column=2, value=cnt)
        c3 = ws2.cell(row=r, column=3, value=bar)
        c3.font = Font(name='Courier New', size=9, color=YAS_BLUE)
        style_row(ws2, r, 2, alt=(i%2==1))
        ws2.cell(row=r, column=3).fill   = a_fill() if i%2==1 else w_fill()
        ws2.cell(row=r, column=3).border = border

    bc2 = BarChart()
    bc2.type = 'bar'; bc2.title = 'Récurrence des Sites'; bc2.style = 10
    bc2.width = 18; bc2.height = 12
    d2 = Reference(ws2, min_col=2, min_row=4, max_row=4+len(sites_top10))
    l2 = Reference(ws2, min_col=1, min_row=5, max_row=4+len(sites_top10))
    bc2.add_data(d2, titles_from_data=True); bc2.set_categories(l2)
    bc2.series[0].graphicalProperties.solidFill = YAS_YELLOW
    ws2.add_chart(bc2, 'E4')

    # ── ONGLET 3 : OUTAGE MÉTIER ─────────────────────────────────────────
    ws3 = wb.create_sheet('🥧 Outage Métier')
    ws3.sheet_view.showGridLines = False
    for col, w in enumerate([30,14,10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    add_banner(ws3, 'Outage par Métier', ncols=3)
    for c, h in enumerate(['Métier / Escalade','Outage (h)','% Total'], 1):
        ws3.cell(row=4, column=c, value=h)
    style_hdr(ws3, 4, 3)

    for i, (name, h, pct) in enumerate(outage_data):
        r = 5 + i
        ws3.cell(row=r, column=1, value=name)
        ws3.cell(row=r, column=2, value=h)
        ws3.cell(row=r, column=3, value=f'{pct}%')
        style_row(ws3, r, 3, alt=(i%2==1))

    tr3 = 5 + len(outage_data)
    for c, v in enumerate(['TOTAL', round(total_outage_sec/3600,1), '100%'], 1):
        ws3.cell(row=tr3, column=c, value=v)
    style_total(ws3, tr3, 3)

    if outage_data:
        pie = PieChart3D()
        pie.title = 'Outage par Métier'; pie.style = 10
        pie.width = 18; pie.height = 16
        n_rows = len(outage_data)
        dr = Reference(ws3, min_col=2, min_row=4, max_row=4+n_rows)
        lr = Reference(ws3, min_col=1, min_row=5, max_row=4+n_rows)
        pie.add_data(dr, titles_from_data=True); pie.set_categories(lr)
        series = pie.series[0]
        for idx in range(n_rows):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = SLICE_COLORS[idx % len(SLICE_COLORS)]
            series.dPt.append(pt)
        series.dLbls = openpyxl.chart.label.DataLabelList()
        series.dLbls.showCatName = True; series.dLbls.showPercent = True
        series.dLbls.showVal = False; series.dLbls.showLegendKey = False
        series.dLbls.separator = '\n'
        ws3.add_chart(pie, 'E4')

    # ── ONGLET 4 : CAUSES DURÉE ──────────────────────────────────────────
    ws4 = wb.create_sheet('🔧 Causes Durée')
    ws4.sheet_view.showGridLines = False
    for col, w in enumerate([42,14,10], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    add_banner(ws4, 'Outage par Cause — Durée (Top 10)', ncols=3)
    add_section(ws4, 4, '🔧  Classement par durée cumulée d\'outage', 3, ORANGE)
    for c, h in enumerate(['Cause / Root Cause','Durée (h)','% du Max'], 1):
        ws4.cell(row=5, column=c, value=h)
    style_hdr(ws4, 5, 3)

    for i, (cause, dur_sec) in enumerate(causes_dur_top10):
        r = 6 + i
        dur_h = round(dur_sec/3600, 1)
        pct   = round(dur_sec/max_cause_dur*100, 1)
        ws4.cell(row=r, column=1, value=cause)
        ws4.cell(row=r, column=2, value=dur_h)
        ws4.cell(row=r, column=3, value=f'{pct}%')
        style_row(ws4, r, 3, alt=(i%2==1))

    if causes_dur_top10:
        bc4 = BarChart()
        bc4.type = 'bar'; bc4.title = 'Causes — Durée Outage (h)'; bc4.style = 10
        bc4.width = 20; bc4.height = 14
        n4 = len(causes_dur_top10)
        d4 = Reference(ws4, min_col=2, min_row=5, max_row=5+n4)
        l4 = Reference(ws4, min_col=1, min_row=6, max_row=5+n4)
        bc4.add_data(d4, titles_from_data=True); bc4.set_categories(l4)
        bc4.series[0].graphicalProperties.solidFill = ORANGE
        ws4.add_chart(bc4, 'E5')

    # ── ONGLET 5 : CAUSES NOMBRE ─────────────────────────────────────────
    ws5 = wb.create_sheet('🔢 Causes Nombre')
    ws5.sheet_view.showGridLines = False
    for col, w in enumerate([42,14,10], 1):
        ws5.column_dimensions[get_column_letter(col)].width = w

    add_banner(ws5, 'Outage par Cause — Nombre d\'incidents (Top 10)', ncols=3)
    add_section(ws5, 4, '🔢  Classement par nombre d\'incidents', 3, PURPLE)
    for c, h in enumerate(['Cause / Root Cause','Incidents','% du Max'], 1):
        ws5.cell(row=5, column=c, value=h)
    style_hdr(ws5, 5, 3)

    for i, (cause, nb) in enumerate(causes_nb_top10):
        r = 6 + i
        pct = round(nb/max_cause_nb*100, 1)
        ws5.cell(row=r, column=1, value=cause)
        ws5.cell(row=r, column=2, value=nb)
        ws5.cell(row=r, column=3, value=f'{pct}%')
        style_row(ws5, r, 3, alt=(i%2==1))

    if causes_nb_top10:
        bc5 = BarChart()
        bc5.type = 'bar'; bc5.title = 'Causes — Nombre d\'Incidents'; bc5.style = 10
        bc5.width = 20; bc5.height = 14
        n5 = len(causes_nb_top10)
        d5 = Reference(ws5, min_col=2, min_row=5, max_row=5+n5)
        l5 = Reference(ws5, min_col=1, min_row=6, max_row=5+n5)
        bc5.add_data(d5, titles_from_data=True); bc5.set_categories(l5)
        bc5.series[0].graphicalProperties.solidFill = PURPLE
        ws5.add_chart(bc5, 'E5')

    # ── ONGLET 6 : SITES DÉGRADÉS ────────────────────────────────────────
    ws6 = wb.create_sheet('🔴 Dégradés')
    ws6.sheet_view.showGridLines = False
    for col, w in enumerate([34,16,14], 1):
        ws6.column_dimensions[get_column_letter(col)].width = w

    add_banner(ws6, 'Sites les Plus Dégradés (Top 10)', ncols=3)
    for c, h in enumerate(['Site','Durée totale (h)','Criticité'], 1):
        ws6.cell(row=4, column=c, value=h)
    style_hdr(ws6, 4, 3)

    for i, (site, dur_sec) in enumerate(degraded_top10):
        r = 5 + i
        dur_h = round(dur_sec/3600, 1)
        pct   = dur_sec / max_degraded if max_degraded else 0
        if pct >= 0.8:   crit, bg = '🔴 Critique', 'FFEBEE'
        elif pct >= 0.5: crit, bg = '🟠 Élevé',    'FFF3E0'
        else:            crit, bg = '🟡 Modéré',   'FFFDE7'
        ws6.cell(row=r, column=1, value=site)
        ws6.cell(row=r, column=2, value=dur_h)
        ws6.cell(row=r, column=3, value=crit)
        for col in range(1, 4):
            cell = ws6.cell(row=r, column=col)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = l_align() if col == 1 else c_align()
            cell.border    = border
        ws6.row_dimensions[r].height = 20

    if degraded_top10:
        bc6 = BarChart()
        bc6.type = 'bar'; bc6.title = 'Sites les Plus Dégradés'; bc6.style = 10
        bc6.width = 18; bc6.height = 12
        n6 = len(degraded_top10)
        d6 = Reference(ws6, min_col=2, min_row=4, max_row=4+n6)
        l6 = Reference(ws6, min_col=1, min_row=5, max_row=4+n6)
        bc6.add_data(d6, titles_from_data=True); bc6.set_categories(l6)
        bc6.series[0].graphicalProperties.solidFill = 'E53E3E'
        ws6.add_chart(bc6, 'E4')

    # ── ONGLET 7 : DISPONIBILITÉ ─────────────────────────────────────────
    if semaine_labels_exp:
        ws7 = wb.create_sheet('📈 Disponibilité')
        ws7.sheet_view.showGridLines = False
        n_weeks   = len(semaine_labels_exp)
        all_cols7 = 1 + n_weeks + 2
        ws7.column_dimensions['A'].width = 16
        for i in range(1, n_weeks+3):
            ws7.column_dimensions[get_column_letter(i+1)].width = 15

        add_banner(ws7, 'Disponibilité Équipements (%)', ncols=all_cols7)
        add_section(ws7, 4, '📶  Taux de disponibilité par équipement et par semaine', all_cols7)

        ws7.cell(row=5, column=1, value='Équipement')
        for j, lbl in enumerate(semaine_labels_exp, 2):
            ws7.cell(row=5, column=j, value=lbl)
        ws7.cell(row=5, column=n_weeks+2, value='Min (%)')
        ws7.cell(row=5, column=n_weeks+3, value='Moy (%)')
        style_hdr(ws7, 5, all_cols7)

        all_dispo_vals = []
        for i, esc in enumerate(DISPO_ESCS):
            r = 6 + i
            ws7.cell(row=r, column=1, value=esc).font = Font(name='Calibri', bold=True, size=11, color=ESC_COLORS.get(esc, YAS_BLUE))
            ws7.cell(row=r, column=1).fill      = s_fill()
            ws7.cell(row=r, column=1).alignment = l_align()
            ws7.cell(row=r, column=1).border    = border
            for j, lbl in enumerate(semaine_labels_exp, 2):
                val = dispo_table_exp.get(esc, {}).get(lbl)
                if val is not None:
                    all_dispo_vals.append(val)
                    c = ws7.cell(row=r, column=j, value=round(val, 4))
                    c.number_format = '0.0000"%"'
                    if val >= 99.9:    c.fill = PatternFill('solid', fgColor='E8F5E9')
                    elif val >= 99.5:  c.fill = PatternFill('solid', fgColor='FFF8E1')
                    else:              c.fill = PatternFill('solid', fgColor='FFEBEE')
                    c.font      = Font(name='Calibri', size=10, bold=(val < 99.5))
                    c.alignment = c_align(); c.border = border
            # Min et Moy
            c1 = get_column_letter(2)
            cn = get_column_letter(n_weeks+1)
            for col_idx, formula in [(n_weeks+2, f'=MIN({c1}{r}:{cn}{r})'), (n_weeks+3, f'=AVERAGE({c1}{r}:{cn}{r})')]:
                c = ws7.cell(row=r, column=col_idx, value=formula)
                c.number_format = '0.0000"%"'
                c.font          = Font(name='Calibri', size=10, bold=True, color=ESC_COLORS.get(esc, YAS_BLUE))
                c.alignment     = c_align(); c.border = border; c.fill = s_fill()
            ws7.row_dimensions[r].height = 22

        # Légende
        leg_row = 6 + len(DISPO_ESCS) + 2
        ws7.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=min(all_cols7, 6))
        ws7.cell(row=leg_row, column=1, value='  Légende :   🟢 ≥ 99.9% Excellent    🟡 ≥ 99.5% Bon    🔴 < 99.5% Critique')
        ws7.cell(row=leg_row, column=1).font = Font(name='Calibri', size=9, italic=True, color='555555')
        ws7.row_dimensions[leg_row].height = 18

        # Graphique courbes
        lc = LineChart()
        lc.title = 'Disponibilité Équipements (%)'; lc.style = 10
        lc.y_axis.title = 'Disponibilité (%)'; lc.x_axis.title = 'Semaine'
        lc.width = 26; lc.height = 16
        lc.y_axis.numFmt = '0.00"%"'; lc.x_axis.tickLblPos = 'low'

        for i, esc in enumerate(DISPO_ESCS):
            data_row = 6 + i
            dr = Reference(ws7, min_col=2, max_col=1+n_weeks, min_row=data_row, max_row=data_row)
            serie = Series(dr, title=esc)
            serie.graphicalProperties.line.solidFill        = ESC_COLORS.get(esc, YAS_BLUE)
            serie.graphicalProperties.line.width            = 25000
            serie.smooth                                    = True
            serie.marker.symbol                             = 'circle'
            serie.marker.size                               = 7
            serie.marker.graphicalProperties.solidFill      = ESC_COLORS.get(esc, YAS_BLUE)
            serie.marker.graphicalProperties.line.solidFill = ESC_COLORS.get(esc, YAS_BLUE)
            lc.series.append(serie)

        cats = Reference(ws7, min_col=2, max_col=1+n_weeks, min_row=5)
        lc.set_categories(cats)
        if all_dispo_vals:
            lc.y_axis.scaling.min = round(min(all_dispo_vals) - 0.2, 1)
            lc.y_axis.scaling.max = 100.05
        lc.legend.position = 'b'
        ws7.add_chart(lc, f'A{leg_row + 2}')

    # ── Export ───────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_period = period_label.replace(' ', '_').replace('/', '-').replace('→', 'to')[:40]
    filename    = f"Statistiques_YAS_{date.today().strftime('%Y%m%d')}_{safe_period}.xlsx"
    response    = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _calc_disponibilite(reports, cutoff_date=None, cutoff_end=None):
    import datetime as _dt
    from collections import defaultdict

    MOIS_FR = ['', 'Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin',
               'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc']

    def _week_label(year, week):
        lundi    = _dt.date.fromisocalendar(year, week, 1)
        dimanche = lundi + _dt.timedelta(days=6)
        if lundi.month == dimanche.month:
            return f"{lundi.day}-{dimanche.day} {MOIS_FR[dimanche.month]}"
        return f"{lundi.day} {MOIS_FR[lundi.month]}-{dimanche.day} {MOIS_FR[dimanche.month]}"

    outage_par_jour = defaultdict(lambda: defaultdict(float))
    for r in reports:
        for esc, jours in (r.outage_journalier_json or {}).items():
            for day_str, sec in jours.items():
                outage_par_jour[esc][day_str] += sec

    if not outage_par_jour:
        return [], {}, {}

    all_days = set()
    for jours in outage_par_jour.values():
        all_days.update(jours.keys())

    semaines_set = set()
    for day_str in all_days:
        try:
            d = _dt.date.fromisoformat(day_str)
            if cutoff_date and d < cutoff_date:
                continue
            if cutoff_end and d > cutoff_end:
                continue
            iso = d.isocalendar()
            semaines_set.add((iso[0], iso[1]))
        except Exception:
            pass

    if not semaines_set:
        return [], {}, {}

    semaines_sorted = sorted(semaines_set)
    semaine_labels  = [_week_label(y, w) for y, w in semaines_sorted]
    semaine_keys    = [f"S{w:02d}-{y}" for y, w in semaines_sorted]

    outage_sec_table = defaultdict(lambda: defaultdict(float))
    for esc, jours in outage_par_jour.items():
        for day_str, sec in jours.items():
            try:
                d   = _dt.date.fromisoformat(day_str)
                iso = d.isocalendar()
                key = f"S{iso[1]:02d}-{iso[0]}"
                outage_sec_table[esc][key] += sec
            except Exception:
                pass

    dispo_table  = {}
    outage_table = {}
    for esc in NB_SITES:
        nb = NB_SITES[esc]
        dispo_table[esc]  = {}
        outage_table[esc] = {}
        for lbl, key in zip(semaine_labels, semaine_keys):
            outage_sec = outage_sec_table[esc].get(key, 0)
            total_sec  = nb * 7 * 24 * 3600
            dispo_pct  = round((1 - outage_sec / total_sec) * 100, 4) if total_sec else None
            h  = int(outage_sec // 3600)
            m  = int((outage_sec % 3600) // 60)
            dispo_table[esc][lbl]  = dispo_pct
            outage_table[esc][lbl] = f"{h}:{m:02d}"

    return semaine_labels, dispo_table, outage_table


def statistiques(request):
    from collections import defaultdict
    from datetime import timedelta
    import datetime as _dt

    platform = request.GET.get('platform', 'mobile')
    PLATFORM_LABELS = {
        'all':          '🌐 TOUTES LES PLATEFORMES',
        'mobile':       '📡 RÉSEAU MOBILE',
        'fixe':         '🔌 RÉSEAU FIXE',
        'core':         '⚙️ CORE ET IGW',
        'transmission': '📶 TRANSMISSION',
    }
    # Plateformes sans données : retour rapide avec carte "bientôt"
    if platform not in ('mobile', 'all'):
        return render(request, 'reports/statistiques.html', {
            'platform':       platform,
            'platform_label': PLATFORM_LABELS.get(platform, platform.upper()),
            'platform_labels': PLATFORM_LABELS,
        })

    if request.user.is_superuser:
        base_qs = UploadedReport.objects.filter(processed=True).order_by('-uploaded_at')
    else:
        base_qs = UploadedReport.objects.filter(processed=True, user=request.user).order_by('-uploaded_at')

    report_pk     = request.GET.get('report')
    single_report = None
    period_filter = request.GET.get('period', 'latest')

    if report_pk:
        single_report = get_object_or_404(base_qs, pk=report_pk)
        reports       = base_qs.filter(pk=report_pk)
        period_filter = 'report'
    elif period_filter == 'latest' or period_filter not in (
            'day','3days','week','2weeks','month','quarter','half','year','all','custom'):
        single_report = base_qs.first()
        reports       = base_qs.filter(pk=single_report.pk) if single_report else base_qs.none()
        period_filter = 'latest'
    else:
        today = date.today()
        if period_filter == 'day':
            reports = base_qs.filter(uploaded_at__date=today)
        elif period_filter == '3days':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=3))
        elif period_filter == 'week':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=7))
        elif period_filter == '2weeks':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=14))
        elif period_filter == 'month':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=30))
        elif period_filter == 'quarter':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=90))
        elif period_filter == 'half':
            reports = base_qs.filter(uploaded_at__date__gte=today - timedelta(days=180))
        elif period_filter == 'year':
            reports = base_qs.filter(uploaded_at__year=today.year)
        elif period_filter == 'custom':
            date_from = request.GET.get('date_from')
            date_to   = request.GET.get('date_to')
            if date_from and date_to:
                try:
                    from django.utils import timezone as _tz
                    dt_from = _tz.make_aware(_dt.datetime.fromisoformat(date_from))
                    dt_to   = _tz.make_aware(_dt.datetime.fromisoformat(date_to))
                    reports = base_qs.filter(uploaded_at__gte=dt_from, uploaded_at__lte=dt_to)
                except (ValueError, TypeError):
                    reports = base_qs
            else:
                reports = base_qs
        else:
            reports = base_qs

    def parse_hms(s):
        try:
            parts = str(s).split(':')
            if len(parts) == 3:
                return int(float(parts[0])) * 3600 + int(float(parts[1])) * 60 + int(float(parts[2]))
        except Exception:
            pass
        return 0

    escalade_data = defaultdict(lambda: {'count': 0, 'outage_sec': 0, 'duree_sec': 0})
    for r in reports:
        if not r.synthesis_json:
            continue
        for row in r.synthesis_json:
            esc = row.get('Escalade', '')
            if esc == 'TOTAL' or not esc:
                continue
            escalade_data[esc]['count']      += row.get('Inc count', 0)
            escalade_data[esc]['outage_sec'] += parse_hms(row.get('OUTAGE', '0:00:00'))
            escalade_data[esc]['duree_sec']  += parse_hms(row.get('DUREE',  '0:00:00'))

    escalades_sorted = sorted(escalade_data.items(), key=lambda x: x[1]['count'], reverse=True)
    max_esc          = escalades_sorted[0][1]['count'] if escalades_sorted else 1
    escalades_chart  = [
        {
            'name':   k,
            'count':  v['count'],
            'pct':    round(v['count'] / max_esc * 100),
            'outage': round(v['outage_sec'] / 3600, 1),
        }
        for k, v in escalades_sorted if v['count'] > 0
    ]

    site_data = defaultdict(int)
    for r in reports:
        if r.top_sites_json:
            for s in r.top_sites_json:
                site_data[s['name']] += s['count']
    sites_top10 = sorted(site_data.items(), key=lambda x: x[1], reverse=True)[:10]
    max_site    = sites_top10[0][1] if sites_top10 else 1
    sites_chart = [
        {'name': k, 'count': v, 'pct': round(v / max_site * 100)}
        for k, v in sites_top10
    ]

    total_outage_sec = sum(v['outage_sec'] for v in escalade_data.values())

    SEUIL_AUTRES = 3
    outage_chart = []
    autres_sec   = 0.0
    for k, v in escalades_sorted:
        if v['outage_sec'] > 0:
            pct = round(v['outage_sec'] / total_outage_sec * 100) if total_outage_sec else 0
            if pct < SEUIL_AUTRES:
                autres_sec += v['outage_sec']
            else:
                outage_chart.append({
                    'name':     k,
                    'outage_h': round(v['outage_sec'] / 3600, 1),
                    'pct':      pct,
                })
    if autres_sec > 0:
        autres_pct = max(0, 100 - sum(d['pct'] for d in outage_chart))
        outage_chart.append({
            'name':      'Autres',
            'outage_h':  round(autres_sec / 3600, 1),
            'pct':       autres_pct,
            'is_autres': True,
        })

    total_outage_h = round(total_outage_sec / 3600, 1)

    site_duration = defaultdict(float)
    for r in reports:
        if not r.detailed_file:
            continue
        file_name = r.detailed_file.name or ''
        if not (('results/' in file_name or 'results\\' in file_name) and file_name.endswith('_detailed.xlsx')):
            continue
        try:
            import pandas as pd
            df = pd.read_excel(r.detailed_file.path)
            site_col = next((c for c in df.columns if c.strip().lower() == 'site name'), None)
            if site_col and 'Duration' in df.columns:
                for _, row in df.iterrows():
                    site = str(row[site_col]).strip()
                    dur  = parse_hms(row['Duration'])
                    if site and site != 'nan':
                        site_duration[site] += dur
        except Exception:
            continue

    degraded_top10 = sorted(site_duration.items(), key=lambda x: x[1], reverse=True)[:10]
    max_deg        = degraded_top10[0][1] if degraded_top10 else 1
    degraded_chart = [
        {'name': k, 'duration_h': round(v / 3600, 1), 'pct': round(v / max_deg * 100)}
        for k, v in degraded_top10
    ]

    # ── Incident par Cause (Durée) ──────────────────────────────────────────
    cause_data = defaultdict(float)
    for r in reports:
        for c in (r.top_causes_json or []):
            cause_data[c['name']] += c['duration_sec']
    causes_top10 = sorted(cause_data.items(), key=lambda x: x[1], reverse=True)[:10]
    max_cause    = causes_top10[0][1] if causes_top10 else 1
    causes_chart = [
        {'name': k, 'duree_h': round(v / 3600, 1), 'pct': round(v / max_cause * 100)}
        for k, v in causes_top10
    ]

    # ── Incidents par Cause (Nombre) ──────────────────────────────────────
    cause_count_data = defaultdict(int)
    for r in reports:
        if not r.detailed_file:
            continue
        file_name = r.detailed_file.name or ''
        if not (('results/' in file_name or 'results\\' in file_name) and file_name.endswith('_detailed.xlsx')):
            continue
        try:
            import pandas as pd
            df = pd.read_excel(r.detailed_file.path)
            cause_col2 = next((c for c in df.columns if c.strip() in ('Root Cause', 'Cause')), None)
            if cause_col2:
                for val in df[cause_col2].dropna().astype(str):
                    val = val.strip()
                    if val and val != 'nan':
                        cause_count_data[val] += 1
        except Exception:
            continue
    causes_count_top10 = sorted(cause_count_data.items(), key=lambda x: x[1], reverse=True)[:10]
    max_cause_count    = causes_count_top10[0][1] if causes_count_top10 else 1
    causes_count_chart = [
        {'name': k, 'count': v, 'pct': round(v / max_cause_count * 100)}
        for k, v in causes_count_top10
    ]

    donut_svg = _make_donut_svg(outage_chart, total_outage_h)
    outage_chart_colored = [
        {**d, 'color': DONUT_COLORS[i % len(DONUT_COLORS)]}
        for i, d in enumerate(outage_chart)
    ]

    total_reports          = reports.count()
    total_sites_impacted   = len(site_data)
    evolution_reports = list(reports.order_by('uploaded_at'))
    evolution_labels  = [r.uploaded_at.strftime('%d/%m') for r in evolution_reports]
    evolution_incidents = [r.total_incidents for r in evolution_reports]
    evolution_outage    = [round(r.total_duration_sec / 3600, 1) for r in evolution_reports]

    # ── Disponibilité dynamique selon filtre ──────────────────────────────
    if period_filter == 'custom':
        date_from  = request.GET.get('date_from')
        date_to    = request.GET.get('date_to')
        cutoff     = _dt.datetime.fromisoformat(date_from).date() if date_from else None
        cutoff_end = _dt.datetime.fromisoformat(date_to).date()   if date_to   else None
        semaine_labels, dispo_table, outage_table = _calc_disponibilite(
            base_qs, cutoff_date=cutoff, cutoff_end=cutoff_end
        )
    elif period_filter in ('report', 'latest') and single_report:
        date_debut = single_report.date_rapport
        cutoff_end = single_report.date_fin or date_debut
        semaine_labels, dispo_table, outage_table = _calc_disponibilite(
            reports, cutoff_date=date_debut, cutoff_end=cutoff_end
        )
    else:
        period_days = {
            '3days': 3, 'week': 7, '2weeks': 14, 'month': 30,
            'quarter': 90, 'half': 180, 'year': 365,
        }
        nb_days = period_days.get(period_filter)
        cutoff  = (date.today() - _dt.timedelta(days=nb_days)) if nb_days else None
        semaine_labels, dispo_table, outage_table = _calc_disponibilite(base_qs, cutoff_date=cutoff)

    DISPO_COLORS = {
        'ENERGIE':  '#003087',
        'RAN':      '#4CAF50',
        'TRANS FH': '#FFC72C',
        'TRANS IP': '#2196F3',
    }
    dispo_series = []
    for esc, color in DISPO_COLORS.items():
        if esc in dispo_table and any(v for v in dispo_table[esc].values()):
            dispo_series.append({
                'name':  esc,
                'color': color,
                'data':  [dispo_table[esc].get(lbl) for lbl in semaine_labels],
            })

    import json as _json
    dispo_series_js = [
        {
            'name':  s['name'],
            'color': s['color'],
            'data':  [dispo_table[s['name']].get(lbl) for lbl in semaine_labels],
        }
        for s in dispo_series
    ]

    # ── Date range pour les inputs dynamiques ────────────────────────────────
    first_up = base_qs.order_by('uploaded_at').values_list('uploaded_at', flat=True).first()
    last_up  = base_qs.order_by('-uploaded_at').values_list('uploaded_at', flat=True).first()
    date_min_str = first_up.strftime('%Y-%m-%dT%H:%M') if first_up else ''
    date_max_str = last_up.strftime('%Y-%m-%dT%H:%M')  if last_up  else ''
    date_from_val = request.GET.get('date_from') or date_min_str
    date_to_val   = request.GET.get('date_to')   or date_max_str

    return render(request, 'reports/statistiques.html', {
        'period_filter':        period_filter,
        'single_report':        single_report,
        'escalades_chart':      escalades_chart,
        'sites_chart':          sites_chart,
        'outage_chart':         outage_chart_colored,
        'degraded_chart':       degraded_chart,
        'causes_chart':         causes_chart,
        'causes_count_chart':   causes_count_chart,
        'total_outage_h':       total_outage_h,
        'total_reports':        total_reports,
        'total_sites_impacted': total_sites_impacted,
        'donut_svg':            donut_svg,
        'show_evolution_chart': len(evolution_labels) > 1,
        'evolution_labels':     mark_safe(json.dumps(evolution_labels)),
        'evolution_incidents':  mark_safe(json.dumps(evolution_incidents)),
        'evolution_outage':     mark_safe(json.dumps(evolution_outage)),
        'semaine_labels':       semaine_labels,
        'dispo_table':          dispo_table,
        'outage_table':         outage_table,
        'dispo_series':         dispo_series,
        'nb_sites':             NB_SITES,
        'semaine_labels_js':    mark_safe(_json.dumps(semaine_labels)),
        'dispo_series_js':      mark_safe(_json.dumps(dispo_series_js)),
        'date_from_val':        date_from_val,
        'date_to_val':          date_to_val,
        'date_min_str':         date_min_str,
        'date_max_str':         date_max_str,
        'platform':             platform,
        'platform_label':       PLATFORM_LABELS.get(platform, '📡 RÉSEAU MOBILE'),
        'platform_labels':      PLATFORM_LABELS,
    })


def sites_instables(request):
    return render(request, 'reports/sites_instables.html')


from .reporting_config import PLATFORMS as REPORTING_PLATFORMS

# Alias de compatibilité (gardé pour _build_network_section, etc.)
REPORTING_NETWORKS = {
    k: {'label': v['label'], 'domains': v['domains'], 'icon': v['icon']}
    for k, v in REPORTING_PLATFORMS.items()
}

# Clé de groupement par domaine (colonne utilisée pour la synthèse)
DOMAIN_GROUP_FIELD = {
    'mobile':    'escalade',
    'dr2':       'escalade',
    'fixe':      'escalade',
    'transport': 'escalade',
    'igw':       'escalade',
    'core':      'escalade',
}


def _fmt_sec(secs):
    if not secs:
        return '0:00:00'
    secs = int(secs)
    return f"{secs // 3600}:{(secs % 3600) // 60:02d}:{secs % 60:02d}"


def reporting_import(request):
    """Upload du fichier multi-feuilles BASES DES INCIDENTS et import dans Incident."""
    from .models import Incident
    from reports.management.commands.import_incidents import PARSERS

    if request.method != 'POST':
        return redirect('reporting')

    uploaded = request.FILES.get('incidents_file')
    if not uploaded:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('reporting')

    ext = uploaded.name.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls'):
        messages.error(request, 'Format non supporté. Utilisez un fichier Excel (.xlsx).')
        return redirect('reporting')

    clear_mois = request.POST.get('clear_mois') == '1'
    domains_sel = request.POST.getlist('domains') or list(PARSERS.keys())

    # Sauvegarder le fichier temporairement
    import tempfile, os
    suffix = f'.{ext}'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    source = uploaded.name
    total_created = 0
    errors = []

    try:
        for domain in domains_sel:
            if domain not in PARSERS:
                continue
            try:
                incidents = PARSERS[domain](tmp_path, source)
            except Exception as e:
                errors.append(f'{domain}: {e}')
                continue

            if not incidents:
                continue

            mois = incidents[0].mois_rapport
            if clear_mois and mois:
                Incident.objects.filter(domain=domain, mois_rapport=mois).delete()

            Incident.objects.bulk_create(incidents, batch_size=500)
            total_created += len(incidents)
    finally:
        os.unlink(tmp_path)

    if errors:
        messages.warning(request, f'{total_created} incidents importés. Erreurs : ' + ' | '.join(errors))
    else:
        messages.success(request, f'{total_created} incidents importés avec succès depuis « {source} ».')

    return redirect('reporting')


def reporting(request):
    """Page d'accueil du reporting — liste des plateformes."""
    from .models import Incident
    from django.db.models import Count, Sum

    platforms_ctx = []
    for key, cfg in REPORTING_PLATFORMS.items():
        qs = Incident.objects.filter(domain__in=cfg['domains'])
        dernier_mois = (
            qs.exclude(mois_rapport__isnull=True)
            .values_list('mois_rapport', flat=True)
            .order_by('-mois_rapport').first()
        )
        agg = qs.filter(mois_rapport=dernier_mois).aggregate(
            total=Count('id'), outage=Sum('duration_sec')
        ) if dernier_mois else {'total': 0, 'outage': 0}

        nb_excel = len(cfg.get('excel_reports', []))
        nb_pptx  = len(cfg.get('pptx_reports', []))

        platforms_ctx.append({
            'key':          key,
            'label':        cfg['label'],
            'icon':         cfg['icon'],
            'color':        cfg['color'],
            'color2':       cfg['color2'],
            'total':        agg['total'] or 0,
            'outage_h':     round((agg['outage'] or 0) / 3600, 1),
            'dernier_mois': dernier_mois,
            'nb_excel':     nb_excel,
            'nb_pptx':      nb_pptx,
        })

    return render(request, 'reports/reporting.html', {
        'platforms': platforms_ctx,
    })


def reporting_platform(request, platform):
    """Page plateforme — liste des rapports Excel + PowerPoint + Import."""
    from .models import Incident
    from django.db.models import Count, Sum
    from django.urls import reverse

    cfg = REPORTING_PLATFORMS.get(platform)
    if not cfg:
        raise Http404('Plateforme inconnue')

    # Stats du dernier mois disponible
    qs = Incident.objects.filter(domain__in=cfg['domains'])
    dernier_mois = (
        qs.exclude(mois_rapport__isnull=True)
        .values_list('mois_rapport', flat=True)
        .order_by('-mois_rapport').first()
    )
    agg = qs.filter(mois_rapport=dernier_mois).aggregate(
        total=Count('id'), outage=Sum('duration_sec')
    ) if dernier_mois else {'total': 0, 'outage': 0}

    # Construire les URLs des rapports
    def _build_url(rep):
        try:
            return reverse(rep['url_name'], kwargs=rep.get('url_kwargs') or {})
        except Exception:
            return '#'

    excel_reports = []
    for rep in cfg.get('excel_reports', []):
        excel_reports.append({**rep, 'url': _build_url(rep)})

    pptx_reports = []
    for rep in cfg.get('pptx_reports', []):
        pptx_reports.append({**rep, 'url': _build_url(rep)})

    tool_reports = []
    for rep in cfg.get('tool_reports', []):
        tool_reports.append({**rep, 'url': _build_url(rep)})

    return render(request, 'reports/reporting_platform.html', {
        'platform':      platform,
        'cfg':           cfg,
        'excel_reports': excel_reports,
        'pptx_reports':  pptx_reports,
        'tool_reports':  tool_reports,
        'dernier_mois':  dernier_mois,
        'total':         agg['total'] or 0,
        'outage_h':      round((agg['outage'] or 0) / 3600, 1),
    })


def reporting_platform_import(request, platform):
    """Import de données lié à une plateforme spécifique."""
    from .reporting_config import PLATFORMS
    from .models import Incident

    cfg = PLATFORMS.get(platform)
    if not cfg:
        raise Http404('Plateforme inconnue')

    if request.method != 'POST':
        return redirect('reporting_platform', platform=platform)

    uploaded = request.FILES.get('incidents_file')
    if not uploaded:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('reporting_platform', platform=platform)

    ext = uploaded.name.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls'):
        messages.error(request, 'Format non supporté. Utilisez un fichier Excel (.xlsx).')
        return redirect('reporting_platform', platform=platform)

    import tempfile, os
    with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}') as tmp:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    source = uploaded.name
    clear_mois = request.POST.get('clear_mois') == '1'
    total_created = 0
    errors = []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]
        wb.close()

        # Détection automatique : fichier BASES DES INCIDENTS vs fichier brut
        is_bases_format = any(s in sheet_names for s in ['reseau mobile ', 'dr2', 'reseau fixe', 'transport', 'igw', 'core'])

        if is_bases_format:
            # Parseurs existants (format BASES DES INCIDENTS)
            from reports.management.commands.import_incidents import PARSERS
            domains_to_import = cfg['import']['domains']
            for domain in domains_to_import:
                if domain not in PARSERS:
                    continue
                try:
                    incidents = PARSERS[domain](tmp_path, source)
                except Exception as e:
                    errors.append(f'{domain}: {e}')
                    continue
                if not incidents:
                    continue
                mois = incidents[0].mois_rapport
                if clear_mois and mois:
                    Incident.objects.filter(domain=domain, mois_rapport=mois).delete()
                Incident.objects.bulk_create(incidents, batch_size=500)
                total_created += len(incidents)
        else:
            # Fichier brut ticketing (1 feuille, headers ligne 1)
            from .bases_incidents import parse_raw_mobile, _parse_generic_raw, _parse_dur_str, _parse_dt, _clean, _fmt_dur
            from reports.management.commands.import_incidents import _rows_to_incidents
            import pandas as pd

            domain = cfg['import']['domains'][0]

            if domain == 'mobile':
                rows = parse_raw_mobile(tmp_path, mois_filter=None)
                from django.utils import timezone as _tz

                def _to_aware(dt):
                    if dt is None:
                        return None
                    try:
                        return _tz.make_aware(dt, _tz.get_current_timezone())
                    except Exception:
                        return dt

                from reports.models import Incident as _Inc
                from datetime import date as _d
                incidents = []
                for r in rows:
                    at  = r.get('alarm_time')
                    ct  = r.get('cancel_time')
                    mois = _d(at.year, at.month, 1) if at else None
                    incidents.append(_Inc(
                        domain='mobile', mois_rapport=mois, source_file=source,
                        alarm_time=_to_aware(at), cancel_time=_to_aware(ct),
                        duration_sec=r.get('duration_sec'),
                        numero_ticket=r.get('numero_ticket', ''),
                        nature=r.get('nature', ''),
                        site_parent=r.get('site_parent', ''),
                        site_name=r.get('site_name', ''),
                        site_id=r.get('site_id', ''),
                        region=r.get('region', ''),
                        base=r.get('base', ''),
                        impact_equipement=r.get('impact_equipement', ''),
                        impact_service=r.get('impact_service', ''),
                        plateforme=r.get('plateforme', ''),
                        technologies=r.get('technologies', ''),
                        escalade=r.get('escalade', ''),
                        cause=r.get('cause', ''),
                        root_cause=r.get('root_cause', ''),
                        action=r.get('action', ''),
                        technicien_informe=r.get('technicien_informe', ''),
                        technicien_maint=r.get('technicien_maint', ''),
                        point_bloquant=r.get('point_bloquant', ''),
                        observation=r.get('observation', ''),
                        status=r.get('status', ''),
                    ))
                if incidents:
                    mois_val = incidents[0].mois_rapport
                    if clear_mois and mois_val:
                        Incident.objects.filter(domain='mobile', mois_rapport=mois_val).delete()
                    Incident.objects.bulk_create(incidents, batch_size=500, ignore_conflicts=False)
                    total_created = len(incidents)
            else:
                errors.append(f'Import brut non supporté pour la plateforme « {platform} » — utilisez le format Bases des Incidents.')
    finally:
        os.unlink(tmp_path)

    if errors:
        messages.warning(request, f'{total_created} incidents importés. Erreurs : ' + ' | '.join(errors))
    else:
        messages.success(request, f'{total_created} incidents importés depuis « {source} ».')

    return redirect('reporting_platform', platform=platform)


def generate_pptx_platform(request, platform):
    """Génère le rapport PowerPoint pour une plateforme (redirige vers generate_pptx_report)."""
    from django.urls import reverse
    return redirect(reverse('generate_pptx') + f'?platform={platform}')


DR2_REGION_TARGETS = {
    'LOME':     372,
    'MARITIME': 182,
    'PLATEAUX': 159,
    'CENTRALE': 134,
    'KARA':     161,
    'SAVANES':  118,
}

DR2_ESCALADE_ORDER = [
    'ENERGIE', 'RAN-FIELD O', 'TRANS FH-FIELD O', 'TRANS IP',
    'TRANS FO', 'TRANS FTTM', 'PROJET', 'BSS', 'INFRA',
]


def _parse_dr2_dates(request):
    """Parse debut/fin depuis GET params, avec défaut = mois dernier."""
    today = date.today()
    last_month_first = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    debut_str = request.GET.get('debut', '')
    fin_str   = request.GET.get('fin', '')
    try:
        debut = date.fromisoformat(debut_str) if debut_str else last_month_first
    except ValueError:
        debut = last_month_first
    try:
        fin = date.fromisoformat(fin_str) if fin_str else today
    except ValueError:
        fin = today
    return debut, fin


def _build_dr2_data(debut, fin):
    """Construit le dataset DR2 pour une période. Partagé entre view et export."""
    from .models import Incident
    from django.utils import timezone
    import datetime

    today = date.today()
    period_days = (fin - debut).days + 1

    debut_dt = timezone.make_aware(datetime.datetime.combine(debut, datetime.time.min))
    fin_dt   = timezone.make_aware(datetime.datetime.combine(fin,   datetime.time.max))

    qs = Incident.objects.filter(domain='dr2', alarm_time__gte=debut_dt, alarm_time__lte=fin_dt)
    total_dr2 = qs.count()
    moyenne   = round(total_dr2 / period_days, 2) if period_days else 0

    yesterday = today - timedelta(days=1)
    y_s = timezone.make_aware(datetime.datetime.combine(yesterday, datetime.time.min))
    y_e = timezone.make_aware(datetime.datetime.combine(yesterday, datetime.time.max))
    nbre_j1 = Incident.objects.filter(domain='dr2', alarm_time__gte=y_s, alarm_time__lte=y_e).count()

    db_escs = list(qs.exclude(escalade='').values_list('escalade', flat=True).distinct())
    escalade_vals = [e for e in DR2_ESCALADE_ORDER if e in db_escs]
    for e in db_escs:
        if e not in escalade_vals:
            escalade_vals.append(e)

    region_rows = []
    for region in DR2_REGION_TARGETS:
        rqs      = qs.filter(region__iexact=region)
        dr2      = rqs.count()
        tget     = DR2_REGION_TARGETS[region]
        pct_reg  = round(dr2 / total_dr2 * 100) if total_dr2 else 0
        pct_tget = round(dr2 / tget * 100)       if tget else 0
        metier   = rqs.exclude(cancel_time__isnull=True).count()
        color    = 'red' if pct_tget >= 100 else ('yellow' if pct_tget >= 70 else 'green')
        region_rows.append({
            'region':    region,
            'tget':      tget,
            'dr2':       dr2,
            'metier':    metier,
            'pct_reg':   pct_reg,
            'pct_tget':  pct_tget,
            'escalades': {e: rqs.filter(escalade=e).count() for e in escalade_vals},
            'color':     color,
        })

    metier_qs   = qs.exclude(cancel_time__isnull=True)
    metier_tots = {e: metier_qs.filter(escalade=e).count() for e in escalade_vals}
    total_metier = sum(metier_tots.values())

    esc_stats = []
    for e in escalade_vals:
        tot    = qs.filter(escalade=e).count()
        metier = metier_tots.get(e, 0)
        esc_stats.append({
            'escalade': e,
            'total':    tot,
            'metier':   metier,
            'pct_tgt':  round(metier / tot * 100) if tot else None,
            'pct_tt':   round(metier / total_dr2 * 100) if total_dr2 else 0,
        })

    return {
        'debut':         debut,
        'fin':           fin,
        'period_days':   period_days,
        'total_dr2':     total_dr2,
        'total_tget':    sum(DR2_REGION_TARGETS.values()),
        'moyenne':       moyenne,
        'nbre_j1':       nbre_j1,
        'escalade_vals': escalade_vals,
        'region_rows':   region_rows,
        'esc_stats':     esc_stats,
        'total_metier':  total_metier,
    }


def dr2_daily_report(request):
    debut, fin = _parse_dr2_dates(request)
    ctx = _build_dr2_data(debut, fin)
    return render(request, 'reports/dr2_daily.html', ctx)


def dr2_daily_export(request):
    """Génère l'export Excel du DR2 Daily Report."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    debut, fin = _parse_dr2_dates(request)
    d = _build_dr2_data(debut, fin)

    wb = Workbook()
    ws = wb.active
    ws.title = 'DR2 Daily Report'

    # ── Couleurs ──────────────────────────────────────────────────────────────
    C_BLUE_DARK  = 'FF003087'
    C_BLUE_MED   = 'FF1E3A6E'
    C_BLUE_LIGHT = 'FF2A4A80'
    C_YELLOW     = 'FFFFC72C'
    C_WHITE      = 'FFFFFFFF'
    C_GREEN_BG   = 'FFC6EFCE';  C_GREEN_FG   = 'FF276221'
    C_YELLOW_BG  = 'FFFFEB9C';  C_YELLOW_FG  = 'FF9C6500'
    C_RED_BG     = 'FFFFC7CE';  C_RED_FG     = 'FF9C0006'
    C_METIER_BG  = 'FFF0F4FF'
    C_PCT_BG     = 'FFFFF7ED'
    C_PCT2_BG    = 'FFF0FDF4'
    C_GRAY       = 'FFF5F5F5'

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def font(bold=False, color=C_WHITE, size=11):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    def align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(border_style='thin', color='FFD8E0F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    escs = d['escalade_vals']
    n_esc = len(escs)
    # Columns: [Region, TGET, DR2, %REG, %TGET, STAT_label] + [esc×N]
    n_fixed = 6
    total_cols = n_fixed + n_esc

    # ── Largeurs de colonnes ──────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 4   # STAT vertical
    for i in range(n_esc):
        col = get_column_letter(n_fixed + 1 + i)
        ws.column_dimensions[col].width = 16

    # ── Hauteurs de lignes ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22   # titre
    ws.row_dimensions[2].height = 16   # période
    ws.row_dimensions[3].height = 40   # entêtes (wrap)
    ws.row_dimensions[4].height = 18   # tgets escalades

    def _set_row(row_idx, values, bg, fg_color, bold=True, h_align='center', row_height=None):
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill    = fill(bg)
            cell.font    = font(bold=bold, color=fg_color)
            cell.alignment = align(h=h_align)
            cell.border  = border
        if row_height:
            ws.row_dimensions[row_idx].height = row_height

    # ── Ligne 1 : Titre ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value='DR2 / DAILY REPORT')
    c.fill      = fill(C_BLUE_DARK)
    c.font      = Font(bold=True, color=C_WHITE, size=14, name='Calibri')
    c.alignment = align()
    c.border    = border

    # ── Ligne 2 : Période ─────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    period_str = f"Du {d['debut'].strftime('%d/%m/%Y')} au {d['fin'].strftime('%d/%m/%Y')}"
    c = ws.cell(row=2, column=1, value=period_str)
    c.fill      = fill(C_BLUE_DARK)
    c.font      = Font(bold=True, color=C_YELLOW, size=11, name='Calibri')
    c.alignment = align()
    c.border    = border

    # ── Ligne 3 : En-têtes colonnes ───────────────────────────────────────────
    headers_left = ['Région', 'TGET', 'DR2 REG', '% REG', '% TGET\nREG', 'STAT.']
    for c_idx, hdr in enumerate(headers_left, 1):
        cell = ws.cell(row=3, column=c_idx, value=hdr)
        cell.fill      = fill(C_BLUE_MED)
        cell.font      = Font(bold=True, color=C_WHITE, size=10, name='Calibri')
        cell.alignment = align(wrap=True)
        cell.border    = border
    for i, esc in enumerate(escs):
        cell = ws.cell(row=3, column=n_fixed + 1 + i, value=esc)
        cell.fill      = fill(C_BLUE_MED)
        cell.font      = Font(bold=True, color=C_WHITE, size=10, name='Calibri')
        cell.alignment = align(wrap=True)
        cell.border    = border

    # ── Ligne 4 : Totaux TGET par escalade ───────────────────────────────────
    for c_idx in range(1, n_fixed + 1):
        cell = ws.cell(row=4, column=c_idx, value='')
        cell.fill   = fill(C_BLUE_LIGHT)
        cell.border = border
    for i, stat in enumerate(d['esc_stats']):
        cell = ws.cell(row=4, column=n_fixed + 1 + i, value=stat['total'])
        cell.fill      = fill(C_BLUE_LIGHT)
        cell.font      = Font(bold=True, color=C_WHITE, size=11, name='Calibri')
        cell.alignment = align()
        cell.border    = border

    # ── Lignes par région (à partir de la ligne 5) ────────────────────────────
    COLOR_MAP = {
        'green':  (C_GREEN_BG,  C_GREEN_FG),
        'yellow': (C_YELLOW_BG, C_YELLOW_FG),
        'red':    (C_RED_BG,    C_RED_FG),
    }
    row_idx = 5
    for row in d['region_rows']:
        bg_c, fg_c = COLOR_MAP[row['color']]
        ws.row_dimensions[row_idx].height = 18

        cells_left = [
            row['region'],
            row['tget'],
            row['dr2'],
            f"{row['pct_reg']}%",
            f"{row['pct_tget']}%",
            '',
        ]
        for c_idx, val in enumerate(cells_left, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            if c_idx in (3, 5):
                cell.fill = fill(bg_c)
                cell.font = Font(bold=True, color=fg_c, size=11, name='Calibri')
            elif c_idx == 1:
                cell.font = Font(bold=True, color='FF1E3A6E', size=11, name='Calibri')
                cell.fill = fill('FFF8FAFF')
            else:
                cell.font = Font(size=11, name='Calibri')
                cell.fill = fill('FFFFFFFF')
            cell.alignment = align(h='left' if c_idx == 1 else 'center')
            cell.border    = border

        for i, esc in enumerate(escs):
            val  = row['escalades'].get(esc, 0)
            cell = ws.cell(row=row_idx, column=n_fixed + 1 + i, value=val if val else 0)
            cell.fill      = fill('FFFFFFFF' if val else C_GRAY)
            cell.font      = Font(bold=bool(val), size=11, name='Calibri', color='FF1E3A6E' if val else 'FFBBBBBB')
            cell.alignment = align()
            cell.border    = border

        row_idx += 1

    # ── Ligne TOTAL DR2 ───────────────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 20
    total_pct = round(d['total_dr2'] / d['total_tget'] * 100) if d['total_tget'] else 0
    _set_row(row_idx,
             ['TOTAL DR2', d['total_tget'], d['total_dr2'], '100%', f"{total_pct}%", ''] +
             [s['total'] for s in d['esc_stats']],
             C_BLUE_DARK, C_WHITE, bold=True, row_height=20)
    row_idx += 1

    # ── Ligne MOYENNE DR2 ─────────────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    avg_cells = ['MOYENNE DR2', '', str(d['moyenne']), '', '% TGET/MÉTIER', '']
    for i, stat in enumerate(d['esc_stats']):
        if stat['pct_tgt'] is None:
            avg_cells.append('#DIV/0!')
        else:
            avg_cells.append(f"{stat['pct_tgt']}%")
    _set_row(row_idx, avg_cells, 'FF0047CC', C_WHITE, bold=True, row_height=18)
    row_idx += 1

    # ── Ligne DR2/MÉTIER ──────────────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    _set_row(row_idx,
             ['DR2 / MÉTIER', '', '', '', '', ''] + [s['metier'] for s in d['esc_stats']],
             C_METIER_BG, 'FF1E3A6E', bold=True, row_height=18)
    row_idx += 1

    # ── Ligne %MÉTIER/TT DR2 ─────────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    _set_row(row_idx,
             ['% MÉTIER / TT DR2', '', '', '', '', ''] + [f"{s['pct_tt']}%" for s in d['esc_stats']],
             C_PCT2_BG, 'FF1E3A6E', bold=True, row_height=18)

    # ── Réponse HTTP ──────────────────────────────────────────────────────────
    filename = f"DR2_Daily_{d['debut'].strftime('%Y%m%d')}_{d['fin'].strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_network_section(domain_filter, mois_sel=None):
    """Construit les stats (synth, top_sites, top_causes, top_regions, statuts) pour un domaine."""
    from .models import Incident
    from django.db.models import Count, Sum

    qs_all = Incident.objects.filter(**domain_filter)

    # Tous les mois disponibles pour ce domaine (non filtrés) → dropdown
    mois_list = list(dict.fromkeys(
        qs_all.exclude(mois_rapport__isnull=True)
        .values_list('mois_rapport', flat=True)
        .distinct().order_by('-mois_rapport')
    ))

    # Mois effectif : explicite ou plus récent
    mois_effectif = mois_sel or (mois_list[0] if mois_list else None)

    # qs filtré sur le mois effectif
    qs = qs_all.filter(mois_rapport=mois_effectif) if mois_effectif else qs_all

    synth_rows, total_nb, total_duree = [], 0, 0
    for row in (qs.exclude(escalade='').values('escalade')
                  .annotate(nb=Count('id'), duree_sec=Sum('duration_sec'))
                  .order_by('-nb')):
        nb    = row['nb']
        duree = row['duree_sec'] or 0
        ouvert = qs.filter(escalade=row['escalade'], status__iexact='OUVERT').count()
        synth_rows.append({
            'escalade': row['escalade'],
            'nb': nb,
            'duree': _fmt_sec(duree),
            'mttr':  _fmt_sec(duree / nb if nb else 0),
            'outage': _fmt_sec(duree),
            'status': f"{ouvert} Non résolu" if ouvert else 'Résolu',
            'ouvert': ouvert,
        })
        total_nb += nb; total_duree += duree

    top_sites = list(
        qs.exclude(site_name='').values('site_name', 'region')
        .annotate(nb=Count('id'), duree=Sum('duration_sec')).order_by('-nb')[:15]
    )
    for s in top_sites:
        s['duree_h'] = round((s['duree'] or 0) / 3600, 1)

    top_causes  = list(qs.exclude(cause='').values('cause').annotate(nb=Count('id')).order_by('-nb')[:10])
    top_regions = list(qs.exclude(region='').values('region').annotate(nb=Count('id')).order_by('-nb')[:8])
    statuts     = list(qs.exclude(status='').values('status').annotate(nb=Count('id')).order_by('-nb'))

    open_cnt = sum(s['nb'] for s in statuts if s['status'].upper() == 'OUVERT')

    return {
        'mois_list':    mois_list,
        'mois_sel':     mois_effectif,
        'synth_rows':   synth_rows,
        'total_row': {
            'nb':     total_nb,
            'duree':  _fmt_sec(total_duree),
            'mttr':   _fmt_sec(total_duree / total_nb if total_nb else 0),
            'outage': _fmt_sec(total_duree),
            'ouvert': open_cnt,
        },
        'top_sites':    top_sites,
        'top_causes':   top_causes,
        'top_regions':  top_regions,
        'statuts':      statuts,
        'total_all':    qs.count(),
    }


def reporting_network(request, platform):
    from .models import Incident
    from django.db.models import Count, Sum

    meta = REPORTING_NETWORKS.get(platform)
    if not meta:
        raise Http404('Réseau de reporting inconnu')

    domains = meta['domains']

    # Mois sélectionné via GET (optionnel)
    mois_sel_str = request.GET.get('mois', '')
    mois_sel = None
    if mois_sel_str:
        try:
            from datetime import datetime as _dt
            mois_sel = _dt.strptime(mois_sel_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # ── Cas spécial : mobile-dr2 → deux sections indépendantes ───────────────
    if platform == 'mobile-dr2':
        mobile_data = _build_network_section({'domain': 'mobile'}, mois_sel)
        dr2_data    = _build_network_section({'domain': 'dr2'}, mois_sel)
        return render(request, 'reports/reporting_network.html', {
            'network':        platform,
            'platform':       platform,
            'network_label':  meta['label'],
            'network_icon':   meta['icon'],
            'is_mobile_dr2':  True,
            'mobile':         mobile_data,
            'dr2':            dr2_data,
        })

    # ── Cas standard : un seul domaine ───────────────────────────────────────
    section_std = _build_network_section({'domain__in': domains}, mois_sel)
    return render(request, 'reports/reporting_network.html', {
        'network':       platform,
        'platform':      platform,
        'network_label': meta['label'],
        'network_icon':  meta['icon'],
        'mois_list':     section_std['mois_list'],
        'mois_sel':      section_std['mois_sel'],
        'section_std':   section_std,
        'is_mobile_dr2': False,
    })


def site_info(request):
    from .models import Site
    query = request.GET.get('q', '').strip()
    site  = None
    results = []

    if query:
        exact = Site.objects.filter(site_name__iexact=query).first()
        if exact:
            site = exact
        else:
            results = list(
                Site.objects.filter(site_name__icontains=query)
                .values('site_name', 'site_id', 'region')[:20]
            )

    return render(request, 'reports/site_info.html', {
        'query':   query,
        'site':    site,
        'results': results,
    })


def site_search_api(request):
    from .models import Site
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse([], safe=False)
    hits = list(
        Site.objects.filter(site_name__icontains=q)
        .values('site_name', 'site_id', 'region')[:15]
    )
    return JsonResponse(hits, safe=False)


# ── Import API manuel ─────────────────────────────────────────────────────────

def api_import_view(request):
    """Récupère les données API, sauvegarde en Excel et redirige vers process_report."""
    import datetime as _dt
    from django.utils import timezone as _tz
    from .api_import import fetch_and_save_api

    VALID_NETWORKS = {"mobile", "fixe", "transmission", "core", "all"}

    if request.method == 'POST':
        date_debut = request.POST.get('date_debut', '').strip()
        date_fin   = request.POST.get('date_fin', '').strip()
        network    = request.POST.get('network', 'mobile').strip().lower()
        if network not in VALID_NETWORKS:
            network = 'mobile'

        # Défaut : hier 00:00 → aujourd'hui 23:59
        if not date_debut or not date_fin:
            now        = _tz.now()
            date_fin   = now.strftime('%Y-%m-%dT23:59')
            date_debut = (now - _dt.timedelta(days=1)).strftime('%Y-%m-%dT00:00')

        # Normalise : extrait juste la partie date
        try:
            date_debut_d = _dt.datetime.fromisoformat(date_debut).strftime('%Y-%m-%d')
            date_fin_d   = _dt.datetime.fromisoformat(date_fin).strftime('%Y-%m-%d')
        except ValueError:
            messages.error(request, "Format de date invalide.")
            return redirect('/upload/?tab=api')

        try:
            report = fetch_and_save_api(date_debut_d, date_fin_d, user=request.user, network=network)
            return redirect('process_report', pk=report.pk)
        except Exception as exc:
            messages.error(request, f"Erreur import API : {exc}")
            return redirect('/upload/?tab=api')

    return redirect('/upload/?tab=api')


# ── Audit admin ───────────────────────────────────────────────────────────────

def audit_view(request):
    """Audit admin : ré-importe une période depuis l'API (mois par mois si > 1 mois)."""
    if not request.user.is_superuser:
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('home')

    if request.method == 'POST':
        from .api_import import run_import_months
        import datetime as _dt

        date_debut = request.POST.get('date_debut', '').strip()
        date_fin   = request.POST.get('date_fin', '').strip()
        overwrite  = request.POST.get('overwrite') == '1'

        if not date_debut or not date_fin:
            messages.error(request, "Les deux dates sont obligatoires.")
            return redirect('audit')

        try:
            _dt.date.fromisoformat(date_debut)
            _dt.date.fromisoformat(date_fin)
        except ValueError:
            messages.error(request, "Format de date invalide (YYYY-MM-DD attendu).")
            return redirect('audit')

        if date_fin < date_debut:
            messages.error(request, "La date de fin doit être >= la date de début.")
            return redirect('audit')

        try:
            result = run_import_months(date_debut, date_fin, triggered_by=request.user, overwrite=overwrite)
            msg_parts = []
            if result['created']:
                msg_parts.append(f"{result['created']} rapport(s) importé(s)")
            if result['skipped']:
                msg_parts.append(f"{result['skipped']} ignoré(s) (déjà existants)")
            if result['errors']:
                msg_parts.append(f"{len(result['errors'])} erreur(s)")
                messages.error(request, f"Audit terminé avec erreurs : {'; '.join(result['errors'][:3])}")
            if msg_parts:
                messages.success(request, "Audit terminé — " + ", ".join(msg_parts) + ".")
            else:
                messages.info(request, "Audit terminé — aucune donnée retournée.")
        except Exception as exc:
            messages.error(request, f"Erreur durant l'audit : {exc}")

        return redirect('audit')


# ─────────────────────────────────────────────────────────────────────────────
# Génération du rapport PowerPoint
# ─────────────────────────────────────────────────────────────────────────────

def generate_pptx_report(request):
    """Génère et télécharge le rapport PowerPoint Comité GDI."""
    from datetime import date
    from django.http import HttpResponse
    from .pptx_report import generate_report

    # Récupère le mois demandé (GET ?mois=2026-05-01)
    mois_str = request.GET.get('mois', '')
    mois = None
    if mois_str:
        try:
            mois = date.fromisoformat(mois_str)
        except ValueError:
            pass

    # Si pas de mois fourni, cherche le plus récent disponible pour mobile
    if not mois:
        from .models import Incident
        mois = (
            Incident.objects.filter(domain='mobile')
            .exclude(mois_rapport__isnull=True)
            .order_by('-mois_rapport')
            .values_list('mois_rapport', flat=True)
            .first()
        )

    generated_on = date.today().strftime('%d/%m/%Y')

    buf = generate_report(
        mois_mobile=mois,
        mois_fixe=mois,
        mois_transport=mois,
        mois_igw=mois,
        mois_core=mois,
        generated_on=generated_on,
    )

    label = mois.strftime('%Y-%m') if mois else 'rapport'
    filename = f'GDI_{label}.pptx'

    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# Bases des Incidents — export multi-onglets
# ─────────────────────────────────────────────────────────────────────────────

def bases_incidents_view(request):
    """Page de génération du fichier Bases des Incidents."""
    from datetime import date
    from .models import Incident

    mois_list = list(dict.fromkeys(
        Incident.objects.exclude(mois_rapport__isnull=True)
        .values_list('mois_rapport', flat=True)
        .distinct().order_by('-mois_rapport')
    ))

    mois_sel_str = request.GET.get('mois', '')
    mois_sel = None
    if mois_sel_str:
        try:
            mois_sel = date.fromisoformat(mois_sel_str)
        except ValueError:
            pass
    if not mois_sel and mois_list:
        mois_sel = mois_list[0]

    return render(request, 'reports/bases_incidents.html', {
        'mois_list': mois_list,
        'mois_sel': mois_sel,
    })


def bases_incidents_export(request):
    """Génère et télécharge le fichier Bases des Incidents."""
    from datetime import date
    from django.http import HttpResponse
    from .bases_incidents import generate_bases_incidents

    if request.method == 'POST':
        mois_str = request.POST.get('mois', '')
    else:
        mois_str = request.GET.get('mois', '')

    mois = None
    if mois_str:
        try:
            mois = date.fromisoformat(mois_str)
        except ValueError:
            pass

    if not mois:
        from .models import Incident
        mois = (
            Incident.objects.exclude(mois_rapport__isnull=True)
            .order_by('-mois_rapport')
            .values_list('mois_rapport', flat=True)
            .first()
        )
        if not mois:
            return HttpResponse('Aucune donnée disponible.', status=400)

    # Fichiers uploadés (optionnels)
    import tempfile, os

    def _save_tmp(f):
        if not f:
            return None
        ext = f.name.rsplit('.', 1)[-1].lower()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        for chunk in f.chunks():
            tmp.write(chunk)
        tmp.close()
        return tmp.name

    files = {}
    tmp_paths = []
    for domain in ('mobile', 'fixe', 'transport', 'igw', 'core'):
        f = request.FILES.get(f'file_{domain}')
        path = _save_tmp(f)
        files[domain] = path
        if path:
            tmp_paths.append(path)

    try:
        buf, nb_mobile, nb_dr2 = generate_bases_incidents(
            mois=mois,
            mobile_file=files.get('mobile'),
            fixe_file=files.get('fixe'),
            transport_file=files.get('transport'),
            igw_file=files.get('igw'),
            core_file=files.get('core'),
        )
    finally:
        for p in tmp_paths:
            try:
                os.unlink(p)
            except Exception:
                pass

    from calendar import month_name
    _MOIS_FR = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE',
    }
    label = f'{_MOIS_FR.get(mois.month, str(mois.month))}_{mois.year}'
    filename = f'BASES_INCIDENTS_{label}.xlsx'

    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# Bases des Incidents — par plateforme
# ─────────────────────────────────────────────────────────────────────────────

def platform_bases_incidents(request, platform):
    """Page de génération Bases des Incidents pour une plateforme donnée."""
    from datetime import date as date_
    from .reporting_config import PLATFORMS as REPORTING_PLATFORMS
    from .models import Incident

    cfg = REPORTING_PLATFORMS.get(platform)
    if not cfg:
        raise Http404

    # Domaines DB à interroger (dr2 n'est pas stocké séparément, dérivé de mobile)
    domains = [d for d in cfg.get('domains', []) if d != 'dr2']
    if not domains:
        domains = ['mobile']

    mois_list = list(dict.fromkeys(
        Incident.objects.filter(domain__in=domains)
        .exclude(mois_rapport__isnull=True)
        .values_list('mois_rapport', flat=True)
        .distinct().order_by('-mois_rapport')
    ))

    mois_sel = mois_list[0] if mois_list else date_.today().replace(day=1)

    SHEETS_INFO = {
        'mobile-dr2':  {'sheets': ['Réseau Mobile', 'DR2'], 'hint': 'RESEAU_MOBILE_*.xlsx'},
        'fixe':        {'sheets': ['Réseau Fixe'],           'hint': 'RESEAU_FIXE_*.xlsx'},
        'transmission': {'sheets': ['Transport'],            'hint': 'TRANSPORT_*.xlsx'},
        'igw':         {'sheets': ['IGW'],                   'hint': 'IGW_*.xlsx'},
        'core':        {'sheets': ['Core'],                  'hint': 'CORE_*.xlsx'},
    }
    info = SHEETS_INFO.get(platform, {'sheets': [], 'hint': '*.xlsx'})

    return render(request, 'reports/platform_bases_incidents.html', {
        'platform': platform,
        'cfg': cfg,
        'mois_list': mois_list,
        'mois_sel': mois_sel,
        'sheets': info['sheets'],
        'file_hint': info['hint'],
    })


def platform_bases_incidents_export(request, platform):
    """Génère et télécharge les Bases des Incidents pour une plateforme."""
    from datetime import date as date_
    import tempfile, os
    from django.http import HttpResponse
    from .reporting_config import PLATFORMS as REPORTING_PLATFORMS
    from .bases_incidents import generate_platform_bases_incidents, _MOIS_FR

    if request.method != 'POST':
        return redirect('platform_bases_incidents', platform=platform)

    cfg = REPORTING_PLATFORMS.get(platform)
    if not cfg:
        raise Http404

    # Mois sélectionné
    mois_str = request.POST.get('mois', '')
    try:
        mois = date_.fromisoformat(mois_str).replace(day=1)
    except Exception:
        from .models import Incident
        domains = [d for d in cfg.get('domains', []) if d != 'dr2'] or ['mobile']
        mois = (
            Incident.objects.filter(domain__in=domains)
            .exclude(mois_rapport__isnull=True)
            .order_by('-mois_rapport')
            .values_list('mois_rapport', flat=True)
            .first()
        ) or date_.today().replace(day=1)

    # Fichier source uploadé (optionnel)
    source_file = None
    tmp_path = None
    uploaded = request.FILES.get('source_file')
    if uploaded:
        ext = uploaded.name.rsplit('.', 1)[-1].lower()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp.close()
        tmp_path = tmp.name
        source_file = tmp_path

    try:
        buf, nb = generate_platform_bases_incidents(platform, mois, source_file)
    except Exception as e:
        messages.error(request, f'Erreur lors de la génération : {e}')
        return redirect('platform_bases_incidents', platform=platform)
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    if buf is None:
        messages.error(request, 'Plateforme non reconnue.')
        return redirect('platform_bases_incidents', platform=platform)

    label = f'{_MOIS_FR.get(mois.month, str(mois.month))}_{mois.year}'
    plat_slug = platform.upper().replace('-', '_')
    filename = f'BASES_INCIDENTS_{plat_slug}_{label}.xlsx'

    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ── Outils interactifs ────────────────────────────────────────────────────────

def igw_rapport_noc(request):
    from .reporting_config import PLATFORMS
    cfg = PLATFORMS['igw']
    return render(request, 'reports/igw_rapport_noc.html', {'cfg': cfg, 'platform': 'igw'})


def igw_trafic_international(request):
    from .reporting_config import PLATFORMS
    cfg = PLATFORMS['igw']
    return render(request, 'reports/igw_trafic_international.html', {'cfg': cfg, 'platform': 'igw'})


def transport_rapport_noc(request):
    from .reporting_config import PLATFORMS
    cfg = PLATFORMS['transmission']
    return render(request, 'reports/transport_rapport_noc.html', {'cfg': cfg, 'platform': 'transmission'})


def transport_rapport_fo(request):
    from .reporting_config import PLATFORMS
    cfg = PLATFORMS['transmission']
    return render(request, 'reports/transport_rapport_fo.html', {'cfg': cfg, 'platform': 'transmission'})


def fixe_rapport_ftth(request):
    from .reporting_config import PLATFORMS
    cfg = PLATFORMS['fixe']
    return render(request, 'reports/fixe_rapport_ftth.html', {'cfg': cfg, 'platform': 'fixe'})