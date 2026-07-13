"""Traitement automatique « SITE DOWN » (micro-coupures réseau mobile).

Version intégrée du script autonome ``site_down_auto.py`` :

- Les fichiers d'alarmes NetAct proviennent soit du partage réseau ISOC
  (collecte automatique planifiée), soit d'un upload manuel via la page web.
- Les régions sont lues depuis le modèle :class:`reports.models.Site`
  (plus besoin de ``BASE TECH.csv``).
- Les Cause / Escalade sont lues depuis le modèle
  :class:`reports.models.Incident` (domaine mobile) — plus besoin de relire
  les rapports journaliers Excel.
- Les alarmes traitées sont stockées dans :class:`reports.models.SiteDownAlarm`
  (contrainte unique site + alarm_time → pas de doublons).
- Un fichier Excel consolidé ``SITE_DOWN_AAAA-MM.xlsx`` est produit par mois
  dans ``MEDIA_ROOT/site_down/traites/`` (onglet Données + onglet Cumul).
"""

import logging
import os
import re
import shutil
from datetime import datetime, time

import pandas as pd
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Répertoires de travail (sous MEDIA_ROOT)
# ──────────────────────────────────────────────────────────────────────────────
def _base_dir():
    return os.path.join(settings.MEDIA_ROOT, 'site_down')

def folder_a_traiter():
    return os.path.join(_base_dir(), 'a_traiter')

def folder_traites():
    return os.path.join(_base_dir(), 'traites')

def folder_erreurs():
    return os.path.join(_base_dir(), 'erreurs')

def _journal_alarmes():
    return os.path.join(_base_dir(), '.journal_alarmes.txt')

def ensure_dirs():
    for d in (folder_a_traiter(), folder_traites(), folder_erreurs()):
        os.makedirs(d, exist_ok=True)


# ──────────────────────────────────────────────────────────────────────────────
# Constantes
# ──────────────────────────────────────────────────────────────────────────────
ALARM_FILTER = 'WCDMA BASE STATION OUT OF USE'

NOMS_MOIS_COURTS = {
    '01': 'janv', '02': 'févr', '03': 'mars', '04': 'avr',
    '05': 'mai',  '06': 'juin', '07': 'juil', '08': 'août',
    '09': 'sept', '10': 'oct',  '11': 'nov',  '12': 'déc',
}
NOMS_MOIS_COMPLETS = {
    '01': 'Janvier', '02': 'Février', '03': 'Mars',      '04': 'Avril',
    '05': 'Mai',     '06': 'Juin',    '07': 'Juillet',   '08': 'Août',
    '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre',
}
NOMS_MOIS_DOSSIER = {
    1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
    5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
    9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE',
}

JAUNE_HEADER     = 'FFFF00'
NOIR_TEXTE       = '000000'
GRIS_LIGNE_PAIRE = 'F2F2F2'
VERT_TOTAL       = '92D050'
ORANGE_TOTAL     = 'FFA500'
ROUGE_TOTAL      = 'FF4040'
BLEU_FONCE       = '003087'
JAUNE_YAS        = 'FFC72C'

BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)


# ──────────────────────────────────────────────────────────────────────────────
# Collecte réseau (facultative — dépend de SITE_DOWN_NETWORK_BASES)
# ──────────────────────────────────────────────────────────────────────────────
def _network_bases():
    """Liste des racines réseau candidates (ex: \\\\10.228.15.80\\isoc)."""
    return [b for b in getattr(settings, 'SITE_DOWN_NETWORK_BASES', []) if b]


def _source_alarmes_candidates():
    now = datetime.now()
    mois, annee = NOMS_MOIS_DOSSIER[now.month], now.year
    return [
        os.path.join(base, 'RAPPORT RESEAU MOBILE', 'ALARME NETACT SITES DOWN',
                     f'SITE DOWN {mois} {annee}')
        for base in _network_bases()
    ]


def _lire_journal(path):
    if not os.path.exists(path):
        return set()
    with open(path, 'r', encoding='utf-8') as f:
        return {l.strip() for l in f if l.strip()}


def _ecrire_journal(path, noms):
    with open(path, 'a', encoding='utf-8') as f:
        for nom in noms:
            f.write(nom + '\n')


_DATE_PATTERNS = [
    # JJ-MM-AAAA (séparateurs - _ . ou espace)
    (re.compile(r'(\d{2})[-_. ](\d{2})[-_. ](\d{4})'), lambda m: (m.group(3), m.group(2), m.group(1))),
    # AAAA-MM-JJ
    (re.compile(r'(\d{4})[-_. ](\d{2})[-_. ](\d{2})'), lambda m: (m.group(1), m.group(2), m.group(3))),
    # JJ-MM-AA (année sur 2 chiffres)
    (re.compile(r'(\d{2})[-_. ](\d{2})[-_. ](\d{2})(?!\d)'), lambda m: ('20' + m.group(3), m.group(2), m.group(1))),
]


def extraire_date(filename):
    """Extrait (AAAA, MM, JJ) du nom de fichier, ou None."""
    for pattern, extract in _DATE_PATTERNS:
        m = pattern.search(filename)
        if m:
            annee, mois, jour = extract(m)
            if 1 <= int(mois) <= 12 and 1 <= int(jour) <= 31:
                return annee, mois, jour
    return None


def extraire_mois_annee(filename):
    """`... 05-07-2026 ...` → ``2026-07`` (None si pas de date dans le nom)."""
    d = extraire_date(filename)
    return f"{d[0]}-{d[1]}" if d else None


def _dates_deja_traitees_cache():
    """Dates (YYYY-MM-DD) déjà présentes dans les fichiers consolidés, par mois."""
    cache = {}

    def get(mois_annee):
        if mois_annee in cache:
            return cache[mois_annee]
        path = os.path.join(folder_traites(), f'SITE_DOWN_{mois_annee}.xlsx')
        dates = set()
        if os.path.exists(path):
            try:
                df = pd.read_excel(path, sheet_name='Données', usecols=['Alarm Time'])
                df['Alarm Time'] = pd.to_datetime(df['Alarm Time'], errors='coerce')
                dates = set(df['Alarm Time'].dt.strftime('%Y-%m-%d').dropna().unique())
            except Exception:
                logger.warning("site_down : lecture dates existantes %s échouée", path,
                               exc_info=True)
        cache[mois_annee] = dates
        return dates

    return get


def collecter_alarmes():
    """Copie les nouveaux fichiers d'alarmes du partage réseau vers ``a_traiter``.

    Returns:
        int: nombre de fichiers copiés (0 si le réseau est inaccessible).
    """
    ensure_dirs()
    source = next((s for s in _source_alarmes_candidates() if os.path.exists(s)), None)
    if source is None:
        logger.warning("site_down : source alarmes inaccessible (%s)",
                       _source_alarmes_candidates())
        return 0

    deja_copies = _lire_journal(_journal_alarmes())
    dates_traitees = _dates_deja_traitees_cache()

    fichiers = [
        f for f in os.listdir(source)
        if os.path.isfile(os.path.join(source, f))
        and f.lower().endswith(('.xlsx', '.xls', '.csv'))
        and not f.startswith(('~$', '.'))
    ]

    candidats = []
    for f in sorted(fichiers):
        ma = extraire_mois_annee(f)
        d = extraire_date(f)
        date_fichier = f"{d[0]}-{d[1]}-{d[2]}" if d else None
        if ma and date_fichier:
            if date_fichier not in dates_traitees(ma):
                candidats.append(f)
        elif f not in deja_copies:
            candidats.append(f)

    copies = []
    for nom in candidats:
        try:
            shutil.copy2(os.path.join(source, nom), os.path.join(folder_a_traiter(), nom))
            copies.append(nom)
        except Exception:
            logger.exception("site_down : erreur copie %s", nom)

    if copies:
        _ecrire_journal(_journal_alarmes(), copies)
        logger.info("site_down : %d fichier(s) alarme copié(s)", len(copies))
    return len(copies)


# ──────────────────────────────────────────────────────────────────────────────
# Mappings depuis la base Django
# ──────────────────────────────────────────────────────────────────────────────
def charger_regions_map():
    """SITE NAME (majuscules) → région, depuis le modèle Site."""
    from .models import Site
    return {
        (name or '').strip().upper(): (region or '').strip()
        for name, region in Site.objects.exclude(region='').values_list('site_name', 'region')
    }


def charger_causes_escalades_map(mois_annee=None):
    """(site, alarm_time arrondi à la minute) → {Cause, Escalade}.

    Construit depuis le modèle Incident (domaine mobile), au lieu de relire
    les rapports journaliers Excel.
    """
    from .models import Incident
    qs = Incident.objects.filter(domain=Incident.DOMAIN_MOBILE).exclude(alarm_time=None)
    if mois_annee:
        annee, mois = mois_annee.split('-')
        qs = qs.filter(alarm_time__year=int(annee), alarm_time__month=int(mois))

    mapping = {}
    for site, alarm_time, cause, escalade in qs.values_list(
            'site_name', 'alarm_time', 'cause', 'escalade'):
        if not site:
            continue
        at = alarm_time
        if at.tzinfo is not None:
            at = at.astimezone().replace(tzinfo=None)
        key = (site.strip(), at.replace(second=0, microsecond=0))
        mapping[key] = {'Cause': cause or '', 'Escalade': escalade or ''}
    return mapping


def charger_causes_escalades_rapport(file_or_path, filename=''):
    """Mapping (site, alarm_time à la minute) → {Cause, Escalade} depuis un
    rapport journalier Excel (onglet ``INCIDENTS MOB J-1``).

    Accepte un chemin ou un objet fichier (upload Django).
    """
    df = pd.read_excel(file_or_path, sheet_name='INCIDENTS MOB J-1')
    required = ['Site Name', 'Alarm Time', 'Cause', 'Escalade']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"{filename or 'rapport journalier'} : colonnes manquantes {missing} "
            "dans l'onglet INCIDENTS MOB J-1")

    df['Alarm Time'] = pd.to_datetime(df['Alarm Time'], errors='coerce', dayfirst=True)
    mapping = {}
    for _, row in df.iterrows():
        site = str(row['Site Name']).strip()
        at = row['Alarm Time']
        if not site or pd.isna(at):
            continue
        key = (site, at.to_pydatetime().replace(second=0, microsecond=0))
        mapping[key] = {
            'Cause':    str(row['Cause'])    if pd.notna(row['Cause'])    else '',
            'Escalade': str(row['Escalade']) if pd.notna(row['Escalade']) else '',
        }
    return mapping


def ajouter_cause_escalade(df, mapping):
    causes, escalades = [], []
    for name, at in zip(df['Name'], df['Alarm Time']):
        cause = escalade = ''
        if pd.notna(at):
            key = (str(name).strip(), at.to_pydatetime().replace(second=0, microsecond=0))
            entry = mapping.get(key)
            if entry:
                cause, escalade = entry['Cause'], entry['Escalade']
        causes.append(cause)
        escalades.append(escalade)
    df['Cause'] = causes
    df['Escalade'] = escalades
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Lecture d'un fichier d'alarmes brut
# ──────────────────────────────────────────────────────────────────────────────
def _float_to_duration(val):
    if pd.isna(val) or val == 0:
        return '00:00:00'
    try:
        total = float(val) * 24 * 3600
        return f"{int(total // 3600):02d}:{int((total % 3600) // 60):02d}:{int(total % 60):02d}"
    except (TypeError, ValueError):
        return str(val)


def traiter_fichier(file_path):
    """Lit un fichier d'alarmes (xlsx/xls/csv) et normalise ses colonnes.

    Returns:
        DataFrame avec colonnes Name, Alarm Time, Cancel Time, Duration, Alarm Text.
    """
    ext = file_path.lower().rsplit('.', 1)[-1]
    if ext in ('xlsx', 'xls'):
        # Détection de la ligne d'en-tête sur les 10 premières lignes seulement
        df_raw = pd.read_excel(file_path, header=None, nrows=10)
        header_row = 0
        for i, row in df_raw.iterrows():
            vals = ' '.join(str(v).strip().lower() for v in row.tolist())
            if any(k in vals for k in ('alarm time', 'cancel time', 'name', 'site')):
                header_row = i
                break
        df = pd.read_excel(file_path, header=header_row)
    elif ext == 'csv':
        df = pd.read_csv(file_path, sep=None, engine='python', encoding='utf-8-sig')
    else:
        raise ValueError(f"Format non supporté : {ext}")

    df.columns = df.columns.str.strip().str.replace('\ufeff', '')

    mapping = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl == 'site' or cl == 'site name':
            mapping['site_name'] = col
        elif cl == 'name':
            mapping.setdefault('site_name', col)
        elif 'alarm' in cl and 'text' in cl:
            mapping['alarm_text'] = col
        elif 'alarm' in cl and 'time' in cl:
            mapping['alarm_time'] = col
        elif 'cancel' in cl and 'time' in cl:
            mapping['cancel_time'] = col
        elif 'duration' in cl or 'durée' in cl or cl == 'd':
            mapping.setdefault('duration', col)

    missing = [c for c in ('site_name', 'alarm_time', 'cancel_time', 'duration')
               if c not in mapping]
    if missing:
        raise ValueError(f"Colonnes manquantes : {missing}")

    cols      = [mapping['site_name'], mapping['alarm_time'],
                 mapping['cancel_time'], mapping['duration']]
    col_names = ['Name', 'Alarm Time', 'Cancel Time', 'Duration']
    if 'alarm_text' in mapping:
        cols.append(mapping['alarm_text'])
        col_names.append('Alarm Text')

    df = df[cols].copy()
    df.columns = col_names
    if 'Alarm Text' not in df.columns:
        df['Alarm Text'] = ''
    df = df.dropna(subset=['Name'])
    df['Alarm Time']  = pd.to_datetime(df['Alarm Time'],  errors='coerce', dayfirst=True)
    df['Cancel Time'] = pd.to_datetime(df['Cancel Time'], errors='coerce', dayfirst=True)

    if df['Duration'].dtype == 'float64':
        df['Duration'] = df['Duration'].apply(_float_to_duration)
    elif pd.api.types.is_timedelta64_dtype(df['Duration']):
        df['Duration'] = df['Duration'].apply(
            lambda x: str(x).split(' days ')[-1] if pd.notna(x) else '')
    elif df['Duration'].apply(lambda x: isinstance(x, time)).any():
        df['Duration'] = df['Duration'].apply(
            lambda x: x.strftime('%H:%M:%S') if isinstance(x, time) else str(x))
    return df


def dur_to_min(val):
    if pd.isna(val) or val in ('', '00:00:00'):
        return 0
    try:
        parts = str(val).split(':')
        if len(parts) == 3:
            return int(parts[0]) * 60 + int(parts[1]) + int(parts[2]) / 60
    except (TypeError, ValueError):
        pass
    return 0


def _clean_str(val):
    s = str(val).strip() if pd.notna(val) else ''
    return '' if s.lower() in ('nan', 'none', '') else s


# ──────────────────────────────────────────────────────────────────────────────
# Feuille Cumul (pivot sites × jours)
# ──────────────────────────────────────────────────────────────────────────────
def _duree_reelle_par_site_jour(df_site):
    """Durée réelle (min) par jour pour un site : découpe par minuit puis
    fusionne les plages chevauchantes pour éviter le double comptage."""
    plages_par_jour = {}
    for alarm_dt, cancel_dt in zip(df_site['Alarm Time'], df_site['Cancel Time']):
        if pd.isna(alarm_dt) or pd.isna(cancel_dt) or cancel_dt <= alarm_dt:
            continue
        current = alarm_dt
        while current < cancel_dt:
            minuit_suivant = pd.Timestamp(current.date()) + pd.Timedelta(days=1)
            fin_tranche = min(cancel_dt, minuit_suivant)
            if fin_tranche > current:
                plages_par_jour.setdefault(current.day, []).append((current, fin_tranche))
            current = minuit_suivant

    result = {}
    for jour, plages in plages_par_jour.items():
        merged = []
        for start, end in sorted(plages, key=lambda x: x[0]):
            if merged and start <= merged[-1][1]:
                merged[-1][1] = max(merged[-1][1], end)
            else:
                merged.append([start, end])
        result[jour] = int(round(sum((e - s).total_seconds() / 60 for s, e in merged)))
    return result


def creer_feuille_cumul(df_consolide, mois_annee, regions_map=None):
    base_cols = ['Name', 'Alarm Time', 'Cancel Time', 'Duration', 'Cause', 'Escalade']
    if 'Alarm Text' in df_consolide.columns:
        base_cols.append('Alarm Text')
    df = df_consolide[base_cols].copy()
    if 'Alarm Text' not in df.columns:
        df['Alarm Text'] = ''

    # Si Alarm Text absent, toutes les lignes comptent dans Nb
    if not (df['Alarm Text'].notna().any() and (df['Alarm Text'].astype(str).str.strip() != '').any()):
        df['Alarm Text'] = ALARM_FILTER

    df['Alarm Time']  = pd.to_datetime(df['Alarm Time'],  errors='coerce')
    df['Cancel Time'] = pd.to_datetime(df['Cancel Time'], errors='coerce')
    df = df.sort_values('Alarm Time')

    mask = df['Alarm Time'].notna()
    df_nb = pd.DataFrame({
        'Name':     df.loc[mask, 'Name'],
        'jour':     df.loc[mask, 'Alarm Time'].dt.day,
        'is_wcdma': df.loc[mask, 'Alarm Text'].astype(str).str.strip().str.upper()
                      .str.contains(ALARM_FILTER, regex=False),
    })
    df_wcdma = df_nb[df_nb['is_wcdma']] if len(df_nb) else df_nb

    all_sites = df['Name'].unique()
    groupes = dict(tuple(df.groupby('Name', sort=False)))
    durees_par_site, all_jours_set = {}, set()
    for site in all_sites:
        d = _duree_reelle_par_site_jour(groupes[site])
        durees_par_site[site] = d
        all_jours_set.update(d.keys())
    all_jours = sorted(all_jours_set)

    if len(df_wcdma) > 0:
        pivot_count = df_wcdma.pivot_table(index='Name', columns='jour',
                                           values='is_wcdma', aggfunc='count', fill_value=0)
    else:
        pivot_count = pd.DataFrame()
    pivot_count = pivot_count.reindex(index=all_sites, columns=all_jours, fill_value=0)

    pivot_duree = pd.DataFrame(
        [{**{'Name': s}, **{j: durees_par_site[s].get(j, 0) for j in all_jours}}
         for s in all_sites]
    ).set_index('Name') if len(all_sites) else pd.DataFrame()
    if len(pivot_duree):
        pivot_duree = pivot_duree.reindex(index=all_sites, columns=all_jours, fill_value=0)

    _, mois_num = mois_annee.split('-')
    nom_mois = NOMS_MOIS_COURTS.get(mois_num, mois_num)

    result = pd.DataFrame()
    result['Site'] = list(all_sites)
    result['Région'] = result['Site'].apply(
        lambda s: regions_map.get(str(s).strip().upper(), '') if regions_map else '')

    for jour in all_jours:
        prefix = f"{int(jour)}-{nom_mois}"
        nb_vals, dur_vals = [], []
        for site in all_sites:
            count = int(pivot_count.loc[site, jour]) if site in pivot_count.index else 0
            duree = int(round(pivot_duree.loc[site, jour])) if site in pivot_duree.index else 0
            nb_vals.append(count or None)
            dur_vals.append(duree or None)
        result[f"{prefix} Nb"] = nb_vals
        result[f"{prefix} Durée"] = dur_vals

    total_nb, total_dur, col_cause, col_escalade = [], [], [], []
    for site in all_sites:
        t_count = int(pivot_count.loc[site].sum()) if site in pivot_count.index else 0
        t_duree = int(round(pivot_duree.loc[site].sum())) if site in pivot_duree.index else 0
        site_data = groupes[site]
        derniere_cause    = _clean_str(site_data.iloc[-1]['Cause'])    if len(site_data) else ''
        derniere_escalade = _clean_str(site_data.iloc[-1]['Escalade']) if len(site_data) else ''
        total_nb.append(t_count or None)
        total_dur.append(t_duree or None)
        col_cause.append(derniere_cause or None)
        col_escalade.append(derniere_escalade or None)

    result['Total Nb']    = total_nb
    result['Total Durée'] = total_dur
    result['Cause']       = col_cause
    result['Escalade']    = col_escalade

    result['_sort'] = result['Total Nb'].fillna(0)
    return (result.sort_values('_sort', ascending=False)
                  .drop(columns=['_sort']).reset_index(drop=True))


# ──────────────────────────────────────────────────────────────────────────────
# Formatage Excel
# ──────────────────────────────────────────────────────────────────────────────
# Styles partagés (créés une seule fois — la création par cellule est coûteuse)
_FONT_HEADER   = Font(name='Arial', bold=True, color=NOIR_TEXTE, size=10)
_FILL_HEADER   = PatternFill('solid', start_color=JAUNE_HEADER)
_ALIGN_HEADER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
_FONT_DATA     = Font(name='Arial', size=9)
_ALIGN_LEFT    = Alignment(horizontal='left',   vertical='center')
_ALIGN_CENTER  = Alignment(horizontal='center', vertical='center')


def _style_header(cell):
    cell.font      = _FONT_HEADER
    cell.fill      = _FILL_HEADER
    cell.alignment = _ALIGN_HEADER
    cell.border    = BORDER_THIN


def formater_feuille(ws, df, sheet_name):
    col_widths = ({'Name': 28, 'Alarm Time': 22, 'Cancel Time': 22,
                   'Duration': 14, 'Cause': 25, 'Escalade': 18}
                  if sheet_name == 'Données' else {})

    n_cols = len(df.columns)
    n_rows = ws.max_row          # calculés une seule fois (coûteux par appel)

    for col_idx in range(1, n_cols + 1):
        _style_header(ws.cell(row=1, column=col_idx))
    ws.row_dimensions[1].height = 30

    # Alignement précalculé par colonne ; le zébrage et les bordures sont
    # assurés par le style de tableau Excel (bien plus rapide que cellule à cellule)
    aligns = [
        _ALIGN_LEFT if (i == 0 or 'Cause' in c or 'Escalade' in c) else _ALIGN_CENTER
        for i, c in enumerate(df.columns)
    ]
    for row in ws.iter_rows(min_row=2, max_row=n_rows, max_col=n_cols):
        for cell, align in zip(row, aligns):
            cell.font      = _FONT_DATA
            cell.alignment = align
    for row_idx in range(2, n_rows + 1):
        ws.row_dimensions[row_idx].height = 16

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name in col_widths:
            ws.column_dimensions[letter].width = col_widths[col_name]
        elif sheet_name == 'Données':
            ws.column_dimensions[letter].width = 15
        elif col_idx == 1:
            ws.column_dimensions[letter].width = 30
        elif col_name == 'Région':
            ws.column_dimensions[letter].width = 16
        elif 'Cause' in col_name or 'Escalade' in col_name:
            ws.column_dimensions[letter].width = 22
        elif 'Nb' in col_name:
            ws.column_dimensions[letter].width = 8
        elif 'Durée' in col_name:
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 14

    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', sheet_name)
    table = Table(displayName=f"Tbl_{safe_name}",
                  ref=f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = 'B2'

    if sheet_name == 'Données':
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name == 'Alarm Text':
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True
                break


def colorier_totaux_cumul(ws, df):
    if 'Total Nb' not in df.columns:
        return

    jours = []
    for col in df.columns:
        m = re.match(r'^(\d+)-\w+ Nb$', col)
        if m:
            jours.append((int(m.group(1)), col))
    jours.sort(key=lambda x: x[0])
    if not jours:
        return

    dernier_jour = max(j for j, _ in jours)
    cols_rouge  = [c for j, c in jours if dernier_jour - 3 <= j <= dernier_jour]
    cols_orange = [c for j, c in jours if dernier_jour - 2 <= j <= dernier_jour]

    col_total_nb  = next((i for i, c in enumerate(df.columns, 1) if c == 'Total Nb'), None)
    col_total_dur = next((i for i, c in enumerate(df.columns, 1) if c == 'Total Durée'), None)
    if col_total_nb is None:
        return

    def tous_en_panne(row, cols):
        return bool(cols) and all(pd.notna(row.get(c)) and row.get(c, 0) > 0 for c in cols)

    def aucune_panne(row, cols):
        return all(not (pd.notna(row.get(c)) and row.get(c, 0) > 0) for c in cols)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        if len(cols_rouge) == 4 and tous_en_panne(row, cols_rouge):
            couleur, font_c = ROUGE_TOTAL, Font(name='Arial', size=9, bold=True, color='FFFFFF')
        elif len(cols_orange) == 3 and tous_en_panne(row, cols_orange):
            couleur, font_c = ORANGE_TOTAL, Font(name='Arial', size=9, bold=True, color='000000')
        elif aucune_panne(row, cols_orange):
            couleur, font_c = VERT_TOTAL, Font(name='Arial', size=9, bold=True, color='000000')
        else:
            couleur, font_c = None, Font(name='Arial', size=9)

        for col_idx in (col_total_nb, col_total_dur):
            if col_idx:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill('solid', start_color=couleur) if couleur else PatternFill()
                cell.font = font_c
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER_THIN


def ajouter_heatmap_cumul(ws, df):
    """Dégradé blanc → jaune → rouge sur les colonnes « Durée » journalières."""
    if len(df) == 0:
        return
    for col_idx, col_name in enumerate(df.columns, start=1):
        if re.match(r'^\d+-\w+ Durée$', col_name):
            letter = get_column_letter(col_idx)
            ws.conditional_formatting.add(
                f'{letter}2:{letter}{len(df) + 1}',
                ColorScaleRule(
                    start_type='num', start_value=0,   start_color='FFFFFF',
                    mid_type='num',   mid_value=120,   mid_color=JAUNE_YAS,
                    end_type='num',   end_value=720,   end_color=ROUGE_TOTAL))


# ──────────────────────────────────────────────────────────────────────────────
# Onglet Synthèse (KPI + graphiques)
# ──────────────────────────────────────────────────────────────────────────────
def _fmt_duree(total_min):
    total_min = int(total_min)
    if total_min >= 1440:
        return f"{total_min // 1440} j {(total_min % 1440) // 60} h"
    if total_min >= 60:
        return f"{total_min // 60} h {total_min % 60:02d} min"
    return f"{total_min} min"


def creer_feuille_synthese(wb, df, df_cumul, mois_annee):
    """Insère un onglet Synthèse en première position : KPI + 3 graphiques."""
    annee, mois_num = mois_annee.split('-')
    nom_mois = NOMS_MOIS_COMPLETS.get(mois_num, mois_num)

    if 'Synthèse' in wb.sheetnames:
        del wb['Synthèse']
    if '_data' in wb.sheetnames:
        del wb['_data']
    ws = wb.create_sheet('Synthèse', 0)
    ws.sheet_view.showGridLines = False

    # ── Données dérivées ───────────────────────────────────────────
    mask = df['Alarm Time'].notna()
    total_alarmes = int(mask.sum())
    nb_sites      = int(df['Name'].nunique())
    total_min     = int(df_cumul['Total Durée'].fillna(0).sum()) if len(df_cumul) else 0
    par_jour      = (df.loc[mask].groupby(df.loc[mask, 'Alarm Time'].dt.day).size()
                     if total_alarmes else pd.Series(dtype=int))
    jour_pic      = int(par_jour.idxmax()) if len(par_jour) else None
    top10         = df_cumul.head(10) if len(df_cumul) else df_cumul

    esc = df['Escalade'].astype(str).str.strip() if 'Escalade' in df.columns else pd.Series(dtype=str)
    par_escalade = esc[esc != ''].value_counts()
    if len(par_escalade) == 0 and 'Région' in df_cumul.columns:
        reg = df_cumul['Région'].astype(str).str.strip()
        par_escalade = reg[reg != ''].value_counts()
        pie_titre = 'Répartition par région'
    else:
        pie_titre = 'Répartition par escalade'

    # ── Feuille de données cachée pour les graphiques ──────────────────────
    wsd = wb.create_sheet('_data')
    wsd.sheet_state = 'hidden'
    wsd['A1'], wsd['B1'] = 'Jour', 'Alarmes'
    for i, (jour, nb) in enumerate(sorted(par_jour.items()), start=2):
        wsd.cell(row=i, column=1, value=int(jour))
        wsd.cell(row=i, column=2, value=int(nb))
    wsd['D1'], wsd['E1'], wsd['F1'] = 'Site', 'Nb', 'Durée (min)'
    for i, (_, r) in enumerate(top10.iterrows(), start=2):
        wsd.cell(row=i, column=4, value=str(r['Site']))
        wsd.cell(row=i, column=5, value=int(r['Total Nb'] or 0))
        wsd.cell(row=i, column=6, value=int(r['Total Durée'] or 0))
    wsd['H1'], wsd['I1'] = 'Catégorie', 'Nb'
    for i, (label, nb) in enumerate(par_escalade.head(8).items(), start=2):
        wsd.cell(row=i, column=8, value=str(label))
        wsd.cell(row=i, column=9, value=int(nb))

    # ── Titre ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:P2')
    c = ws['A1']
    c.value = f"📉 SITE DOWN — {nom_mois} {annee}"
    c.font = Font(name='Arial', bold=True, size=18, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=BLEU_FONCE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A3:P3')
    c = ws['A3']
    c.value = f"Généré le {datetime.now():%d/%m/%Y %H:%M} — consolidation mensuelle des micro-coupures"
    c.font = Font(name='Arial', size=10, color='888888', italic=True)
    c.alignment = Alignment(horizontal='center')

    # ── Cartes KPI ────────────────────────────────────────────────────
    kpis = [
        ('Alarmes',        f"{total_alarmes:,}".replace(',', ' ')),
        ('Sites impactés', f"{nb_sites:,}".replace(',', ' ')),
        ('Durée cumulée',  _fmt_duree(total_min)),
        ('Durée moyenne',  _fmt_duree(total_min / total_alarmes) if total_alarmes else '—'),
        ('Jour le plus impacté', f"{jour_pic} {NOMS_MOIS_COURTS.get(mois_num, '')}" if jour_pic else '—'),
    ]
    start_cols = [1, 4, 7, 10, 13]   # A, D, G, J, M
    for (label, value), sc in zip(kpis, start_cols):
        l1, l2 = get_column_letter(sc), get_column_letter(sc + 2)
        ws.merge_cells(f'{l1}5:{l2}5')
        ws.merge_cells(f'{l1}6:{l2}6')
        lab = ws[f'{l1}5']; lab.value = label.upper()
        lab.font = Font(name='Arial', size=9, bold=True, color='888888')
        lab.alignment = Alignment(horizontal='center')
        val = ws[f'{l1}6']; val.value = value
        val.font = Font(name='Arial', size=16, bold=True, color=BLEU_FONCE)
        val.alignment = Alignment(horizontal='center')
        for row in (5, 6):
            for cc in range(sc, sc + 3):
                ws.cell(row=row, column=cc).fill = PatternFill('solid', start_color='F0F4FF')
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 26

    # ── Graphiques ────────────────────────────────────────────────────
    n_top = len(top10)
    if n_top:
        bar = BarChart()
        bar.type, bar.style = 'col', 10
        bar.title = 'Top 10 sites (nombre de coupures)'
        bar.y_axis.title = 'Nb'
        bar.legend = None
        bar.width, bar.height = 17, 9
        bar.add_data(Reference(wsd, min_col=5, min_row=1, max_row=n_top + 1),
                     titles_from_data=True)
        bar.set_categories(Reference(wsd, min_col=4, min_row=2, max_row=n_top + 1))
        ws.add_chart(bar, 'A9')

    if len(par_jour):
        line = LineChart()
        line.title = 'Alarmes par jour'
        line.y_axis.title = 'Nb'
        line.x_axis.title = 'Jour'
        line.legend = None
        line.width, line.height = 17, 9
        line.add_data(Reference(wsd, min_col=2, min_row=1, max_row=len(par_jour) + 1),
                      titles_from_data=True)
        line.set_categories(Reference(wsd, min_col=1, min_row=2, max_row=len(par_jour) + 1))
        ws.add_chart(line, 'J9')

    if len(par_escalade):
        pie = PieChart()
        pie.title = pie_titre
        pie.width, pie.height = 12, 9
        n_pie = min(len(par_escalade), 8)
        pie.add_data(Reference(wsd, min_col=9, min_row=1, max_row=n_pie + 1),
                     titles_from_data=True)
        pie.set_categories(Reference(wsd, min_col=8, min_row=2, max_row=n_pie + 1))
        ws.add_chart(pie, 'A28')

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 9
    wb.active = 0


def _ecrire_fichier_mensuel(output_file, df_consolide, df_cumul, mois_annee, nom_feuille_cumul):
    """Écrit le fichier mensuel complet : Synthèse + Données + Cumul formatés."""
    cols_donnees = [c for c in df_consolide.columns if c not in ('Cause', 'Escalade')]
    df_donnees = df_consolide[cols_donnees]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_donnees.to_excel(writer, sheet_name='Données', index=False)
        df_cumul.to_excel(writer, sheet_name=nom_feuille_cumul, index=False)

    wb = load_workbook(output_file)
    formater_feuille(wb['Données'], df_donnees, 'Données')
    formater_feuille(wb[nom_feuille_cumul], df_cumul, nom_feuille_cumul)
    colorier_totaux_cumul(wb[nom_feuille_cumul], df_cumul)
    ajouter_heatmap_cumul(wb[nom_feuille_cumul], df_cumul)
    creer_feuille_synthese(wb, df_consolide, df_cumul, mois_annee)
    wb.save(output_file)
    return df_donnees


# ──────────────────────────────────────────────────────────────────────────────
# Persistance ORM
# ──────────────────────────────────────────────────────────────────────────────
def _sauvegarder_orm(df, source_file, regions_map):
    """Upsert en masse des alarmes dans SiteDownAlarm (clé : site + alarm_time)."""
    from django.utils import timezone as tz
    from .models import SiteDownAlarm

    current_tz = tz.get_current_timezone()

    def aware(ts):
        if pd.isna(ts):
            return None
        dt = ts.to_pydatetime()
        return tz.make_aware(dt, current_tz) if dt.tzinfo is None else dt

    objs, cles = [], []
    for name, alarm, cancel, duration, alarm_text, cause, escalade in zip(
            df['Name'], df['Alarm Time'], df['Cancel Time'], df.get('Duration', []),
            df.get('Alarm Text', ['']*len(df)), df.get('Cause', ['']*len(df)),
            df.get('Escalade', ['']*len(df))):
        alarm_time = aware(alarm)
        if alarm_time is None:
            continue
        site = str(name).strip()
        cles.append((site, alarm_time))
        objs.append(SiteDownAlarm(
            site_name=site,
            alarm_time=alarm_time,
            cancel_time=aware(cancel),
            duration_min=dur_to_min(duration) or None,
            alarm_text=_clean_str(alarm_text)[:255],
            cause=_clean_str(cause),
            escalade=_clean_str(escalade)[:80],
            region=(regions_map or {}).get(site.upper(), '')[:50],
            source_file=source_file[:255],
        ))
    if not objs:
        return 0, 0

    # Comptage créés/mis à jour avant l'upsert en masse
    alarm_times = [c[1] for c in cles]
    existants = set(SiteDownAlarm.objects.filter(
        alarm_time__gte=min(alarm_times), alarm_time__lte=max(alarm_times),
    ).values_list('site_name', 'alarm_time'))
    updated = sum(1 for c in cles if c in existants)
    created = len(objs) - updated

    SiteDownAlarm.objects.bulk_create(
        objs,
        update_conflicts=True,
        unique_fields=['site_name', 'alarm_time'],
        update_fields=['cancel_time', 'duration_min', 'alarm_text',
                       'cause', 'escalade', 'region', 'source_file'],
        batch_size=1000,
    )
    return created, updated


# ──────────────────────────────────────────────────────────────────────────────
# Pipeline principal
# ──────────────────────────────────────────────────────────────────────────────
def fichiers_mensuels():
    """Liste [(nom, chemin, taille, mtime)] des fichiers consolidés produits."""
    ensure_dirs()
    out = []
    for f in sorted(os.listdir(folder_traites()), reverse=True):
        if re.fullmatch(r'SITE_DOWN_\d{4}-\d{2}\.xlsx', f):
            p = os.path.join(folder_traites(), f)
            st = os.stat(p)
            out.append({'name': f, 'path': p, 'size': st.st_size,
                        'mtime': datetime.fromtimestamp(st.st_mtime)})
    return out


def process_pending_files(extra_causes_map=None):
    """Traite tous les fichiers présents dans ``a_traiter``.

    Args:
        extra_causes_map: mapping Cause/Escalade supplémentaire (ex. rapport
            journalier uploadé) — prioritaire sur celui de la base Incident.

    Returns:
        dict résumé : {processed, errors, months, created, updated, messages}
    """
    ensure_dirs()
    regions_map = charger_regions_map()

    files = sorted(f for f in os.listdir(folder_a_traiter())
                   if os.path.isfile(os.path.join(folder_a_traiter(), f)))

    summary = {'processed': 0, 'errors': 0, 'months': [],
               'created': 0, 'updated': 0, 'messages': []}

    if not files:
        summary['messages'].append("Aucun nouveau fichier à traiter.")
        return summary

    fichiers_par_mois = {}
    for f in files:
        ma = extraire_mois_annee(f)
        if ma:
            fichiers_par_mois.setdefault(ma, []).append(f)
        else:
            summary['errors'] += 1
            summary['messages'].append(
                f"❌ {f} : aucune date (JJ-MM-AAAA) trouvée dans le nom du fichier — "
                "déplacé vers le dossier erreurs.")
            try:
                shutil.move(os.path.join(folder_a_traiter(), f),
                            os.path.join(folder_erreurs(), f))
            except OSError:
                pass

    for mois_annee in sorted(fichiers_par_mois):
        causes_map = charger_causes_escalades_map(mois_annee)
        if extra_causes_map:
            causes_map = {**causes_map, **extra_causes_map}
        annee, mois_num = mois_annee.split('-')
        nom_feuille_cumul = f"Cumul_{NOMS_MOIS_COMPLETS.get(mois_num, mois_num)}_{annee}"
        output_file = os.path.join(folder_traites(), f'SITE_DOWN_{mois_annee}.xlsx')

        df_existant = None
        if os.path.exists(output_file):
            try:
                df_existant = pd.read_excel(output_file, sheet_name='Données')
                df_existant['Alarm Time']  = pd.to_datetime(df_existant['Alarm Time'],  errors='coerce')
                df_existant['Cancel Time'] = pd.to_datetime(df_existant['Cancel Time'], errors='coerce')
            except Exception:
                logger.warning("site_down : lecture existant %s échouée", output_file, exc_info=True)

        dfs_mois, fichiers_ok = [], []
        for file in sorted(fichiers_par_mois[mois_annee]):
            file_path = os.path.join(folder_a_traiter(), file)
            try:
                df = traiter_fichier(file_path)
                df = ajouter_cause_escalade(df, causes_map)
                dfs_mois.append(df)
                fichiers_ok.append((file, file_path))
                summary['processed'] += 1
            except Exception as exc:
                logger.exception("site_down : erreur traitement %s", file)
                summary['errors'] += 1
                summary['messages'].append(f"Erreur {file} : {exc}")
                try:
                    shutil.move(file_path, os.path.join(folder_erreurs(), file))
                except OSError:
                    pass

        if not dfs_mois:
            continue

        df_nouvelles = pd.concat(dfs_mois, ignore_index=True)

        if df_existant is not None and len(df_existant) > 0:
            df_existant = ajouter_cause_escalade(df_existant, causes_map)
            df_consolide = pd.concat([df_existant, df_nouvelles], ignore_index=True)
            df_consolide = df_consolide.drop_duplicates(subset=['Name', 'Alarm Time'], keep='last')
        else:
            df_consolide = df_nouvelles

        df_consolide = df_consolide.sort_values('Alarm Time').reset_index(drop=True)
        df_cumul = creer_feuille_cumul(df_consolide, mois_annee, regions_map)

        df_donnees = _ecrire_fichier_mensuel(
            output_file, df_consolide, df_cumul, mois_annee, nom_feuille_cumul)

        # Persistance ORM (uniquement les nouvelles lignes traitées)
        created, updated = _sauvegarder_orm(
            df_nouvelles, ', '.join(f for f, _ in fichiers_ok)[:255], regions_map)
        summary['created'] += created
        summary['updated'] += updated
        summary['months'].append(mois_annee)
        summary['messages'].append(
            f"{mois_annee} : {len(df_donnees)} lignes, {len(df_cumul)} sites "
            f"({created} créées, {updated} mises à jour en base)")

        for _, fp in fichiers_ok:
            try:
                os.remove(fp)
            except OSError:
                pass

    return summary


def actualiser_fichiers_existants(extra_causes_map=None):
    """Ré-applique Cause/Escalade + cumul sur les fichiers mensuels existants."""
    ensure_dirs()
    regions_map = charger_regions_map()
    summary = {'refreshed': 0, 'errors': 0, 'messages': []}

    for info in fichiers_mensuels():
        m = re.search(r'SITE_DOWN_(\d{4})-(\d{2})\.xlsx', info['name'])
        if not m:
            continue
        mois_annee = f"{m.group(1)}-{m.group(2)}"
        nom_feuille_cumul = f"Cumul_{NOMS_MOIS_COMPLETS.get(m.group(2), m.group(2))}_{m.group(1)}"
        try:
            df = pd.read_excel(info['path'], sheet_name='Données')
            df['Alarm Time']  = pd.to_datetime(df['Alarm Time'],  errors='coerce')
            df['Cancel Time'] = pd.to_datetime(df['Cancel Time'], errors='coerce')
            causes_map = charger_causes_escalades_map(mois_annee)
            if extra_causes_map:
                causes_map = {**causes_map, **extra_causes_map}
            df = ajouter_cause_escalade(df, causes_map)
            df_cumul = creer_feuille_cumul(df, mois_annee, regions_map)

            _ecrire_fichier_mensuel(
                info['path'], df, df_cumul, mois_annee, nom_feuille_cumul)
            summary['refreshed'] += 1
        except Exception as exc:
            logger.exception("site_down : actualisation %s échouée", info['name'])
            summary['errors'] += 1
            summary['messages'].append(f"Erreur {info['name']} : {exc}")

    return summary


def run_auto():
    """Point d'entrée complet : collecte réseau puis traitement (scheduler)."""
    nb = collecter_alarmes()
    summary = process_pending_files()
    summary['collected'] = nb
    if nb == 0 and summary['processed'] == 0:
        # Rien de neuf : rafraîchit quand même Cause/Escalade des fichiers du mois
        refresh = actualiser_fichiers_existants()
        summary['messages'].append(
            f"Actualisation : {refresh['refreshed']} fichier(s), {refresh['errors']} erreur(s)")
    return summary
