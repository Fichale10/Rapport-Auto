"""
Parser pour le fichier BASES DES INCIDENTS (format CGI multi-plateforme).
Onglets attendus : Reseau mobile, DR2, Reseau Fixe, Transport, IGW, Core
"""
import openpyxl
from datetime import datetime as _dt, timedelta
from collections import defaultdict


# ── Helpers bas niveau ────────────────────────────────────────────────────────

def _norm(h):
    return (str(h or '').strip().upper()
            .replace('É','E').replace('È','E').replace('Ê','E')
            .replace('À','A').replace('Î','I').replace('\xa0',' '))


def _find_header_row(ws, min_cols=3):
    for ridx, row in enumerate(ws.iter_rows(max_row=12, values_only=True), 1):
        non_none = [c for c in row if c is not None and str(c).strip()]
        if len(non_none) >= min_cols:
            return ridx
    return 1


def _parse_duration(val):
    if val is None:
        return None
    if isinstance(val, timedelta):
        return val.total_seconds()
    s = str(val).strip()
    if not s or s.upper() in ('N/A', 'EN COURS', 'NULL', '', 'NONE'):
        return None
    # Format HH:MM:SS or large H:MM:SS
    parts = s.split(':')
    try:
        if len(parts) == 3:
            h, m, sec = int(parts[0]), int(parts[1]), float(parts[2])
            return abs(h) * 3600 + m * 60 + sec
        elif len(parts) == 2:
            m, sec = int(parts[0]), float(parts[1])
            return m * 60 + sec
    except (ValueError, TypeError):
        pass
    return None


def _parse_dt(val):
    if val is None:
        return None
    if isinstance(val, _dt):
        return val
    s = str(val).strip()
    for fmt in ('%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S',
                '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M', '%d-%m-%Y'):
        try:
            return _dt.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _fmt_dur(seconds):
    if not seconds:
        return '—'
    s = int(seconds)
    h = s // 3600; m = (s % 3600) // 60
    return f'{h}h {m:02d}m'


def _sheet_rows(ws, header_row):
    rows = list(ws.iter_rows(values_only=True))
    headers = [_norm(h) for h in rows[header_row - 1]]
    for row in rows[header_row:]:
        if not any(v for v in row if v is not None):
            continue
        yield {headers[i]: row[i] for i in range(min(len(headers), len(row)))}


def _is_closed(val):
    return str(val or '').strip().upper() in ('FERME', 'CLOSED', 'RESOLU', 'FERME ')


# ── Parseurs par onglet ───────────────────────────────────────────────────────

def parse_fixe(ws):
    hr = _find_header_row(ws)
    rows = []
    for d in _sheet_rows(ws, hr):
        dur = _parse_duration(d.get('DURATION'))
        rows.append({
            'nature':       str(d.get("NATURE DE L'INCIDENT", d.get('NATURE', '')) or '')[:80],
            'alarm':        _parse_dt(d.get('ALARM TIME')),
            'site_name':    str(d.get('SITE NAME', '') or '').strip(),
            'plateforme':   str(d.get('PLATEFORME', '') or '').strip(),
            'impact_svc':   str(d.get('IMPACT - SERVICE', '') or '').strip(),
            'nb_clients':   str(d.get('NBRE DE CLIENT IMPACTES', d.get('NBRE DE CLIENTS IMPACTES', '')) or '').strip(),
            'escalade':     str(d.get('ESCALADE', '') or '').strip(),
            'root_cause':   str(d.get('ROOT CAUSE', '') or '').strip()[:60],
            'cancel':       _parse_dt(d.get('CANCEL TIME')),
            'duration_sec': dur,
            'duration_fmt': _fmt_dur(dur),
            'status':       str(d.get('STATUS', '') or '').strip(),
            'is_closed':    _is_closed(d.get('STATUS')),
        })
    return rows


def parse_transport(ws):
    hr = _find_header_row(ws)
    rows = []
    seen = set()
    for d in _sheet_rows(ws, hr):
        ticket = str(d.get('NUMERO DU TICKET', '') or '').strip()
        # Dédoublonner par ticket
        if ticket and ticket in seen:
            continue
        if ticket:
            seen.add(ticket)
        dur = _parse_duration(d.get('DURATION'))
        impact_svc = str(d.get('IMPACT - SERVICE', '') or '').strip()
        avec_impact = bool(impact_svc) and impact_svc.upper() not in ('N/A', 'SANS IMPACT', '')
        rows.append({
            'ticket':       ticket,
            'nature':       str(d.get("NATURE DE L'INCIDENT", '') or '')[:80],
            'alarm':        _parse_dt(d.get('ALARM TIME')),
            'site_name':    str(d.get('SITE NAME', '') or '').strip(),
            'region':       str(d.get('REGION', d.get('REGION', '')) or '').strip(),
            'plateforme':   str(d.get('PLATEFORME', '') or '').strip(),
            'technologies': str(d.get('TECHNOLOGIES', '') or '').strip(),
            'cause':        str(d.get('CAUSE', '') or '').strip()[:60],
            'escalade':     str(d.get('ESCALADE', '') or '').strip(),
            'impact_svc':   impact_svc,
            'avec_impact':  avec_impact,
            'cancel':       _parse_dt(d.get('CANCEL TIME')),
            'duration_sec': dur,
            'duration_fmt': _fmt_dur(dur),
            'status':       str(d.get('STATUS', '') or '').strip().upper(),
            'is_closed':    _is_closed(d.get('STATUS')),
        })
    return rows


def parse_igw(ws):
    hr = _find_header_row(ws)
    rows = []
    for d in _sheet_rows(ws, hr):
        dur = _parse_duration(d.get('DURATION'))
        rows.append({
            'alarm':       _parse_dt(d.get('ALARM TIME')),
            'nature':      str(d.get("NATURE DE L'INCIDENT", d.get("NATURE DE L'INCIDENT ", '')) or '')[:80],
            'lien':        str(d.get('LIEN', '') or '').strip(),
            'lien_internet': str(d.get('LIEN INTERNET', '') or '').strip(),
            'impacts':     str(d.get('IMPACTS', d.get('IMPACTS ', '')) or '').strip(),
            'escalade':    str(d.get('ESCALADE', '') or '').strip(),
            'cause':       str(d.get("CAUSES DE L'INCIDENT", d.get("CAUSES DE L'INCIDENT ", '')) or '')[:60],
            'priorite':    str(d.get('PRIORITES', '') or '').strip(),
            'cancel':      _parse_dt(d.get('CANCEL TIME')),
            'duration_sec': dur,
            'duration_fmt': _fmt_dur(dur),
            'status':      str(d.get('STATUS', d.get('STATUS ', '')) or '').strip().upper(),
            'is_closed':   _is_closed(d.get('STATUS', d.get('STATUS ', ''))),
        })
    return rows


def parse_core(ws):
    hr = _find_header_row(ws)
    rows = []
    for d in _sheet_rows(ws, hr):
        dur = _parse_duration(d.get('DURATION'))
        rows.append({
            'nature':      str(d.get("NATURE DE L'INCIDENT", '') or '')[:80],
            'alarm':       _parse_dt(d.get('ALARM TIME')),
            'espc':        str(d.get('ESPC', '') or '').strip(),
            'impact_svc':  str(d.get('IMPACT - SERVICE', '') or '').strip(),
            'escalade':    str(d.get('ESCALADE', '') or '').strip(),
            'root_cause':  str(d.get('ROOT CAUSE', '') or '')[:60],
            'cancel':      _parse_dt(d.get('CANCEL TIME')),
            'duration_sec': dur,
            'duration_fmt': _fmt_dur(dur),
            'status':      str(d.get('STATUS', '') or '').strip().upper(),
            'is_closed':   _is_closed(d.get('STATUS')),
        })
    return rows


# ── Agrégations ───────────────────────────────────────────────────────────────

def stats_fixe(rows):
    total = len(rows)
    closed = sum(1 for r in rows if r['is_closed'])
    total_dur = sum(r['duration_sec'] for r in rows if r['duration_sec'])
    mttr = total_dur / closed if closed else 0

    by_plat = defaultdict(lambda: {'nb': 0, 'dur': 0})
    for r in rows:
        k = r['plateforme'] or 'Autre'
        by_plat[k]['nb'] += 1
        by_plat[k]['dur'] += r['duration_sec'] or 0

    by_esc = defaultdict(lambda: {'nb': 0, 'dur': 0})
    for r in rows:
        k = r['escalade'] or 'Non défini'
        by_esc[k]['nb'] += 1
        by_esc[k]['dur'] += r['duration_sec'] or 0

    return {
        'total': total, 'closed': closed, 'open': total - closed,
        'mttr_fmt': _fmt_dur(mttr),
        'by_plateforme': sorted(by_plat.items(), key=lambda x: -x[1]['nb']),
        'by_escalade':   sorted(by_esc.items(),  key=lambda x: -x[1]['nb']),
    }


def stats_transport(rows):
    total = len(rows)
    avec = sum(1 for r in rows if r['avec_impact'])
    sans = total - avec
    closed = sum(1 for r in rows if r['is_closed'])
    total_dur = sum(r['duration_sec'] for r in rows if r['duration_sec'])
    mttr = total_dur / closed if closed else 0

    by_esc = defaultdict(lambda: {'nb': 0, 'avec': 0, 'dur': 0})
    for r in rows:
        k = r['escalade'] or 'Non défini'
        by_esc[k]['nb'] += 1
        if r['avec_impact']:
            by_esc[k]['avec'] += 1
        by_esc[k]['dur'] += r['duration_sec'] or 0

    by_region = defaultdict(lambda: {'nb': 0, 'dur': 0})
    for r in rows:
        k = r['region'] or 'Non défini'
        by_region[k]['nb'] += 1
        by_region[k]['dur'] += r['duration_sec'] or 0

    return {
        'total': total, 'avec_impact': avec, 'sans_impact': sans,
        'closed': closed, 'open': total - closed,
        'mttr_fmt': _fmt_dur(mttr),
        'by_escalade': sorted(by_esc.items(), key=lambda x: -x[1]['nb']),
        'by_region':   sorted(by_region.items(), key=lambda x: -x[1]['nb']),
        'top3': sorted([r for r in rows if r['avec_impact']],
                       key=lambda r: -(r['duration_sec'] or 0))[:3],
    }


def stats_igw(rows):
    total = len(rows)
    closed = sum(1 for r in rows if r['is_closed'])
    critiques = [r for r in rows if r['priorite'].upper() == 'CRITIQUE']

    by_lien = defaultdict(lambda: {'nb': 0, 'dur': 0})
    for r in rows:
        k = r['lien'] or r['lien_internet'] or 'Autre'
        by_lien[k]['nb'] += 1
        by_lien[k]['dur'] += r['duration_sec'] or 0

    top3 = sorted(rows, key=lambda r: -(r['duration_sec'] or 0))[:3]

    return {
        'total': total, 'closed': closed, 'open': total - closed,
        'critiques': len(critiques),
        'by_lien': sorted(by_lien.items(), key=lambda x: -x[1]['nb']),
        'top3': top3,
    }


def stats_core(rows):
    total = len(rows)
    closed = sum(1 for r in rows if r['is_closed'])

    by_esc = defaultdict(int)
    for r in rows:
        k = r['escalade'] or 'Non défini'
        by_esc[k] += 1

    by_espc = defaultdict(int)
    for r in rows:
        k = r['espc'] or 'Non défini'
        by_espc[k] += 1

    return {
        'total': total, 'closed': closed, 'open': total - closed,
        'by_escalade': sorted(by_esc.items(), key=lambda x: -x[1]),
        'by_espc':     sorted(by_espc.items(), key=lambda x: -x[1]),
        'rows': rows,
    }


# ── Réseau Mobile ────────────────────────────────────────────────────────────

def parse_mobile(ws):
    hr = _find_header_row(ws)
    rows = []
    for d in _sheet_rows(ws, hr):
        dur = _parse_duration(d.get('DURATION', d.get('DUR\xc9E', d.get('DUREE'))))
        if dur is None:
            # Try timedelta
            raw = d.get('DURATION', d.get('DUR\xc9E', d.get('DUREE')))
            if hasattr(raw, 'total_seconds'):
                dur = raw.total_seconds()
        alarm = _parse_dt(d.get('ALARM TIME'))
        cancel = _parse_dt(d.get('CANCEL TIME'))
        rows.append({
            'ticket':      str(d.get('NUMERO DU TICKET', d.get('NUMERO DU TICKET', '')) or '').strip(),
            'nature':      str(d.get("NATURE DE L'INCIDENT", d.get("NATURE DE L’INCIDENT", '')) or '')[:80],
            'alarm':       alarm,
            'site_name':   str(d.get('SITE NAME', '') or '').strip(),
            'site_id':     str(d.get('SITE ID', '') or '').strip(),
            'region':      str(d.get('R\xc9GION', d.get('REGION', '')) or '').strip(),
            'base':        str(d.get('BASE', '') or '').strip(),
            'plateforme':  str(d.get('PLATEFORME', '') or '').strip(),
            'technologies':str(d.get('TECHNOLOGIES', '') or '').strip(),
            'cause':       str(d.get('CAUSE', '') or '').strip()[:60],
            'root_cause':  str(d.get('ROOT CAUSE', '') or '').strip()[:60],
            'escalade':    str(d.get('ESCALADE', '') or '').strip(),
            'point_bloquant': str(d.get('POINT BLOQUANT', '') or '').strip(),
            'cancel':      cancel,
            'duration_sec': dur,
            'duration_fmt': _fmt_dur(dur),
            'status':      str(d.get('STATUS', '') or '').strip().upper(),
            'is_closed':   _is_closed(d.get('STATUS')),
        })
    return rows


def _is_dr2(alarm, cancel, duration_sec):
    """Calcul DR2 ARCEP : site HS >= 3h après la prochaine heure pleine suivant l'alarme."""
    if not alarm or not duration_sec:
        return False
    partial = alarm.minute * 60 + alarm.second
    secs_to_next = (3600 - partial) if partial > 0 else 3600
    return duration_sec >= secs_to_next + 10800


def stats_mobile(rows):
    total = len(rows)
    closed = sum(1 for r in rows if r['is_closed'])
    total_dur = sum(r['duration_sec'] for r in rows if r['duration_sec'])
    mttr = total_dur / closed if closed else 0

    # DR2 par incident
    dr2_rows = [r for r in rows if _is_dr2(r['alarm'], r['cancel'], r['duration_sec'])]
    nb_dr2 = len(dr2_rows)

    # DR1 : sites avec >= 2 incidents
    from collections import Counter
    site_counts = Counter(r['site_name'] for r in rows if r['site_name'])
    dr1_sites = {s: c for s, c in site_counts.items() if c >= 2}
    nb_dr1 = len(dr1_sites)

    # Top sites récurrents (DR1)
    top_sites_dr1 = []
    site_causes = defaultdict(list)
    site_regions = {}
    for r in rows:
        if r['site_name'] in dr1_sites:
            site_causes[r['site_name']].append(r['cause'] or r['root_cause'] or '')
            site_regions[r['site_name']] = r['region']
    for site, count in sorted(dr1_sites.items(), key=lambda x: -x[1])[:15]:
        causes = Counter(site_causes[site])
        top_cause = causes.most_common(1)[0][0] if causes else '—'
        top_sites_dr1.append({
            'site_name': site, 'count': count,
            'region': site_regions.get(site, '—'), 'cause': top_cause,
        })

    # Par escalade (Métier)
    esc_agg = defaultdict(lambda: {'nb': 0, 'dur': 0, 'dr2': 0})
    for r in rows:
        k = r['escalade'] or 'Non défini'
        esc_agg[k]['nb'] += 1
        esc_agg[k]['dur'] += r['duration_sec'] or 0
        if _is_dr2(r['alarm'], r['cancel'], r['duration_sec']):
            esc_agg[k]['dr2'] += 1
    by_escalade = []
    for name, agg in sorted(esc_agg.items(), key=lambda x: -x[1]['nb']):
        nb = agg['nb']
        by_escalade.append({
            'escalade': name, 'nb': nb,
            'mttr_fmt': _fmt_dur(agg['dur'] / nb if nb else 0),
            'nb_dr2': agg['dr2'],
            'eff_pct': round((1 - agg['dr2'] / nb) * 100) if nb else 100,
        })

    # Par région
    reg_agg = defaultdict(lambda: {'nb': 0, 'dur': 0, 'dr2': 0, 'sites': set()})
    for r in rows:
        k = r['region'] or 'Non défini'
        reg_agg[k]['nb'] += 1
        reg_agg[k]['dur'] += r['duration_sec'] or 0
        if r['site_name']:
            reg_agg[k]['sites'].add(r['site_name'])
        if _is_dr2(r['alarm'], r['cancel'], r['duration_sec']):
            reg_agg[k]['dr2'] += 1
    by_region = []
    for name, agg in sorted(reg_agg.items(), key=lambda x: -x[1]['nb']):
        nb = agg['nb']
        by_region.append({
            'region': name, 'nb': nb, 'nb_sites': len(agg['sites']),
            'mttr_fmt': _fmt_dur(agg['dur'] / nb if nb else 0),
            'nb_dr2': agg['dr2'],
            'eff_pct': round((1 - agg['dr2'] / nb) * 100) if nb else 100,
        })

    # Par base
    base_agg = defaultdict(lambda: {'nb': 0, 'dur': 0, 'dr2': 0, 'sites': set(), 'region': ''})
    for r in rows:
        k = r['base'] or r['region'] or 'Autre'
        base_agg[k]['nb'] += 1
        base_agg[k]['dur'] += r['duration_sec'] or 0
        if r['site_name']:
            base_agg[k]['sites'].add(r['site_name'])
        if not base_agg[k]['region']:
            base_agg[k]['region'] = r['region'] or ''
        if _is_dr2(r['alarm'], r['cancel'], r['duration_sec']):
            base_agg[k]['dr2'] += 1
    by_base = []
    for name, agg in sorted(base_agg.items(), key=lambda x: -x[1]['nb']):
        nb = agg['nb']
        by_base.append({
            'base': name, 'nb': nb, 'nb_sites': len(agg['sites']),
            'region': agg['region'],
            'mttr_fmt': _fmt_dur(agg['dur'] / nb if nb else 0),
            'nb_dr2': agg['dr2'],
            'eff_pct': round((1 - agg['dr2'] / nb) * 100) if nb else 100,
        })

    # Points bloquants
    pb_counts = Counter()
    for r in rows:
        pb = r['point_bloquant']
        if pb and pb.upper() not in ('N/A', '', 'NONE', 'NULL'):
            pb_counts[pb[:50]] += 1
    points_bloquants = [{'cause': c, 'count': n} for c, n in pb_counts.most_common(10)]

    # Causes principales
    cause_counts = Counter()
    for r in rows:
        c = r['root_cause'] or r['cause']
        if c and c.upper() not in ('N/A', ''):
            cause_counts[c[:50]] += 1
    by_cause = [{'cause': c, 'nb': n} for c, n in cause_counts.most_common(10)]

    return {
        'total': total, 'closed': closed, 'open': total - closed,
        'mttr_fmt': _fmt_dur(mttr),
        'total_dur': _fmt_dur(total_dur),
        'nb_dr2': nb_dr2, 'nb_dr1': nb_dr1,
        'top_sites_dr1': top_sites_dr1,
        'by_escalade': by_escalade,
        'by_region': by_region,
        'by_base': by_base,
        'points_bloquants': points_bloquants,
        'by_cause': by_cause,
        'dr2_rows': dr2_rows,
    }


# ── Point d'entrée principal ─────────────────────────────────────────────────

def parse_all(fileobj):
    """Parse toutes les feuilles du fichier BASES DES INCIDENTS.
    Retourne un dict { 'fixe': {...}, 'transport': {...}, 'igw': {...}, 'core': {...} }
    """
    wb = openpyxl.load_workbook(fileobj, data_only=True)
    sheets_map = {_norm(s).replace(' ', ''): s for s in wb.sheetnames}

    def get_ws(*keys):
        for key in keys:
            for k, name in sheets_map.items():
                if key in k:
                    return wb[name]
        return None

    result = {}

    ws = get_ws('RESEAUMOBILE', 'MOBILE')
    if ws:
        rows = parse_mobile(ws)
        result['mobile'] = {'rows': rows, 'stats': stats_mobile(rows)}

    ws = get_ws('RESEAUFIXE', 'FIXE')
    if ws:
        rows = parse_fixe(ws)
        result['fixe'] = {'rows': rows, 'stats': stats_fixe(rows)}

    ws = get_ws('TRANSPORT')
    if ws:
        rows = parse_transport(ws)
        result['transport'] = {'rows': rows, 'stats': stats_transport(rows)}

    ws = get_ws('IGW')
    if ws:
        rows = parse_igw(ws)
        result['igw'] = {'rows': rows, 'stats': stats_igw(rows)}

    ws = get_ws('CORE')
    if ws:
        rows = parse_core(ws)
        result['core'] = {'rows': rows, 'stats': stats_core(rows)}

    return result
