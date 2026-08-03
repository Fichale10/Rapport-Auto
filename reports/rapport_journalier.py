# -*- coding: utf-8 -*-
"""
Génération du fichier Excel « RAPPORT JOURNALIER  DD-MM-YYYY .xlsx » qui
rassemble tous les rapports de la plateforme (INCIDENTS MOB J-1, MTTR MOB,
STATISTIQUE J-1, FIXE J-1, MTTR FIXE, DR2 J-1, COMPILATION DR2,
COMPIL OUTAGE MOB) en reproduisant le design / les couleurs du modèle manuel.

NB : la feuille COMPILATION DR2 n'est pas encore automatisée — elle est
générée avec ses en-têtes uniquement.
"""
from __future__ import annotations

import calendar
import io
import logging
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ────────────────────────── Couleurs du modèle ──────────────────────────
FILL_HEADER   = PatternFill('solid', fgColor='4C5A77')   # en-tête bleu ardoise
FILL_GREEN    = PatternFill('solid', fgColor='72AB43')   # lignes résolues / escalade
FILL_YELLOW   = PatternFill('solid', fgColor='F0C715')   # Site Parent renseigné
FILL_RED      = PatternFill('solid', fgColor='FF0000')   # Inc count / DR2
FILL_DGREEN   = PatternFill('solid', fgColor='00B050')   # durées MTTR / labels outage
FILL_ORANGE   = PatternFill('solid', fgColor='FFC000')   # Non résolu / site parent DR2
FILL_DR2_HDR  = PatternFill('solid', fgColor='44546A')   # en-tête DR2
FILL_DR2_TITLE = PatternFill('solid', fgColor='D9D9D9')  # titre DR2
FILL_GREY     = PatternFill('solid', fgColor='BFBFBF')   # DR2 COUNT
FILL_YELLOW2  = PatternFill('solid', fgColor='FFD966')   # catégories DR2
FILL_GREEN2   = PatternFill('solid', fgColor='70AD47')   # % catégories DR2

FONT_HDR_11 = Font(bold=True, size=11, color='FFFFFF')
FONT_HDR_12 = Font(bold=True, size=12, color='FFFFFF')
FONT_HDR_14 = Font(bold=True, size=14, color='FFFFFF')
FMT_DUR  = '[h]:mm:ss;@'
FMT_DATE = 'mm-dd-yy'

# ─────────────────────── Colonnes / correspondances ──────────────────────
# En-têtes FR du modèle (25 colonnes) et équivalents EN des fichiers API.
FR_HEADERS = [
    'Ingénieur NOC', 'Numero du ticket', "Nature de l'incident", 'Alarm Time',
    'Site Parent', 'Site Name', 'Site ID', 'Région', 'Impact - Equipement',
    'Impact - Service', 'Plateforme', 'Technologies', 'Alarm text', 'Cause',
    'Escalade', 'Technicien Informé', 'Durée escalade', 'Action',
    'Technicien de maintenance', 'Root Cause', 'Observation', 'Point bloquant',
    'Cancel Time', 'Duration', 'Status',
]
EN_TO_FR = {
    'NOC Engineer': 'Ingénieur NOC',
    'Ticket Number': 'Numero du ticket',
    'Incident Nature': "Nature de l'incident",
    'Impact Equipement': 'Impact - Equipement',
    'Impact Service': 'Impact - Service',
    'Informed Technician': 'Technicien Informé',
    'Duration Escalade': 'Durée escalade',
    'Maintenance Technician': 'Technicien de maintenance',
    'Point Bloquant': 'Point bloquant',
}

ALARMES_MOBILES = [
    'BTS O&M LINK FAILURE / WCDMA BASE STATION OUT OF USE',
    'WCDMA BASE STATION OUT OF USE',
    'BTS O&M LINK FAILURE',
    'ALL RFMS MISSING',
]

# Ordre des escalades — STATISTIQUE J-1
ESCALADES_STAT = [
    'ENERGIE', 'RAN-FIELD O', 'TRANS FH-FIELD O', 'ENERGIE / TRANS / RAN',
    'TRANS / RAN', 'INFRA', 'PROJET', 'TRANS FO', 'TRANS FTTM', 'TRANS IP',
    'ENVIRONNEMENT', 'BSS',
]
# Libellés courts — COMPIL OUTAGE MOB
ESC_SHORT = {'RAN-FIELD O': 'RAN', 'TRANS FH-FIELD O': 'TRANS FH'}
# Catégories du bloc de synthèse DR2
DR2_CATEGORIES = [
    'RAN-FIELD O', 'ENERGIE', 'TRANS FH-FIELD O', 'TRANS FTTM', 'BSS',
    'INFRA', 'PROJET', 'TRANS FO', 'ENERGIE / TRANS / RAN', 'TRANS IP',
]

DR2_HEADERS = ['N°', 'Numero ticket', 'SITE PARENT', 'Site Name', 'Site ID',
               'Alarm Time', 'DUREE', 'Catégorie', 'CAUSE', 'POINT BLOQUANTS',
               'Cancel Time', 'OBSERVATION', 'DR2']
COMPIL_DR2_HEADERS = ['N°', 'DATE DR2 ', 'Numero ticket', 'SITE PARENT',
                      'Site Name', 'Site ID', 'Alarm Time', 'DUREE',
                      'Catégorie', 'CAUSE', 'POINT BLOQUANTS', 'Cancel Time',
                      'OBSERVATION', 'DR2']


# ─────────────────────────── Helpers données ───────────────────────────
def _parse_hms(val) -> timedelta | None:
    """'95:56:00' / '0:00:00' / '95:56' / secondes numériques → timedelta."""
    if val is None:
        return None
    if isinstance(val, timedelta):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return timedelta(seconds=int(val)) if val >= 0 else None
    try:
        parts = str(val).strip().split(':')
        if len(parts) == 3:
            h, m, s = int(float(parts[0])), int(parts[1]), int(float(parts[2]))
        elif len(parts) == 2:
            h, m, s = int(float(parts[0])), int(float(parts[1])), 0
        else:
            return None
        return timedelta(hours=h, minutes=m, seconds=s)
    except (ValueError, TypeError):
        return None


def _synth_is_valid(synth_list) -> bool:
    """Vrai si la synthèse contient une ligne TOTAL exploitable : dès que
    Inc count > 0, DUREE / OUTAGE doivent être lisibles et > 0 (les anciennes
    versions de l'import API stockaient une synthèse sans durées)."""
    if not synth_list:
        return False
    total = next((r for r in synth_list
                  if str(r.get('Escalade', '')).strip() == 'TOTAL'), None)
    if total is None:
        return False
    try:
        count = int(total.get('Inc count') or 0)
    except (TypeError, ValueError):
        count = 0
    if count == 0:
        return False         # suspect (jour sans aucun incident — à recalculer)
    duree  = _parse_hms(total.get('DUREE'))
    outage = _parse_hms(total.get('OUTAGE'))
    return bool(duree and duree.total_seconds() > 0
                and outage and outage.total_seconds() > 0)


def _mobile_reports_of_month(year: int, month: int):
    """Dernier rapport mobile mono-journée par jour → {date: UploadedReport}."""
    from django.db.models import Q
    from .models import UploadedReport

    qs = (UploadedReport.objects
          .filter(processed=True, date_rapport__year=year, date_rapport__month=month)
          .filter(Q(original_filename__startswith='API_MOBILE_') |
                  ~Q(original_filename__startswith='API_'))
          .order_by('date_rapport', 'uploaded_at'))
    result: dict[date, object] = {}
    for r in qs:
        if r.date_fin and r.date_fin != r.date_rapport:
            continue  # rapports multi-jours exclus
        result[r.date_rapport] = r  # le plus récent (ordre uploaded_at) l'emporte
    return result


def _parse_time_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns=EN_TO_FR)
    for col in ('Alarm Time', 'Cancel Time'):
        if col in df.columns:
            df[f'_{col}'] = pd.to_datetime(df[col], dayfirst=True,
                                           format='mixed', errors='coerce')
        else:
            df[f'_{col}'] = pd.NaT
    return df


def _day_mask(df: pd.DataFrame, day: date) -> pd.Series:
    """Tickets actifs pendant la journée `day`."""
    d0 = pd.Timestamp(f'{day} 00:00:00')
    d1 = pd.Timestamp(f'{day} 23:59:59')
    return (df['_Alarm Time'].notna() & (df['_Alarm Time'] <= d1)
            & (df['_Cancel Time'].isna() | (df['_Cancel Time'] >= d0)))


def _load_mobile_df(report) -> pd.DataFrame | None:
    """Charge les incidents mobiles d'un rapport : fichier brut si présent,
    sinon fichier détaillé (heures bornées à la journée)."""
    df = None
    if report.file:
        try:
            df = pd.read_excel(report.file.path)
        except Exception as exc:
            logger.warning('RJ: lecture brut impossible %s : %s',
                           report.original_filename, exc)
    if df is None and report.detailed_file:
        try:
            df = pd.read_excel(report.detailed_file.path)
        except Exception as exc:
            logger.warning('RJ: lecture détaillé impossible %s : %s',
                           report.original_filename, exc)
    if df is None:
        return None
    df = _parse_time_cols(df)
    if 'Alarm text' in df.columns:
        df = df[df['Alarm text'].astype(str).str.strip().isin(ALARMES_MOBILES)].copy()
    return df[_day_mask(df, report.date_rapport)].copy()


def _dr2_rows(df: pd.DataFrame, day: date) -> pd.DataFrame:
    """Incidents en violation DR2 pour le jour `day` (alarme ≥ J-1 00:00,
    règle : durée ≥ (heure pleine suivante − alarme) + 3 h)."""
    if df is None or df.empty:
        return pd.DataFrame()
    dd = df.copy()
    if 'Numero du ticket' in dd.columns:
        dd = dd.drop_duplicates(subset=['Numero du ticket'], keep='first')
    start = pd.Timestamp(datetime.combine(day - timedelta(days=1), datetime.min.time()))
    end_of_day = pd.Timestamp(f'{day} 23:59:00')
    dd = dd[dd['_Alarm Time'] >= start]

    keep = []
    for _, row in dd.iterrows():
        alarm = row['_Alarm Time']
        end = row['_Cancel Time'] if pd.notna(row['_Cancel Time']) else end_of_day
        next_hour = alarm.ceil('h')
        if next_hour == alarm:
            next_hour = alarm + pd.Timedelta(hours=1)
        allowed = (next_hour - alarm) + pd.Timedelta(hours=3)
        keep.append((end - alarm) >= allowed)
    return dd[pd.Series(keep, index=dd.index)].sort_values('_Alarm Time')


def _fetch_api_df(network: str, date_debut: date, date_fin: date) -> pd.DataFrame | None:
    """Données brutes via l'API ticketing pour un réseau / une période."""
    try:
        from .api_import import fetch_api_excel
        buf, _ = fetch_api_excel(date_debut.isoformat(), date_fin.isoformat(), network)
        df = pd.read_excel(buf)
    except Exception as exc:
        logger.warning('RJ: données %s indisponibles via API : %s', network, exc)
        return None
    return _parse_time_cols(df)


def _fetch_api_raw(network: str, date_debut: date, date_fin: date) -> pd.DataFrame | None:
    """Données brutes (colonnes d'origine, non renommées) via l'API ticketing."""
    try:
        from .api_import import fetch_api_excel
        buf, _ = fetch_api_excel(date_debut.isoformat(), date_fin.isoformat(), network)
        return pd.read_excel(buf)
    except Exception as exc:
        logger.warning('RJ: données brutes %s indisponibles via API : %s', network, exc)
        return None


def _synthesis_from_df(df_raw: pd.DataFrame, d: date) -> list | None:
    """Recalcule la synthèse journalière (mêmes chiffres que le traitement
    manuel) depuis les données brutes — secours quand `synthesis_json`
    est absent en base (ex. serveur de production sans historique)."""
    try:
        from treatement import process_file
        _, _, df_synthese = process_file(df_raw.copy(), d.isoformat())
        return df_synthese.to_dict('records')
    except Exception as exc:
        logger.warning('RJ: synthèse de secours impossible pour %s : %s', d, exc)
        return None


def _cell_str(row, col) -> str:
    val = row.get(col, '')
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return ''
    return str(val)


def _fmt_dt(ts) -> str:
    return ts.strftime('%d-%m-%Y %H:%M:%S') if pd.notna(ts) else ''


# ───────────────────────── Feuilles incidents ─────────────────────────
def _sheet_incidents(wb: Workbook, title: str, df: pd.DataFrame | None):
    ws = wb.create_sheet(title)
    for j, h in enumerate(FR_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = FILL_HEADER, FONT_HDR_11
    ws.column_dimensions['D'].width = 18.0
    ws.column_dimensions['E'].width = 13.9
    ws.column_dimensions['F'].width = 22.2
    if df is None or df.empty:
        return ws
    i = 2
    for _, row in df.iterrows():
        resolved = pd.notna(row.get('_Cancel Time'))
        for j, h in enumerate(FR_HEADERS, start=1):
            if h == 'Alarm Time':
                val = _fmt_dt(row.get('_Alarm Time'))
            elif h == 'Cancel Time':
                val = _fmt_dt(row.get('_Cancel Time'))
            else:
                val = _cell_str(row, h)
            c = ws.cell(row=i, column=j, value=val)
            if h == 'Site Parent' and val.strip():
                c.fill = FILL_YELLOW
            elif resolved:
                c.fill = FILL_GREEN
        i += 1
    return ws


# ──────────────────────────── MTTR MOB / FIXE ────────────────────────────
def _style_mttr_row(ws, i, ncols):
    ws.cell(row=i, column=1).number_format = FMT_DATE
    ws.cell(row=i, column=2).fill = FILL_RED
    for j in range(3, min(ncols, 5) + 1):
        c = ws.cell(row=i, column=j)
        c.fill, c.number_format = FILL_DGREEN, FMT_DUR
    if ncols >= 6:
        ws.cell(row=i, column=6).fill = FILL_RED
    for j in range(1, ncols + 1):
        ws.cell(row=i, column=j).font = Font(size=14)


def _sheet_mttr(wb: Workbook, title: str, headers, day_rows: dict, year, month, last_day):
    """day_rows : {jour(int): tuple valeurs (sans la date)}."""
    ws = wb.create_sheet(title)
    n = len(headers)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = FILL_HEADER, FONT_HDR_14
    nb_days = calendar.monthrange(year, month)[1]
    for d in range(1, nb_days + 1):
        i = d + 1
        _style_mttr_row(ws, i, n)
        if d > last_day:
            continue
        ws.cell(row=i, column=1, value=datetime(year, month, d))
        vals = day_rows.get(d)
        if vals:
            for j, v in enumerate(vals, start=2):
                if v is not None:
                    ws.cell(row=i, column=j, value=v)
    widths = {1: 16.1, 2: 14.3, 3: 14.3, 4: 11.7, 5: 16.1, 6: 12.7}
    for j in range(1, n + 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(j, 12.0)
    return ws


# ───────────────────────────── Construction ─────────────────────────────
def build_rapport_journalier(day: date) -> bytes:
    """Construit le classeur complet pour la journée `day` (J-1)."""
    year, month = day.year, day.month
    reports = _mobile_reports_of_month(year, month)

    # DataFrames bruts par jour (cache) — utilisés pour INCIDENTS + DR2
    df_cache: dict[date, pd.DataFrame | None] = {}
    api_month: list = []          # cache paresseux du mois mobile via API (parsé)
    api_raw: list = []            # cache paresseux du mois mobile via API (brut)
    synth_cache: dict[date, list | None] = {}

    def _mobile_api_raw() -> pd.DataFrame | None:
        if not api_raw:
            api_raw.append(_fetch_api_raw('mobile', date(year, month, 1), day))
        return api_raw[0]

    def _mobile_api_month() -> pd.DataFrame | None:
        if not api_month:
            raw = _mobile_api_raw()
            df = _parse_time_cols(raw) if raw is not None else None
            if df is not None and 'Alarm text' in df.columns:
                df = df[df['Alarm text'].astype(str).str.strip()
                        .isin(ALARMES_MOBILES)].copy()
            api_month.append(df)
        return api_month[0]

    def _synth_of(d: date) -> list | None:
        """Synthèse journalière : recalculée en priorité depuis les données
        LIVE de l'API ticketing (un ticket continue d'évoluer après son
        import initial — incidents tardifs, clôtures, causes mises à jour —
        donc une synthesis_json/un fichier stockés en base par l'import
        automatique peuvent être obsolètes). On ne retombe sur le fichier
        stocké du rapport puis sur la synthèse en base que si l'API est
        injoignable, pour ne jamais produire silencieusement un rapport
        incomplet."""
        if d not in synth_cache:
            rep = reports.get(d)
            db_synth = rep.synthesis_json if (rep and rep.synthesis_json) else None
            synth = None
            live = _mobile_api_raw()
            if live is not None:
                synth = _synthesis_from_df(live, d)
            if not synth and rep is not None and rep.file:
                try:
                    src = pd.read_excel(rep.file.path)
                    synth = _synthesis_from_df(src, d)
                except Exception:
                    pass
            if not synth and db_synth and _synth_is_valid(db_synth):
                logger.warning('RJ: API indisponible pour %s, utilisation de '
                               'la synthèse en base (potentiellement obsolète)', d)
                synth = db_synth
            if not synth:
                synth = db_synth   # au pire : la synthèse en base, même incomplète
            synth_cache[d] = synth
        return synth_cache[d]

    def _df_of(d: date):
        if d not in df_cache:
            dfm = _mobile_api_month()
            df = dfm[_day_mask(dfm, d)].copy() if dfm is not None else None
            if df is None:
                # API injoignable → dernier recours : fichier stocké du rapport
                r = reports.get(d)
                df = _load_mobile_df(r) if r else None
            df_cache[d] = df
        return df_cache[d]

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. INCIDENTS MOB J-1 ──
    df_j = _df_of(day)
    _sheet_incidents(wb, 'INCIDENTS MOB J-1', df_j)

    # ── 2. MTTR MOB ──
    mttr_rows: dict[int, tuple] = {}
    for d in range(1, day.day + 1):
        dt_d = date(year, month, d)
        synth_list = _synth_of(dt_d)
        if not synth_list:
            continue
        total = next((r for r in synth_list
                      if str(r.get('Escalade', '')).strip() == 'TOTAL'), None)
        if not total:
            continue
        try:
            count = int(total.get('Inc count') or 0)
        except (TypeError, ValueError):
            count = 0
        duree  = _parse_hms(total.get('DUREE'))
        outage = _parse_hms(total.get('OUTAGE'))
        mttr = (timedelta(seconds=int((duree / count).total_seconds()))
                if (duree and count) else timedelta(0))
        # DR2 non traité actuellement → colonne laissée vide
        dr2_count = None
        mttr_rows[d] = (count, duree, mttr, outage, dr2_count)
    _sheet_mttr(wb, ' MTTR MOB',
                ['DATE  ', 'Inc count ', 'DUREE', 'MTTR', 'OUTAGE', 'DR2'],
                mttr_rows, year, month, day.day)

    # ── 3. STATISTIQUE J-1 ──
    ws = wb.create_sheet('STATISTIQUE J-1')
    ws.merge_cells('A1:F1')
    t = ws.cell(row=1, column=1, value=f"RAPPORT D'INCIDENT DU {day.strftime('%d-%m-%Y')}")
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(horizontal='center')
    for j, h in enumerate(['Escalade', 'Inc count ', 'DUREE', 'MTTR', 'OUTAGE', 'Status '], 1):
        c = ws.cell(row=2, column=j, value=h)
        c.fill, c.font = FILL_HEADER, FONT_HDR_11
    synth = {str(r.get('Escalade', '')).strip(): r
             for r in (_synth_of(day) or [])}
    i = 3
    for esc in ESCALADES_STAT:
        rec = synth.get(esc, {})
        a = ws.cell(row=i, column=1, value=esc); a.fill = FILL_GREEN
        try:
            cnt = int(rec.get('Inc count') or 0)
        except (TypeError, ValueError):
            cnt = 0
        b = ws.cell(row=i, column=2, value=cnt); b.fill = FILL_RED
        for j, key in ((3, 'DUREE'), (4, 'MTTR'), (5, 'OUTAGE')):
            c = ws.cell(row=i, column=j, value=_parse_hms(rec.get(key)) or timedelta(0))
            c.number_format = '[h]:mm:ss'
        status = str(rec.get('Status') or 'N/A').strip() or 'N/A'
        f = ws.cell(row=i, column=6, value=status)
        f.fill = FILL_ORANGE if 'non' in status.lower() else FILL_GREEN
        i += 1
    # TOTAL
    tot = synth.get('TOTAL', {})
    a = ws.cell(row=i, column=1, value='TOTAL'); a.font = Font(bold=True)
    try:
        tcnt = int(tot.get('Inc count') or 0)
    except (TypeError, ValueError):
        tcnt = 0
    b = ws.cell(row=i, column=2, value=tcnt); b.font = Font(bold=True)
    for j, key in ((3, 'DUREE'), (4, 'MTTR'), (5, 'OUTAGE')):
        c = ws.cell(row=i, column=j, value=_parse_hms(tot.get(key)) or timedelta(0))
        c.number_format, c.font = '[h]:mm:ss', Font(bold=True)
    for col, w in (('A', 35.9), ('B', 10.9), ('C', 11.0), ('D', 8.4), ('E', 10.3), ('F', 13.9)):
        ws.column_dimensions[col].width = w

    # ── 4/5. FIXE J-1 + MTTR FIXE (via API — non stocké en base) ──
    df_fixe_month = _fetch_api_df('fixe', date(year, month, 1), day)
    df_fixe_j = None
    fixe_rows: dict[int, tuple] = {}
    if df_fixe_month is not None and not df_fixe_month.empty:
        df_fixe_j = df_fixe_month[_day_mask(df_fixe_month, day)]
        # stats journalières (regroupées par date d'alarme)
        dd = df_fixe_month
        if 'Numero du ticket' in dd.columns:
            dd = dd.drop_duplicates(subset=['Numero du ticket'], keep='first')
        dd = dd[dd['_Alarm Time'].notna()]
        end_bound = pd.Timestamp(f'{day} 23:59:00')
        for d in range(1, day.day + 1):
            dt_d = date(year, month, d)
            sub = dd[dd['_Alarm Time'].dt.date == dt_d]
            if sub.empty:
                fixe_rows[d] = (0, timedelta(0), timedelta(0))
                continue
            total = timedelta(0)
            for _, row in sub.iterrows():
                end = row['_Cancel Time'] if pd.notna(row['_Cancel Time']) else end_bound
                dur = end - row['_Alarm Time']
                if dur.total_seconds() > 0:
                    total += dur.to_pytimedelta()
            n = len(sub)
            mttr = timedelta(seconds=int(total.total_seconds() / n)) if n else timedelta(0)
            fixe_rows[d] = (n, total, mttr)
    _sheet_incidents(wb, 'FIXE J-1', df_fixe_j)
    _sheet_mttr(wb, 'MTTR FIXE', ['DATE  ', 'Inc count ', 'DUREE', 'MTTR'],
                fixe_rows, year, month, day.day)

    # ── 6. DR2 J-1 ──
    ws = wb.create_sheet(' DR2 J-1')
    ws.merge_cells('A1:M1')
    t = ws.cell(row=1, column=1, value=f'CAS DE VIOLATION DR2  {day.strftime("%d-%m-%Y")} ')
    t.font, t.fill = Font(bold=True, size=18), FILL_DR2_TITLE
    t.alignment = Alignment(horizontal='center')
    for j, h in enumerate(DR2_HEADERS, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.fill, c.font = FILL_DR2_HDR, FONT_HDR_12
    # DR2 non traité actuellement → feuille générée sans données (en-têtes seuls)
    dr2 = pd.DataFrame()
    end_of_day = pd.Timestamp(f'{day} 23:59:00')
    i, cat_counts = 3, {}
    for num, (_, row) in enumerate(dr2.iterrows(), start=1):
        parent = _cell_str(row, 'Site Parent').strip()
        esc = _cell_str(row, 'Escalade').strip()
        cat_counts[esc] = cat_counts.get(esc, 0) + 1
        resolved = pd.notna(row['_Cancel Time'])
        duree = ((row['_Cancel Time'] - row['_Alarm Time']).to_pytimedelta()
                 if resolved else 'ENCOURS')
        values = [num, _cell_str(row, 'Numero du ticket'), parent,
                  _cell_str(row, 'Site Name'), _cell_str(row, 'Site ID'),
                  _fmt_dt(row['_Alarm Time']), duree, esc,
                  _cell_str(row, 'Cause'), _cell_str(row, 'Point bloquant'),
                  _fmt_dt(row['_Cancel Time']) if resolved else 'EN COURS',
                  _cell_str(row, 'Observation'), 'OUI']
        for j, v in enumerate(values, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = FILL_GREEN
            if j == 3 and parent:
                c.fill = FILL_ORANGE
            if j in (4, 7):
                c.font = Font(bold=True)
            if j == 7 and resolved:
                c.number_format = FMT_DUR
        i += 1
    # bloc de synthèse — DR2 non traité : compteur à zéro, pourcentages vides
    i += 1
    total_dr2 = len(dr2)
    b = ws.cell(row=i, column=2, value='DR2 COUNT'); b.fill = FILL_GREY
    c = ws.cell(row=i, column=3, value=total_dr2); c.font = Font(bold=True)
    for cat in DR2_CATEGORIES:
        i += 1
        b = ws.cell(row=i, column=2, value=cat); b.fill = FILL_YELLOW2
        c = ws.cell(row=i, column=3, value=None)
        c.fill, c.font, c.number_format = FILL_GREEN2, Font(bold=True), '0%'
    for col, w in (('A', 3.4), ('B', 25.3), ('C', 14.7), ('D', 18.3), ('E', 9.1),
                   ('F', 17.9), ('G', 12.7), ('H', 19.7), ('I', 56.1), ('J', 41.0),
                   ('K', 18.1), ('L', 19.4), ('M', 5.1)):
        ws.column_dimensions[col].width = w

    # ── 7. COMPILATION DR2 (automatisation non prête → en-têtes seulement) ──
    ws = wb.create_sheet('COMPILATION DR2')
    for j, h in enumerate(COMPIL_DR2_HEADERS, 1):
        ws.cell(row=1, column=j, value=h).font = Font(bold=True)
    for col, w in (('A', 5.7), ('B', 13.4), ('C', 22.0), ('D', 19.0), ('E', 28.0),
                   ('F', 20.3), ('G', 25.9), ('H', 11.7), ('I', 46.0), ('J', 42.0),
                   ('K', 43.4), ('L', 18.0), ('M', 11.4), ('N', 4.7)):
        ws.column_dimensions[col].width = w

    # ── 8. COMPIL OUTAGE MOB ──
    ws = wb.create_sheet('COMPIL OUTAGE  MOB')
    nb_days = calendar.monthrange(year, month)[1]
    ws.merge_cells('A2:T2')
    t = ws.cell(row=2, column=1, value='OUTAGE JOURNALIER')
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(horizontal='center')
    ws.cell(row=3, column=1, value='Escalade').font = Font(bold=True)
    for d in range(1, nb_days + 1):
        c = ws.cell(row=3, column=d + 1, value=datetime(year, month, d))
        c.font, c.number_format = Font(bold=True), FMT_DATE
    # OUTAGE par escalade / jour depuis les synthesis_json quotidiens
    outages: dict[str, dict[int, timedelta]] = {e: {} for e in ESCALADES_STAT}
    for d in range(1, day.day + 1):
        synth_list = _synth_of(date(year, month, d))
        if not synth_list:
            continue
        for rec in synth_list:
            esc = str(rec.get('Escalade', '')).strip()
            if esc in outages:
                outages[esc][d] = _parse_hms(rec.get('OUTAGE')) or timedelta(0)
    i = 4
    for esc in ESCALADES_STAT:
        a = ws.cell(row=i, column=1, value=ESC_SHORT.get(esc, esc))
        a.fill = FILL_DGREEN
        for d in range(1, day.day + 1):
            if d in outages[esc]:
                c = ws.cell(row=i, column=d + 1, value=outages[esc][d])
                c.number_format = FMT_DUR
        i += 1
    a = ws.cell(row=i, column=1, value='TOTAL'); a.font = Font(bold=True, size=14)
    for d in range(1, day.day + 1):
        tot = timedelta(0)
        found = False
        for esc in ESCALADES_STAT:
            if d in outages[esc]:
                tot += outages[esc][d]
                found = True
        if found:
            c = ws.cell(row=i, column=d + 1, value=tot)
            c.font, c.number_format = Font(bold=True, size=14), FMT_DUR
    ws.column_dimensions['A'].width = 25.3
    for j in range(2, nb_days + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14.9

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
