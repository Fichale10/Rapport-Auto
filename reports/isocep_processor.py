import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.axis import DateAxis

class ExcelDataProcessor:
    def __init__(self, file_path_incident, file_path_avail, date_to_process, file_path_alarm = None):
        self.file_path_incident = file_path_incident
        self.file_path_avail = file_path_avail
        self.file_path_alarm = file_path_alarm
        self.first_date_string = date_to_process
        self.df_incident = None
        self.df_alarm = None
        self.df_avail_2G = None
        self.df_avail_3G = None
        self.df_avail_4G = None

    def load_data_from_excel(self):
        df_incident_allRowCol = pd.read_excel(self.file_path_incident, engine="openpyxl")
        df_avail_2G_allCol = pd.read_excel(self.file_path_avail, sheet_name="2G", engine="openpyxl")
        df_avail_3G_allCol = pd.read_excel(self.file_path_avail, sheet_name="3G", engine="openpyxl")
        df_avail_4G_allCol = pd.read_excel(self.file_path_avail, sheet_name="4G", engine="openpyxl")
        if self.file_path_alarm is not None:
            self.df_alarm = pd.read_excel(self.file_path_alarm, engine="openpyxl")
        
        self.df_incident = df_incident_allRowCol.loc[:, ["Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Escalade", "Cause", "Point bloquant", "Duration", "Status", "Site Name", "Impact - Equipement"]]
        self.df_avail_2G = df_avail_2G_allCol.loc[:, ["Period start time", "BCF name", "Cell avail accuracy 1s cellL"]]
        self.df_avail_3G = df_avail_3G_allCol.loc[:, ["Period start time", "WBTS name", "Cell Availability, excluding blocked by user state (BLU)"]]
        self.df_avail_4G = df_avail_4G_allCol.loc[:, ["Period start time", "MRBTS name", "Availability_4G"]]


    def merge_and_filter_data(self):
        df_avail_2G = self.df_avail_2G[self.df_avail_2G["Period start time"] == self.first_date_string]
        df_avail_3G = self.df_avail_3G[self.df_avail_3G["Period start time"] == self.first_date_string]
        df_avail_4G = self.df_avail_4G[self.df_avail_4G["Period start time"] == self.first_date_string]

        df_avail_2G.rename(columns={"BCF name": "Etiquettes de ligne", "Cell avail accuracy 1s cellL": "AVAILABILITY 2G"}, inplace=True)
        df_avail_3G.rename(columns={"WBTS name": "Etiquettes de ligne", "Cell Availability, excluding blocked by user state (BLU)": "AVAILABILITY 3G"}, inplace=True)
        df_avail_4G.rename(columns={"MRBTS name": "Etiquettes de ligne", "Availability_4G": "AVAILABILITY 4G"}, inplace=True)

        df_avail = df_avail_2G.merge(df_avail_3G, on="Etiquettes de ligne", how="outer")
        df_avail = df_avail.merge(df_avail_4G, on="Etiquettes de ligne", how="outer")

        # Liste pour stocker les lignes ajustées
        new_rows = []

        # Parcours des lignes du DataFrame
        for _, row in self.df_incident.iterrows():
            # Ajouter la ligne d'origine
            new_rows.append(row)

            # Si "EXT" est présent dans "Impact - Equipement"
            if "EXT" in str(row["Impact - Equipement"]):
                # Copier la ligne et ajuster "Site Name"
                duplicated_row = row.copy()
                duplicated_row["Site Name"] = f"{row['Site Name']}_EXT"
                
                # Ajouter la ligne dupliquée
                new_rows.append(duplicated_row)

        df_new_rows = pd.DataFrame(new_rows)
        self.df_incident = pd.concat([self.df_incident, df_new_rows], ignore_index=True).drop_duplicates(ignore_index=True)

        df_result = df_avail.merge(self.df_incident, left_on="Etiquettes de ligne", right_on="Site Name", how="outer")
        df_result.drop(columns=["Site Name"], inplace=True)
        df_result = df_result.astype(object).fillna("N/A")

        df_result = df_result[[
            "Etiquettes de ligne", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G",
            "Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Escalade",
            "Cause", "Point bloquant", "Duration", "Status"
        ]]
        
        cols_to_convert = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        for col in cols_to_convert:
            df_result[col] = pd.to_numeric(
                df_result[col].apply(lambda v: str(v).replace(",", ".")), errors="coerce"
            )

        df_res = df_result[(df_result["AVAILABILITY 2G"] <= 99) | (df_result["AVAILABILITY 3G"] <= 99) | (df_result["AVAILABILITY 4G"] <= 99)]
        df_res = df_res.astype(object).fillna("N/A")

        return df_res
    
    def format_duration(seconds):
        if not isinstance(seconds, (int, float)) or seconds < 0:
            return "Invalid duration"

        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        remaining_seconds = int(seconds % 60)

        result = f"{hours} h {minutes} m {remaining_seconds} s"
        return result

    def merge_and_filter_data_triple(self):
        df_avail_2G = self.df_avail_2G[self.df_avail_2G["Period start time"] == self.first_date_string]
        df_avail_3G = self.df_avail_3G[self.df_avail_3G["Period start time"] == self.first_date_string]
        df_avail_4G = self.df_avail_4G[self.df_avail_4G["Period start time"] == self.first_date_string]
        df_alarm = self.df_alarm.groupby('Name').head(3)

        # Conversion des colonnes "Alarm Time" et "Cancel Time" en datetime
        df_alarm["Alarm Time"] = pd.to_datetime(df_alarm["Alarm Time"], errors='coerce')
        df_alarm["Cancel Time"] = pd.to_datetime(df_alarm["Cancel Time"], errors='coerce')

        # Calcul de la durée de l'alarme
        df_alarm["Alarm Duration"] = (df_alarm["Cancel Time"] - df_alarm["Alarm Time"]).astype('timedelta64[s]')  # En minutes
        # df_alarm["Alarm Duration"] = df_alarm["Alarm Duration"].apply(lambda x: self.format_duration(x) if isinstance(x, (int, float)) else "N/A")

        # Remplacer les valeurs NaN par "N/A" en cas de dates manquantes ou invalides
        df_alarm["Alarm Duration"] = df_alarm["Alarm Duration"].astype(object).fillna("N/A")

        df_avail_2G.rename(columns={"BCF name": "Etiquettes de ligne", "Cell avail accuracy 1s cellL": "AVAILABILITY 2G"}, inplace=True)
        df_avail_3G.rename(columns={"WBTS name": "Etiquettes de ligne", "Cell Availability, excluding blocked by user state (BLU)": "AVAILABILITY 3G"}, inplace=True)
        df_avail_4G.rename(columns={"MRBTS name": "Etiquettes de ligne", "Availability_4G": "AVAILABILITY 4G"}, inplace=True)
        df_alarm.rename(columns={"Name": "Etiquettes de ligne"}, inplace=True)
        self.df_incident.rename(columns={"Duration": "Incident Duration"}, inplace=True)

        df_avail = df_avail_2G.merge(df_avail_3G, on="Etiquettes de ligne", how="outer")
        df_avail = df_avail.merge(df_avail_4G, on="Etiquettes de ligne", how="outer")

        # Liste pour stocker les lignes ajustées
        new_rows = []

        # Parcours des lignes du DataFrame
        for _, row in self.df_incident.iterrows():
            # Ajouter la ligne d'origine
            new_rows.append(row)

            # Si "EXT" est présent dans "Impact - Equipement"
            if "EXT" in str(row["Impact - Equipement"]):
                # Copier la ligne et ajuster "Site Name"
                duplicated_row = row.copy()
                duplicated_row["Site Name"] = f"{row['Site Name']}_EXT"
                
                # Ajouter la ligne dupliquée
                new_rows.append(duplicated_row)

        df_new_rows = pd.DataFrame(new_rows)
        self.df_incident = pd.concat([self.df_incident, df_new_rows], ignore_index=True).drop_duplicates(ignore_index=True)

        df_result = df_avail.merge(self.df_incident, left_on="Etiquettes de ligne", right_on="Site Name", how="outer")
        df_result.drop(columns=["Site Name"], inplace=True)
        df_result = df_result.astype(object).fillna("N/A")

        df_result = df_result[[
            "Etiquettes de ligne", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G",
            "Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Escalade",
            "Cause", "Point bloquant", "Incident Duration", "Status"
        ]]

        df_result_final = df_result.merge(df_alarm, on="Etiquettes de ligne", how="outer")
        df_result_final = df_result_final.astype(object).fillna("N/A")
        
        cols_to_convert = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        for col in cols_to_convert:
            df_result_final[col] = pd.to_numeric(
                df_result_final[col].apply(lambda v: str(v).replace(",", ".")), errors="coerce"
            )
        
        df_res = df_result_final[(df_result_final["AVAILABILITY 2G"] <= 99) | (df_result_final["AVAILABILITY 3G"] <= 99) | (df_result_final["AVAILABILITY 4G"] <= 99)]
        df_res = df_res.astype(object).fillna("N/A")

        df_res = df_res[[
            "Etiquettes de ligne", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G",
            "Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Alarm Text", "Object Class", "Escalade",
            "Cause", "Point bloquant", "Alarm Time", "Cancel Time", "Alarm Duration", "Incident Duration", "Status"
        ]]

        return df_res
    
    @staticmethod
    def _duration_to_minutes(val):
        """Convertit une durée (timedelta, time, secondes ou HH:MM:SS) en minutes float."""
        import datetime
        if val is None or val == "N/A":
            return None
        if isinstance(val, float) and val != val:   # NaN
            return None
        if isinstance(val, datetime.timedelta):
            return val.total_seconds() / 60.0
        if isinstance(val, datetime.time):
            return val.hour * 60 + val.minute + val.second / 60.0
        if isinstance(val, (int, float)):
            # La fonction format_duration existante travaille en secondes
            return float(val) / 60.0
        if isinstance(val, str):
            val = val.strip()
            if not val or val == "N/A":
                return None
            parts = val.split(':')
            if len(parts) == 3:
                try:
                    return int(parts[0]) * 60 + int(parts[1]) + float(parts[2]) / 60.0
                except (ValueError, TypeError):
                    pass
            if len(parts) == 2:
                try:
                    return int(parts[0]) * 60 + int(parts[1])
                except (ValueError, TypeError):
                    pass
        return None

    def _write_duration_stats(self, ws, df_final, duration_col, data_end_row):
        """Écrit un tableau récapitulatif (sites durée > 10 min) sous le tableau principal."""
        THRESHOLD_MIN = 10
        thin = Border(
            left=Side(border_style="thin"),  right=Side(border_style="thin"),
            top=Side(border_style="thin"),   bottom=Side(border_style="thin"),
        )

        # ── Collecte et filtrage ────────────────────────────────────────────
        seen_tickets = set()
        stat_rows = []
        for _, row in df_final.iterrows():
            dur_min = self._duration_to_minutes(row.get(duration_col))
            if dur_min is None or dur_min <= THRESHOLD_MIN:
                continue
            ticket = str(row.get("Numero du ticket", "N/A"))
            if ticket != "N/A":
                if ticket in seen_tickets:
                    continue
                seen_tickets.add(ticket)
            stat_rows.append({
                "site":     str(row.get("Etiquettes de ligne", "N/A")),
                "minutes":  dur_min,
                "ticket":   ticket,
                "status":   str(row.get("Status", "N/A")),
                "escalade": str(row.get("Escalade", "N/A")),
            })

        if not stat_rows:
            return

        stat_rows.sort(key=lambda x: x["minutes"], reverse=True)
        r = data_end_row + 2   # 2 lignes vides de séparation

        # ── Titre ──────────────────────────────────────────────────────────
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        tc = ws.cell(row=r, column=3,
                     value=f"📊  SITES AVEC DURÉE > {THRESHOLD_MIN} MIN  —  "
                           f"{len(stat_rows)} incident{'s' if len(stat_rows) > 1 else ''}")
        tc.fill      = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        tc.font      = Font(bold=True, color="FFFFFF", size=11)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

        # ── En-têtes ────────────────────────────────────────────────────────
        for i, h in enumerate(["Site", "Durée", "Ticket", "Statut", "Escalade"]):
            c = ws.cell(row=r, column=3 + i, value=h)
            c.fill      = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            c.font      = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin
        ws.row_dimensions[r].height = 18
        r += 1

        # ── Lignes de données ───────────────────────────────────────────────
        for idx, s in enumerate(stat_rows):
            total_s = int(s["minutes"] * 60)
            h_p = total_s // 3600
            m_p = (total_s % 3600) // 60
            s_p = total_s % 60
            dur_str = (f"{h_p}h {m_p:02d}m {s_p:02d}s" if h_p > 0
                       else f"{m_p}m {s_p:02d}s")

            is_ferme = s["status"].upper() == "FERME"
            row_bg   = "E8F5E9" if is_ferme else ("FFF3E0" if idx % 2 == 0 else "FFFFFF")
            # Couleur durée selon sévérité
            dur_bg   = ("FFCDD2" if s["minutes"] >= 60
                        else "FFE0B2" if s["minutes"] >= 30
                        else "FFF9C4")

            values = [s["site"], dur_str, s["ticket"], s["status"], s["escalade"]]
            for i, val in enumerate(values):
                c = ws.cell(row=r, column=3 + i, value=val)
                c.border    = thin
                c.alignment = Alignment(
                    horizontal="left" if i == 0 else "center",
                    vertical="center",
                )
                if i == 1:
                    c.fill = PatternFill(start_color=dur_bg, end_color=dur_bg, fill_type="solid")
                    c.font = Font(bold=True)
                else:
                    c.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            r += 1

    def generate_excel_report(self, df_final, output_path):
        # Créer un nouveau fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "SITES AVAILABILITY < 99%"

        # Colonnes de regroupement et autres
        group_columns = ["Etiquettes de ligne", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        highlight_columns_availability = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        highlight_columns_other = ["Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Escalade", "Cause", "Point bloquant", "Duration", "Status"]

        # 1. Mettre le titre dans les colonnes C et D (ligne 3)
        ws.merge_cells("C1:D1")
        title_cell = ws.cell(row=1, column=3, value="SITES AVAILABILITY < 99% @ " + self.first_date_string)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4C5A77", end_color="4C5A77", fill_type="solid")

        #  2. Écrire les en-têtes et appliquer les styles (modifié pour commencer à la ligne 3, colonne C)
        row_num = 3  # Ligne 3 pour les données
        for col_num, col_name in enumerate(df_final.columns, 3):  # Commencer à la colonne 3 (C)
            cell = ws.cell(row=row_num, column=col_num, value=col_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")  # Texte en blanc en gras
            cell.fill = PatternFill(start_color="4C5A77", end_color="4C5A77", fill_type="solid")  # Bleu-gris

        row_num = 4

        # Variables pour fusionner les cellules
        prev_values = {col: None for col in group_columns}  # Valeurs précédentes pour chaque colonne de groupe
        start_row = 4  # Début de la première ligne de données

        # Itération sur les lignes du DataFrame
        for i, (_, row) in enumerate(df_final.iterrows(), 4):  # Commence à la ligne 4 (les données commencent après les en-têtes)
            is_ferme = row.get("Status") == "FERME"  # Vérifie si le statut est "FERME"
            
            for col_num, col_name in enumerate(df_final.columns, 3):
                value = row[col_name]

                # Fusionner les cellules des colonnes de groupe si la valeur est identique à la précédente
                if col_name in group_columns:
                    if prev_values[col_name] == value:
                        if start_row is None:
                            start_row = row_num - 1
                    else:
                        if start_row is not None and start_row < row_num - 1:
                            # Fusionner les colonnes de regroupement
                            for col_index in range(3, 7):
                                ws.merge_cells(start_row=start_row, start_column=col_index, end_row=row_num - 1, end_column=col_index)
                        ws.cell(row=row_num, column=col_num, value=value)
                        start_row = row_num
                else:
                    ws.cell(row=row_num, column=col_num, value=value)

                # Appliquer le fond vert uniquement si "Statut" est "FERME"
                if is_ferme and col_name in highlight_columns_other:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="72AB43", end_color="72AB43", fill_type="solid")  # Vert clair
                elif col_name in highlight_columns_availability:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Rouge clair

                prev_values[col_name] = value

            row_num += 1

        # Vérifier et fusionner les dernières lignes de chaque groupe
        if start_row is not None and start_row < row_num - 1:
            for col_index in range(3, 7):  # Fusionner les 4 colonnes de regroupement
                ws.merge_cells(start_row=start_row, start_column=col_index, end_row=row_num - 1, end_column=col_index)

        # Vérification des lignes à colorier en rouge (N/A pour toutes les colonnes spécifiques)
        for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=len(df_final.columns) + 2):
            is_na_row = True
            for cell in row:
                # Si la cellule est fusionnée, on l'ignore et on vérifie les autres cellules
                if isinstance(cell, MergedCell):
                    continue
                if cell.column_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # Colonnes spécifiques
                    if cell.value != "N/A":
                        is_na_row = False
                        break
            if is_na_row:
                for cell in row:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Fond rouge

        # Appliquer l'alignement centré uniquement aux premières cellules des plages fusionnées
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                # Vérifier si la cellule fait partie d'une plage fusionnée
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= cell.row <= merged_range.max_row) and (merged_range.min_col <= cell.column <= merged_range.max_col):
                        # Appliquer l'alignement centré à la première cellule de la plage fusionnée
                        if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                            cell.alignment = Alignment(vertical="center")
                        else:
                            # Appliquer l'alignement centré aux autres cellules fusionnées
                            cell.alignment = Alignment(vertical="center")

        # Appliquer les bordures du tableau (modifié pour ne pas être en gras)
        thin_border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

        for row in ws.iter_rows(min_row=2, max_row=row_num - 1, min_col=3, max_col=len(df_final.columns) + 2):
            for cell in row:
                cell.border = thin_border

        # Définir la largeur des colonnes (exemples)
        column_widths = {
            'C': 20,  # Largeur de la colonne C
            'D': 20,  # Largeur de la colonne D
            'E': 20,  # Largeur de la colonne E
            'F': 20,  # Largeur de la colonne F
            'G': 28,  # Largeur de la colonne G
            'H': 21,  # Largeur de la colonne H
            'I': 35,  # Largeur de la colonne I
            'J': 15,  # Largeur de la colonne J
            'K': 25,  # Largeur de la colonne K
            'L': 20,  # Largeur de la colonne L
            'M': 10,  # Largeur de la colonne M
            'N': 10   # Largeur de la colonne N
        }

        # Appliquer la largeur définie à chaque colonne
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(output_path)


    def generate_excel_report_triple(self, df_final, output_path):
        # Créer un nouveau fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "SITES AVAILABILITY < 99%"

        # Colonnes de regroupement et autres
        group_columns = ["Etiquettes de ligne", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        highlight_columns_availability = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        highlight_columns_other = ["Ingénieur NOC", "Numero du ticket", "Nature de l'incident", "Alarm Text", "Object Class",
                                   "Escalade", "Cause", "Point bloquant", "Alarm Time", "Cancel Time", "Incident Duration", "Status"]

        # 1. Mettre le titre dans les colonnes C et D (ligne 3)
        ws.merge_cells("C1:D1")
        title_cell = ws.cell(row=1, column=3, value="SITES AVAILABILITY < 99% @ " + self.first_date_string)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, color="FFFFFF", name="Raleway")
        title_cell.fill = PatternFill(start_color="4C5A77", end_color="4C5A77", fill_type="solid")

        row_num = 4

        # Variables pour fusionner les cellules
        prev_values = {col: None for col in group_columns}  # Valeurs précédentes pour chaque colonne de groupe
        start_row = 4  # Début de la première ligne de données

        default_font = Font(name="Raleway")

        # Itération sur les lignes du DataFrame
        for i, (_, row) in enumerate(df_final.iterrows(), 4):  # Commence à la ligne 4 (les données commencent après les en-têtes)
            is_ferme = row.get("Status") == "FERME"  # Vérifie si le statut est "FERME"
            
            for col_num, col_name in enumerate(df_final.columns, 3):
                value = row[col_name]

                # Fusionner les cellules des colonnes de groupe si la valeur est identique à la précédente
                if col_name in group_columns:
                    if prev_values[col_name] == value:
                        if start_row is None:
                            start_row = row_num - 1
                    else:
                        if start_row is not None and start_row < row_num - 1:
                            # Fusionner les colonnes de regroupement
                            for col_index in range(3, 7):
                                ws.merge_cells(start_row=start_row, start_column=col_index, end_row=row_num - 1, end_column=col_index)
                        ws.cell(row=row_num, column=col_num, value=value)
                        start_row = row_num
                else:
                    ws.cell(row=row_num, column=col_num, value=value)

                # Appliquer le fond vert uniquement si "Statut" est "FERME"
                if is_ferme and col_name in highlight_columns_other:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="72AB43", end_color="72AB43", fill_type="solid")  # Vert clair
                elif col_name in highlight_columns_availability:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Rouge clair

                prev_values[col_name] = value

            row_num += 1

        # Vérifier et fusionner les dernières lignes de chaque groupe
        if start_row is not None and start_row < row_num - 1:
            for col_index in range(3, 7):  # Fusionner les 4 colonnes de regroupement
                ws.merge_cells(start_row=start_row, start_column=col_index, end_row=row_num - 1, end_column=col_index)

        # Vérification des lignes à colorier en rouge (N/A pour toutes les colonnes spécifiques)
        for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=len(df_final.columns) + 2):
            is_na_row = True
            for cell in row:
                # Si la cellule est fusionnée, on l'ignore et on vérifie les autres cellules
                if isinstance(cell, MergedCell):
                    continue
                if cell.column_letter in ['G', 'H', 'I', 'L', 'M', 'N']:  # Colonnes spécifiques
                    if cell.value != "N/A":
                        is_na_row = False
                        break
            if is_na_row:
                for cell in row:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Fond rouge

        # Appliquer l'alignement centré uniquement aux premières cellules des plages fusionnées
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                # Vérifier si la cellule fait partie d'une plage fusionnée
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= cell.row <= merged_range.max_row) and (merged_range.min_col <= cell.column <= merged_range.max_col):
                        # Appliquer l'alignement centré à la première cellule de la plage fusionnée
                        if cell.row == merged_range.min_row and cell.column == merged_range.min_col:
                            cell.alignment = Alignment(vertical="center")
                        else:
                            # Appliquer l'alignement centré aux autres cellules fusionnées
                            cell.alignment = Alignment(vertical="center")

        # Appliquer les bordures du tableau (modifié pour ne pas être en gras)
        thin_border = Border(left=Side(border_style="thin"),
                            right=Side(border_style="thin"),
                            top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"))

        for row in ws.iter_rows(min_row=2, max_row=row_num - 1, min_col=3, max_col=len(df_final.columns) + 2):
            for cell in row:
                cell.border = thin_border
                cell.font = default_font

        # Écrire les en-têtes et appliquer les styles (modifié pour commencer à la ligne 3, colonne C)
        row_num = 3  # Ligne 3 pour les données
        for col_num, col_name in enumerate(df_final.columns, 3):  # Commencer à la colonne 3 (C)
            cell = ws.cell(row=row_num, column=col_num, value=col_name)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF", name="Raleway")  # Texte en blanc en gras
            cell.fill = PatternFill(start_color="4C5A77", end_color="4C5A77", fill_type="solid")  # Bleu-gris


        # Définir la largeur des colonnes (exemples)
        column_widths = {
            'C': 22,  # Largeur de la colonne C
            'D': 20,  # Largeur de la colonne D
            'E': 20,  # Largeur de la colonne E
            'F': 20,  # Largeur de la colonne F
            'G': 28,  # Largeur de la colonne G
            'H': 24,  # Largeur de la colonne H
            'I': 40,  # Largeur de la colonne I
            'J': 30,  # Largeur de la colonne J
            'K': 15,  # Largeur de la colonne K
            'L': 20,  # Largeur de la colonne L
            'M': 25,  # Largeur de la colonne M
            'N': 20,   # Largeur de la colonne N
            'O': 25,
            'P': 25,
            'Q': 20,
            'R': 20,
            'S': 10,
        }

        # Appliquer la largeur définie à chaque colonne
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(output_path)



class ExcelGraphProcessor:
    def __init__(self, file_path_avail, site_to_process):
        self.file_path_avail = file_path_avail
        self.site_to_process = site_to_process.upper() if site_to_process else None
        self.df_avail_2G = None
        self.df_avail_3G = None
        self.df_avail_4G = None
        self.df_final = None

    def load_data_from_excel(self):
        df_avail_2G_allCol = pd.read_excel(self.file_path_avail, sheet_name="2G", engine="openpyxl")
        df_avail_3G_allCol = pd.read_excel(self.file_path_avail, sheet_name="3G", engine="openpyxl")
        df_avail_4G_allCol = pd.read_excel(self.file_path_avail, sheet_name="4G", engine="openpyxl")
        
        self.df_avail_2G = df_avail_2G_allCol.loc[:, ["Period start time", "BCF name", "Cell avail accuracy 1s cellL"]]
        self.df_avail_3G = df_avail_3G_allCol.loc[:, ["Period start time", "WBTS name", "Cell Availability, excluding blocked by user state (BLU)"]]
        self.df_avail_4G = df_avail_4G_allCol.loc[:, ["Period start time", "MRBTS name", "Availability_4G"]]

    def merge_and_filter_data(self):
        if not self.site_to_process:
            print("Aucun site spécifié. Veuillez fournir un site valide.")
            self.df_final = pd.DataFrame()  # DataFrame vide
            return

        # Renommage des colonnes
        self.df_avail_2G.rename(columns={"BCF name": "Etiquettes de ligne", "Cell avail accuracy 1s cellL": "AVAILABILITY 2G"}, inplace=True)
        self.df_avail_3G.rename(columns={"WBTS name": "Etiquettes de ligne", "Cell Availability, excluding blocked by user state (BLU)": "AVAILABILITY 3G"}, inplace=True)
        self.df_avail_4G.rename(columns={"MRBTS name": "Etiquettes de ligne", "Availability_4G": "AVAILABILITY 4G"}, inplace=True)

        # Mettre en majuscule les étiquettes pour comparaison
        self.df_avail_2G["Etiquettes de ligne"] = self.df_avail_2G["Etiquettes de ligne"].apply(lambda v: str(v).upper() if pd.notna(v) else v)
        self.df_avail_3G["Etiquettes de ligne"] = self.df_avail_3G["Etiquettes de ligne"].apply(lambda v: str(v).upper() if pd.notna(v) else v)
        self.df_avail_4G["Etiquettes de ligne"] = self.df_avail_4G["Etiquettes de ligne"].apply(lambda v: str(v).upper() if pd.notna(v) else v)

        # Filtrage des données
        df_avail_2G = self.df_avail_2G[self.df_avail_2G["Etiquettes de ligne"] == self.site_to_process]
        df_avail_3G = self.df_avail_3G[self.df_avail_3G["Etiquettes de ligne"] == self.site_to_process]
        df_avail_4G = self.df_avail_4G[self.df_avail_4G["Etiquettes de ligne"] == self.site_to_process]

        if df_avail_2G.empty and df_avail_3G.empty and df_avail_4G.empty:
            print(f"Aucune donnée trouvée pour le site {self.site_to_process}.")
            self.df_final = pd.DataFrame()  # DataFrame vide
            return

        # Fusion des DataFrames
        df_result = pd.merge(df_avail_2G, df_avail_3G, on="Period start time", how="outer")
        df_result = pd.merge(df_result, df_avail_4G, on="Period start time", how="outer")

        # Conversion des valeurs en numérique
        cols_to_convert = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        for col in cols_to_convert:
            df_result[col] = pd.to_numeric(
                df_result[col].apply(lambda v: str(v).replace(",", ".")), errors='coerce'
            )

        # Remplacement des valeurs manquantes par "N/A"
        df_result = df_result.astype(object).fillna("N/A")

        df_result = df_result[[
            "Period start time", "AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"
        ]]

        df_result["Period start time"] = pd.to_datetime(df_result["Period start time"], errors="coerce")

        self.df_final = df_result

    def generate_excel_report(self, output_path):
        if self.df_final is None or self.df_final.empty:
            print("Aucune donnée disponible pour générer un rapport.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "GRAPHES AVAILABILITY"

        # Titre du rapport
        ws.merge_cells("C1:G1")
        title_cell = ws.cell(row=1, column=3, value="GRAPHES AVAILABILITY @ " + self.site_to_process)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4C5A77", end_color="4C5A77", fill_type="solid")

        # Écriture des données
        for i, col in enumerate(self.df_final.columns, start=1):
            ws.cell(row=3, column=i, value=col)

        for row_idx, row in enumerate(self.df_final.itertuples(index=False), start=4):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Définition de la plage de données pour le tableau
        min_row = 3
        max_row = len(self.df_final) + 3
        min_col = 1
        max_col = len(self.df_final.columns)

        # Création du tableau Excel
        table = Table(displayName="AvailabilityTable", ref=f"A{min_row}:D{max_row}")

        # Style du tableau
        style = TableStyleInfo(
            name="TableStyleMedium9",  # Style prédéfini d'Excel
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        column_widths = {
            'A': 20,  # Largeur de la colonne A
            'B': 18,  # Largeur de la colonne B
            'C': 18,  # Largeur de la colonne C
            'D': 18,  # Largeur de la colonne D
        }

        # Appliquer la largeur définie à chaque colonne
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


        chart = LineChart()
        
        chart.style = 2

        # Référence dynamique en fonction du tableau
        data = Reference(ws, min_col=2, min_row=min_row, max_col=4, max_row=max_row)
        categories = Reference(ws, min_col=1, min_row=min_row + 1, max_row=max_row)  # En-tête exclue pour les catégories
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.title = f"KPI - {self.site_to_process}"
        chart.y_axis.title = "Disponibilité (%)"
        chart.x_axis.title = "Période de début"
        chart.legend.position = 'b'

        # Position des graphes sous les données
        ws.add_chart(chart, f"F4")

        # # Écriture des données
        # for i, col in enumerate(self.df_final.columns, start=1):
        #     ws.cell(row=3, column=i, value=col)

        # for row_idx, row in enumerate(self.df_final.itertuples(index=False), start=4):
        #     for col_idx, value in enumerate(row, start=1):
        #         ws.cell(row=row_idx, column=col_idx, value=value)

        # # Création des graphes
        # availabilities = ["AVAILABILITY 2G", "AVAILABILITY 3G", "AVAILABILITY 4G"]
        # for idx, avail in enumerate(availabilities, start=1):
        #     chart = LineChart()
        #     chart.title = f"{avail} - {self.site_to_process}"
        #     chart.style = 2
        #     chart.y_axis.title = "Disponibilité (%)"
        #     chart.x_axis.title = "Période de début"

        #     min_row = 4
        #     max_row = len(self.df_final) + 3

        #     # Utilisation des valeurs de "Period start time" comme abscisses (catégories)
        #     categories = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)

        #     # Sélection des données d'AVAILABILITY correspondantes pour les ordonnées (sans la légende par défaut)
        #     data = Reference(ws, min_col=idx + 1, min_row=min_row - 1, max_row=max_row)  # min_row - 1 pour inclure l'en-tête
            
        #     # Ajout des données avec un nom explicite pour la légende
        #     chart.series = [Series(data, title=avail)]
        #     chart.set_categories(categories)

        #     # Positionner les graphes sous les données
        #     ws.add_chart(chart, f"F{10 * idx}")

        wb.save(output_path)